#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K11 平台冒烟 · 扩展项（文档《K11_平台冒烟测试用例》约 48–58 行）

与 ``test_k11_p1_skill_herontest_playwright.py`` 相同：**不经过 L3**，本机 Playwright
``connect_over_cdp`` 附加已由 ``launch_chrome_debug.ps1`` 启动的 Chrome。

覆盖（自动化能力范围内）：
  P1 · 列表完整性、图片资源、静态资源异常（Console MIME/模块脚本等）、响应时间（分类切换）、无数据提示
  P2 · 轻量文案（替换字符）、轻量横向溢出、滚动后高度/稳定性（文档 56–57 浏览器兼容 / 弱网已不纳入本脚本）

前置与用法同 P1 脚本（仓库根 ``.env`` 中 ``KALAROKO_CDP_ENDPOINT`` 等）。

  python scripts/test_k11_platform_smoke_extended_playwright.py
  python scripts/test_k11_platform_smoke_extended_playwright.py -v --json-out out/k11_ext.json
        # -v：每条用例前后打印 page.url；底栏用例额外输出 viewport、各 frame 内 Home/Party 计数与每步策略日志
  python scripts/test_k11_platform_smoke_extended_playwright.py --xlsx-report "C:/Users/Me/Downloads/K11平台测试用例.xlsx"
  python scripts/test_k11_platform_smoke_extended_playwright.py --no-xlsx-report

运行结束后默认将各条「结果 / 备注」写入本地 ``K11平台测试用例.xlsx``（与文档 ``docs/K11平台测试用例_冒烟测试用例.csv`` 列名对齐）：
  - 路径优先 ``--xlsx-report``，否则环境变量 ``K11_XLSX_REPORT``，否则 ``~/Downloads/K11平台测试用例.xlsx``。
  - 需 ``pip install openpyxl``（见 ``core/requirements.txt``）；写入前请关闭 Excel 以免占用文件。

退出码：0 无 FAIL；1 存在 FAIL；2 CDP/环境；3 未捕获异常。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import time
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env", encoding="utf-8")
except ImportError:
    pass
except OSError:
    pass

DEFAULT_TARGET = "https://www.herontest.xin/"

# 底栏常见不在语义 <nav> 内，而在自定义 class 容器
BOTTOM_SHELL_SELECTORS = (
    "nav",
    "footer",
    "[class*='tabbar' i]",
    "[class*='TabBar']",
    "[class*='tab-bar' i]",
    "[class*='bottomNav' i]",
    "[class*='BottomNav' i]",
    "[class*='footer' i]",
)


def _kalaroko_cdp(cli: str | None) -> str:
    raw = (cli or "").strip() or (os.environ.get("KALAROKO_CDP_ENDPOINT") or "").strip()
    if not raw:
        raw = "http://127.0.0.1:9222"
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "http://" + raw.lstrip("/")
    return raw.rstrip("/")


def _host_from_url(url: str) -> str:
    try:
        from urllib.parse import urlparse

        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def _home_feed_url(target: str) -> str:
    from urllib.parse import urlparse, urlunparse

    t = (target or DEFAULT_TARGET).strip() or DEFAULT_TARGET
    p = urlparse(t)
    if p.scheme and p.netloc:
        return urlunparse((p.scheme, p.netloc, "/", "", "", ""))
    return t.rstrip("/") + "/" if t else DEFAULT_TARGET


def _needs_goto_home_feed(current_url: str) -> bool:
    """需整页回到站点根路径（/）再测 Home 列表底、底栏切换等。"""
    u = (current_url or "").lower()
    if "/my/" in u or "/me/" in u:
        return True
    if re.search(r"/(profile|account|wallet|settings)(/|$)", u):
        return True
    # Party Hubs 等子路由无「No More Data」大厅列表底
    if "party-hubs" in u or "/party/hub" in u:
        return True
    return False


# KalaroKo 底栏：文案在 div._item_label_* 内，可点击节点在父级，Playwright click 常超时
_JS_CLICK_TAB_FROM_LABEL = """(el) => {
  /* 文案常在 _item_label_*，其 class 也含子串 _item_；必须从父级起找整块 tab */
  let n = el.parentElement;
  for (let i = 0; i < 12 && n; i++) {
    const tag = (n.tagName || '').toUpperCase();
    const cls = (typeof n.className === 'string') ? n.className : '';
    const role = n.getAttribute && n.getAttribute('role');
    const isLabelOnly = /_item_label_/i.test(cls);
    if (!isLabelOnly && (tag === 'BUTTON' || tag === 'A' || role === 'button' || role === 'tab'
        || /\\bitem\\b/i.test(cls) || (cls.includes('_item_') && !cls.includes('_item_label_')))) {
      n.click();
      return { ok: true, step: i, tag, cls: cls.slice(0, 72) };
    }
    n = n.parentElement;
  }
  if (el.parentElement) {
    el.parentElement.click();
    return { ok: true, fallback: 'parent', tag: el.parentElement.tagName };
  }
  el.click();
  return { ok: true, fallback: 'label-only', tag: el.tagName };
}"""


async def _eval_timeout(
    page: Any,
    expression: str,
    arg: Any = None,
    *,
    timeout: float = 15.0,
) -> Any:
    """避免页面主线程卡死导致 evaluate 永久挂起。"""
    if arg is None:
        return await asyncio.wait_for(page.evaluate(expression), timeout=timeout)
    return await asyncio.wait_for(page.evaluate(expression, arg), timeout=timeout)


def _exc_tail(e: BaseException, lim: int = 200) -> str:
    return f"{type(e).__name__}: {str(e).strip()[:lim]}"


async def _log_bottom_nav_context(page: Any, log: Callable[[str], None], *, tag: str) -> None:
    """打印 viewport、scroll 与各 frame 内 Home/Party 精确匹配数量，便于定位底栏在哪个 frame、是否被判定不可见。"""
    log(f"  [诊断·底栏·{tag}] page.url = {page.url}")
    try:
        vp = await _eval_timeout(
            page,
            "() => ({ w: window.innerWidth, h: window.innerHeight, sy: window.scrollY, "
            "sh: Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0) })",
            timeout=8.0,
        )
        log(
            f"  [诊断·底栏·{tag}] viewport inner={vp.get('w')}×{vp.get('h')} "
            f"scrollY={vp.get('sy')} scrollHeight≈{vp.get('sh')}"
        )
    except Exception as e:
        log(f"  [诊断·底栏·{tag}] 读 viewport 失败：{_exc_tail(e, 120)}")
    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:140]
        except Exception:
            pass
        try:
            nh = await fr.get_by_text("Home", exact=True).count()
            np = await fr.get_by_text("Party", exact=True).count()
            nsh = await fr.get_by_text("Share", exact=True).count()
            nt = await fr.get_by_text("Task", exact=True).count()
            npr = await fr.get_by_text("Profile", exact=True).count()
            log(
                f"  [诊断·底栏·{tag}] frame[{fi}] url≈{fu!r} "
                f"exact: Home={nh} Party={np} Share={nsh} Task={nt} Profile={npr}"
            )
        except Exception as e:
            log(f"  [诊断·底栏·{tag}] frame[{fi}] 计数失败：{_exc_tail(e, 120)}")


async def _click_attached_force(
    loc: Any, *, timeout_ms: int = 8000, force: bool = True
) -> None:
    """不 wait visible（底栏常被判定为 hidden 但仍可 force 点）。"""
    await loc.wait_for(state="attached", timeout=min(6000, timeout_ms))
    try:
        await loc.scroll_into_view_if_needed(timeout=4000)
    except Exception:
        pass
    await loc.click(timeout=timeout_ms, force=force)


async def _ensure_on_home_feed(
    page: Any, target_url: str, log: Callable[[str], None] | None
) -> None:
    home = _home_feed_url(target_url)
    try:
        cur = page.url or ""
    except Exception:
        cur = ""
    if not _needs_goto_home_feed(cur):
        if log:
            log(f"  [诊断] 已在非个人中心路径：{cur!r}")
        return
    if log:
        log(f"  [诊断] goto 大厅：{cur!r} → {home!r}")
    await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(600)


async def _scroll_to_bottom(page: Any) -> None:
    try:
        h = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0, "
            "document.scrollingElement?.scrollHeight||0)",
            timeout=12.0,
        )
        if h is not None:
            await _eval_timeout(
                page, "y => window.scrollTo(0, y)", max(0, int(h)), timeout=12.0
            )
        await page.wait_for_timeout(400)
    except (Exception, asyncio.TimeoutError):
        pass
    try:
        await asyncio.wait_for(page.keyboard.press("End"), timeout=5.0)
        await page.wait_for_timeout(350)
    except (Exception, asyncio.TimeoutError):
        pass


async def _scroll_tabbar_into_view(page: Any) -> None:
    """固定底栏常被遮住时，先滚到页面最底部。"""
    try:
        await _eval_timeout(
            page,
            "() => { const y = Math.max(document.body?.scrollHeight||0, "
            "document.documentElement.scrollHeight||0); window.scrollTo(0, y); }",
            timeout=10.0,
        )
        await page.wait_for_timeout(350)
    except (Exception, asyncio.TimeoutError):
        pass


VERDICT_ZH = {
    "PASS": "通过",
    "FAIL": "未通过",
    "SKIP": "跳过",
}


# 顺序：先首页滚底看「无数据」，再点 Party 测切换耗时（避免离开首页后找不到列表底文案）
CASE_DEFS: list[tuple[str, str, str]] = [
    ("ext_game_list", "P1", "列表完整性（主要游戏名可见）"),
    ("ext_images", "P1", "图片资源（img 裂图抽样）"),
    ("ext_no_more_data", "P1", "无数据提示（No More Data）"),
    ("ext_response_time", "P1", "响应时间（底栏 Party 切换耗时）"),
    ("ext_copy_light", "P2", "文案检查（轻量：Unicode 替换符）"),
    ("ext_layout_light", "P2", "样式检查（轻量：横向溢出提示）"),
    ("ext_scroll_light", "P2", "滚动加载（轻量：滚底后高度与稳定性）"),
    ("ext_static_console", "P1", "静态资源/模块加载（Console MIME/模块脚本抽样，置末以纳入全程）"),
]

CASE_TITLE_ZH = {k: v for k, _, v in CASE_DEFS}

# case_id → Excel「测试项目」列应包含的子串（与《K11_平台冒烟测试用例》/ xlsx 一致）
EXT_CASE_TO_XLSX_TEST_ITEM_KEY: dict[str, str] = {
    "ext_game_list": "列表完整性",
    "ext_images": "图片资源",
    "ext_no_more_data": "无数据提示",
    "ext_response_time": "响应时间",
    "ext_copy_light": "文案检查",
    "ext_layout_light": "样式检查",
    "ext_scroll_light": "滚动加载",
    "ext_static_console": "静态资源",
}

_XLSX_REMARK_MAX_LEN = 32000


def _default_k11_xlsx_report_path() -> Path:
    env = (os.environ.get("K11_XLSX_REPORT") or "").strip()
    if env:
        return Path(env)
    return Path.home() / "Downloads" / "K11平台测试用例.xlsx"


def _find_smoke_sheet_header(ws: Any) -> tuple[int, int, int, int] | None:
    """定位表头行与列号（均为 1-based）：测试项目、结果、备注。"""
    max_r = min(int(ws.max_row or 1), 45)
    max_c = min(int(ws.max_column or 1), 40)
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = str(v).strip() if v is not None else ""
            if s and s not in col_map:
                col_map[s] = c
        if "结果" not in col_map or "备注" not in col_map:
            continue
        item_col = None
        for k in ("测试项目", "测试项", "用例名称"):
            if k in col_map:
                item_col = col_map[k]
                break
        if item_col is None:
            continue
        return r, item_col, col_map["结果"], col_map["备注"]
    return None


def write_k11_extended_results_to_xlsx(
    xlsx_path: Path,
    results: list[dict[str, Any]],
    *,
    log: Callable[[str], None],
) -> None:
    """将扩展冒烟结果写入 xlsx 对应行的「结果」「备注」列。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [xlsx] 未安装 openpyxl，跳过写入（可：pip install openpyxl）")
        return

    if not xlsx_path.is_file():
        log(f"  [xlsx] 文件不存在，跳过：{xlsx_path}")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        log(f"  [xlsx] 打开工作簿失败：{e!s}")
        return

    # 优先名称含「冒烟」的工作表
    ordered_names: list[str] = [n for n in wb.sheetnames if "冒烟" in n]
    ordered_names.extend(n for n in wb.sheetnames if n not in ordered_names)

    verdict_cell = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP", "BLOCKED": "BLOCKED"}

    try:
        for sn in ordered_names:
            ws = wb[sn]
            parsed = _find_smoke_sheet_header(ws)
            if not parsed:
                continue
            hdr, c_item, c_res, c_note = parsed
            last_r = max(int(ws.max_row or hdr), hdr + 5, 80)
            sheet_wrote = 0
            for resrow in results:
                cid = str(resrow.get("case") or "")
                key = EXT_CASE_TO_XLSX_TEST_ITEM_KEY.get(cid)
                if not key:
                    continue
                v = str(resrow.get("verdict") or "")
                detail = str(resrow.get("detail") or "")
                cell_v = verdict_cell.get(v, v)
                remark = (detail or "")[:_XLSX_REMARK_MAX_LEN]
                matched = False
                for r in range(hdr + 1, last_r + 1):
                    raw = ws.cell(row=r, column=c_item).value
                    text = str(raw).strip() if raw is not None else ""
                    if not text:
                        continue
                    if key in text:
                        ws.cell(row=r, column=c_res, value=cell_v)
                        ws.cell(row=r, column=c_note, value=remark)
                        sheet_wrote += 1
                        matched = True
                        break
                if not matched:
                    log(f"  [xlsx] 未匹配行：case={cid} 关键字={key!r} sheet={sn!r}")
            if sheet_wrote:
                try:
                    wb.save(xlsx_path)
                except PermissionError:
                    log(
                        f"  [xlsx] 保存被拒绝（是否正用 Excel 打开该文件？）：{xlsx_path.resolve()}"
                    )
                    return
                log(
                    f"  [xlsx] 已在工作表「{sn}」更新 {sheet_wrote} 行 → {xlsx_path.resolve()}"
                )
                return

        log(
            "  [xlsx] 未找到含「测试项目/结果/备注」表头的工作表，或无任何行被更新；"
            "请对照 docs/K11平台测试用例_冒烟测试用例.csv 列名。"
        )
    except Exception as e:
        log(f"  [xlsx] 写入过程异常：{e!s}")


def _mime_console_failures(console_lines: list[str]) -> list[str]:
    keys = (
        "MIME type",
        "module script",
        "Failed to load module",
        "text/html",
        "Strict MIME",
    )
    return [x for x in console_lines if any(k in x for k in keys)]


async def _broken_images_report(page: Any) -> tuple[int, list[str]]:
    """主文档 + 各 frame 统计 complete 且 naturalWidth==0 的 img src（抽样）。"""
    js = """() => {
      const imgs = Array.from(document.querySelectorAll('img'));
      const bad = [];
      for (const im of imgs) {
        try {
          if (im.complete && im.naturalWidth === 0 && (im.src || im.currentSrc)) {
            const s = (im.currentSrc || im.src || '').slice(0, 160);
            if (s) bad.push(s);
          }
        } catch (e) {}
      }
      return { total: imgs.length, bad: bad.slice(0, 20) };
    }"""

    total = 0
    bad_all: list[str] = []
    for fr in page.frames:
        try:
            r = await asyncio.wait_for(fr.evaluate(js), timeout=12.0)
            total += int(r.get("total") or 0)
            for s in r.get("bad") or []:
                if s not in bad_all:
                    bad_all.append(s)
        except (Exception, asyncio.TimeoutError):
            continue
    return total, bad_all[:25]


async def _visible_any(page: Any, pat: str, *, timeout_ms: float = 3500) -> bool:
    try:
        loc = page.get_by_text(re.compile(pat, re.I)).first
        await loc.wait_for(state="visible", timeout=int(timeout_ms))
        return True
    except Exception:
        return False


async def _run_ext_game_list(page: Any) -> tuple[str, str]:
    try:
        await page.mouse.wheel(0, 900)
        await page.wait_for_timeout(350)
    except Exception:
        pass
    patterns = [
        r"Tongits\s*King",
        r"Royal\s*Pusoy",
        r"Texas\s*Holdem",
        r"Bingo",
        r"Party",
    ]
    found: list[str] = []
    missing: list[str] = []
    for pat in patterns:
        ok = await _visible_any(page, pat, timeout_ms=2800)
        if ok:
            found.append(pat)
        else:
            missing.append(pat)
    if len(found) >= 3:
        return ("PASS", f"至少命中 {len(found)}/{len(patterns)} 组关键词：{', '.join(found)}")
    return (
        "FAIL",
        f"可见游戏/模块文案不足（需≥3）。命中：{found or '无'}；未命中：{', '.join(missing)}。"
        "请先在大厅首页并滚到游戏区。",
    )


async def _run_ext_images(page: Any) -> tuple[str, str]:
    total, bad = await _broken_images_report(page)
    if total == 0:
        return ("SKIP", "页面上未统计到 img 节点（可能图为 background 或 canvas）。")
    if bad:
        return (
            "FAIL",
            f"发现 {len(bad)} 个疑似裂图（complete 且 naturalWidth=0），共扫描约 {total} 个 img。"
            f" 示例：{bad[0][:120]}…" if bad else "",
        )
    return ("PASS", f"抽样检查：{total} 个 img 未发现典型裂图（naturalWidth=0）。")


async def _run_ext_static_console(console_bucket: list[str]) -> tuple[str, str]:
    bad = _mime_console_failures(console_bucket)
    if bad:
        return (
            "FAIL",
            "Console 存在与 MIME/模块脚本相关的错误（可能静态资源被 HTML 替代或缓存异常）："
            + " | ".join(bad[:3]),
        )
    if not console_bucket:
        return ("PASS", "抽样 Console error 中未见典型 MIME/模块脚本类报错。")
    return ("PASS", f"有 {len(console_bucket)} 条 Console error 抽样，但无 MIME/模块脚本关键字命中。")


async def _try_click_bottom_exact_in_frames(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    force: bool = True,
) -> bool:
    """
    底栏点击多策略：0) JS 从 label div 向上找 item/button 再 click（KalaroKo）
    ① class 容器 Home+Share ② nav/footer ③ Playwright force click nth ④ role ⑤ a/button。
    """
    label_re = re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)

    def lg(m: str) -> None:
        if log:
            log("  [诊断·耗时] " + m)

    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:130]
        except Exception:
            pass
        try:
            if "about:blank" in (fr.url or "").lower() and fi > 0:
                continue
        except Exception:
            pass

        # 策略 0：文案节点非可点击层，向上找父级 item / button / [role=tab] 再 DOM click
        try:
            nlab = await fr.get_by_text(label, exact=True).count()
            if nlab >= 1:
                loc0 = fr.get_by_text(label, exact=True).last
                await loc0.wait_for(state="attached", timeout=4000)
                info = await loc0.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                lg(f"frame[{fi}] 策略0 JS 底栏「{label}」→ {info} | {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略0「{label}」：{_exc_tail(e)}")

        # 策略 A：常见 tabbar class + 同时含 Home 与 Share（主导航条）
        for shell_sel in BOTTOM_SHELL_SELECTORS:
            try:
                base = fr.locator(shell_sel)
                bc = await base.count()
                if bc == 0:
                    continue
                shell = base.filter(has=fr.get_by_text("Home", exact=True))
                if await shell.count() == 0:
                    lg(f"frame[{fi}] 策略A {shell_sel!r} 有{bc}个但无 Home 子树")
                    continue
                shell2 = shell.filter(has=fr.get_by_text("Share", exact=True))
                if await shell2.count() > 0:
                    bar = shell2.last
                else:
                    bar = shell.last
                loc = bar.get_by_text(label, exact=True).first
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略A 容器{shell_sel!r} 内 first「{label}」 ok | {fu!r}")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略A {shell_sel!r}「{label}」：{_exc_tail(e)}")

        # 策略 B：nav/footer + Profile/Task 锚定
        try:
            shell = fr.locator("nav, footer").filter(has=fr.get_by_text("Home", exact=True))
            cn = await shell.count()
            lg(f"frame[{fi}] 策略B nav/footer∩Home 容器数={cn}")
            if cn > 0:
                bar = shell.filter(has=fr.get_by_text("Profile", exact=True)).last
                if await bar.count() == 0:
                    bar = shell.filter(has=fr.get_by_text("Task", exact=True)).last
                if await bar.count() == 0:
                    bar = shell.last
                loc = bar.get_by_text(label, exact=True).first
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略B nav/footer 内「{label}」 | {fu!r}")
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略B：{_exc_tail(e)}")

        # 策略 C：精确文案，按 nth 全尝试（last/first/其余）
        try:
            n = await fr.get_by_text(label, exact=True).count()
        except Exception as e:
            lg(f"frame[{fi}] 策略C 统计「{label}」失败：{_exc_tail(e)}")
            continue
        if n < 1:
            lg(f"frame[{fi}] 策略C 无 exact「{label}」，跳过本 frame | {fu!r}")
            continue
        order: list[int] = []
        for i in (n - 1, 0):
            if i >= 0 and i not in order:
                order.append(i)
        for i in range(n):
            if i not in order:
                order.append(i)
        lg(f"frame[{fi}] 策略C exact「{label}」×{n}，尝试 nth 顺序={order[:6]}{'…' if len(order) > 6 else ''}")
        for idx in order:
            try:
                loc = fr.get_by_text(label, exact=True).nth(idx)
                await loc.wait_for(state="attached", timeout=4000)
                try:
                    info = await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                    lg(f"frame[{fi}] 策略C-JS nth({idx})「{label}」→ {info}")
                    await page.wait_for_timeout(350)
                    return True
                except Exception as e_js:
                    lg(f"frame[{fi}] 策略C-JS nth({idx}) 失败：{_exc_tail(e_js, 100)}")
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略C nth({idx})「{label}」成功")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略C nth({idx})：{_exc_tail(e)}")

        # 策略 D：link / button 无障碍名
        for role in ("link", "button"):
            try:
                loc_all = fr.get_by_role(role, name=label_re)
                rn = await loc_all.count()
                if rn == 0:
                    continue
                loc = loc_all.last
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略D role={role} 匹配×{rn} 取 last")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略D role={role}：{_exc_tail(e)}")

        # 策略 E：可点击标签 + 文案
        try:
            row = fr.locator("a,button,[role='button']").filter(has_text=label_re)
            rn = await row.count()
            lg(f"frame[{fi}] 策略E a/button/[role=button] 匹配={rn}")
            if rn > 0:
                await _click_attached_force(row.last, force=force)
                lg(f"frame[{fi}] 策略E .last 成功")
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略E：{_exc_tail(e)}")

    lg(f"全部 frame 均未点到「{label}」")
    return False


# SPA 底栏切换：勿用 networkidle（长连接/轮询会导致几乎永远不达 idle，耗时≈超时）
_WAIT_PARTY_SWITCH_JS = """(before) => {
  const href = location.href;
  if (/party-hubs|party\\/hub/i.test(href)) return true;
  if (typeof before === 'string' && href !== before) return true;
  for (const row of document.querySelectorAll('[class*="tabbar_item"], [class*="_app_tabbar_item"]')) {
    const t = (row.innerText || '').trim();
    if (!t.includes('Party')) continue;
    const c = (row.className && row.className.toString()) || '';
    if (/\\bactive\\b/i.test(c)) return true;
    /* Party 选中态有时只在子节点；勿用 [class*="active"]（会误匹配 inactive） */
    const subs = row.getElementsByTagName('*');
    for (let i = 0; i < subs.length; i++) {
      const cn = (subs[i].className && subs[i].className.toString()) || '';
      if (/\\bactive\\b/i.test(cn)) return true;
    }
  }
  return false;
}"""


async def _wait_party_switch_settled(
    page: Any,
    url_before_click: str,
    *,
    timeout_ms: float,
    log: Callable[[str], None],
) -> tuple[bool, str]:
    """
    等待 Home→Party 在可观测层面的「切换完成」：Party 路由 URL、或 URL 已变、或底栏 Party 带 active。
    与 networkidle 解耦，避免把「网络安静」误当成「Tab 响应」。
    """
    to = max(500.0, float(timeout_ms))
    try:
        # Playwright Python：arg 必须为关键字参数，不可作第二位置参数
        await page.wait_for_function(
            _WAIT_PARTY_SWITCH_JS,
            arg=url_before_click,
            timeout=to,
        )
        log(f"  [诊断·耗时] Party 切换就绪（URL 或底栏 active，非 networkidle），上限等待 {to:.0f} ms")
        return True, "settled"
    except Exception as e:
        log(f"  [诊断·耗时] {to:.0f} ms 内未观察到 Party 路由/active：{_exc_tail(e, 160)}")
        return False, "timeout"


async def _run_ext_response_time(
    page: Any,
    threshold_ms: float,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    """先回站点根路径（含离开 /party-hubs）→ 露底栏 → Home → Party，测可观测切换耗时（非 networkidle）。"""
    log("  [诊断·耗时] 确保离开 party-hubs /my 等非大厅路由…")
    await _ensure_on_home_feed(page, target_url, log)
    await page.wait_for_timeout(400)
    await _log_bottom_nav_context(page, log, tag="响应·起测前")
    await _scroll_tabbar_into_view(page)
    await _log_bottom_nav_context(page, log, tag="响应·滚底后")
    ok_home = await _try_click_bottom_exact_in_frames(page, "Home", log=log, force=True)
    if not ok_home:
        log("  [诊断·耗时] 未能点击 Home（可能已在大厅）；继续尝试 Party…")
    await page.wait_for_timeout(500)
    await _scroll_tabbar_into_view(page)

    url_before = page.url
    t0 = time.monotonic()
    clicked = await _try_click_bottom_exact_in_frames(page, "Party", log=log, force=True)
    if not clicked:
        return ("SKIP", "未能点击底栏「Party」，跳过耗时统计（请先在大厅页或检查底栏是否在 iframe）。")
    ok_wait, _why = await _wait_party_switch_settled(
        page, url_before, timeout_ms=threshold_ms, log=log
    )
    dt_ms = (time.monotonic() - t0) * 1000
    if not ok_wait:
        return (
            "FAIL",
            f"{threshold_ms:.0f} ms 内未完成 Party 切换（未命中 party-hubs 等 URL、URL 未变且底栏 Party 无 active）。",
        )
    if dt_ms <= threshold_ms:
        return (
            "PASS",
            f"点击 Party 后至路由/active 就绪约 {dt_ms:.0f} ms（阈值 {threshold_ms:.0f} ms，非 networkidle）。",
        )
    return (
        "FAIL",
        f"切换耗时 {dt_ms:.0f} ms 超过阈值 {threshold_ms:.0f} ms（弱网或主线程卡顿可调高 --switch-ms）",
    )


async def _run_ext_no_more_data(
    page: Any, target_url: str, log: Callable[[str], None]
) -> tuple[str, str]:
    """「No More Data」仅出现在 Home 大厅列表底部：先 goto/离开 party-hubs 等再滚底。"""
    log("  [诊断·NoMoreData] 回到 Home 大厅根路径（避免当前在 Party Hubs 等子页）…")
    await _ensure_on_home_feed(page, target_url, log)
    await page.wait_for_timeout(450)
    try:
        await _eval_timeout(page, "() => window.scrollTo(0, 0)", timeout=10.0)
        await page.wait_for_timeout(200)
    except (Exception, asyncio.TimeoutError):
        pass
    await _scroll_to_bottom(page)
    ok = await _visible_any(page, r"No\s*More\s*Data", timeout_ms=4000)
    if ok:
        return ("PASS", "列表底部可见「No More Data」类文案。")
    ok2 = await _visible_any(page, r"没有更多|沒有更多|no\s+more", timeout_ms=1500)
    if ok2:
        return ("PASS", "列表底部可见「没有更多」类中文/变体文案。")
    return (
        "FAIL",
        "滚底后未见「No More Data」或常见中文无更多提示（若列表未分页到底则属环境差异）。",
    )


async def _run_ext_copy_light(page: Any) -> tuple[str, str]:
    try:
        txt = await asyncio.wait_for(
            page.evaluate(
                "() => (document.body && document.body.innerText) ? document.body.innerText.slice(0, 80000) : ''"
            ),
            timeout=30.0,
        )
    except asyncio.TimeoutError:
        return ("FAIL", "读取 innerText 超时（30s），页面主线程可能过重或卡死。")
    except Exception as e:
        return ("SKIP", f"无法读取正文：{e}")
    if not txt:
        return ("SKIP", "正文为空。")
    if "\ufffd" in txt or "\uFFFD" in txt:
        return ("FAIL", "正文出现 Unicode 替换字符 U+FFFD（可能编码/乱码）。")
    return ("PASS", "正文前 80k 字符未见 U+FFFD 替换符（非完整错别字审计）。")


async def _run_ext_layout_light(page: Any) -> tuple[str, str]:
    try:
        r = await _eval_timeout(
            page,
            """() => {
              const de = document.documentElement;
              const b = document.body;
              const sw = Math.max(de.scrollWidth, b ? b.scrollWidth : 0);
              const cw = de.clientWidth;
              return { scrollWidth: sw, clientWidth: cw, ratio: cw ? sw / cw : 1 };
            }""",
            timeout=12.0,
        )
        ratio = float(r.get("ratio") or 1)
        if ratio > 1.35:
            return (
                "FAIL",
                f"主文档横向 scrollWidth/clientWidth 比 ≈ {ratio:.2f}，可能存在明显横向溢出（轻量启发式）。",
            )
        return ("PASS", f"横向比例 ≈ {ratio:.2f}（轻量，非视觉回归）。")
    except Exception as e:
        return ("SKIP", f"无法测量布局：{e}")


async def _run_ext_scroll_light(page: Any) -> tuple[str, str]:
    try:
        h0 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        for _ in range(4):
            await _eval_timeout(
                page,
                "() => { window.scrollBy(0, 1400); }",
                timeout=8.0,
            )
            await page.wait_for_timeout(350)
        h1 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        await _scroll_to_bottom(page)
        h2 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        if h2 < (h0 or 0) * 0.5:
            return ("FAIL", f"滚底后 scrollHeight 异常收缩（{h0} → {h2}），可能存在布局闪动。")
        return (
            "PASS",
            f"滚动后高度 {h0} → {h1} → {h2}（轻量：未检测剧烈收缩）。",
        )
    except asyncio.TimeoutError:
        return ("FAIL", "滚动/scrollHeight 检测超时，页面脚本可能无响应（已避免无限挂起）。")
    except Exception as e:
        return ("SKIP", f"滚动测试异常：{e}")


async def _run_case(
    case_id: str,
    page: Any,
    *,
    console_bucket: list[str],
    switch_ms: float,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    if case_id == "ext_game_list":
        return await _run_ext_game_list(page)
    if case_id == "ext_images":
        return await _run_ext_images(page)
    if case_id == "ext_static_console":
        return await _run_ext_static_console(console_bucket)
    if case_id == "ext_response_time":
        return await _run_ext_response_time(page, switch_ms, log, target_url)
    if case_id == "ext_no_more_data":
        return await _run_ext_no_more_data(page, target_url, log)
    if case_id == "ext_copy_light":
        return await _run_ext_copy_light(page)
    if case_id == "ext_layout_light":
        return await _run_ext_layout_light(page)
    if case_id == "ext_scroll_light":
        return await _run_ext_scroll_light(page)
    return ("BLOCKED", f"未知用例：{case_id}")


async def _async_main(args: argparse.Namespace) -> int:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("请先安装：pip install playwright && playwright install chromium", file=sys.stderr)
        return 2

    cdp = _kalaroko_cdp(args.cdp_http or None)
    target_url = (args.target_url or DEFAULT_TARGET).strip()
    host = _host_from_url(target_url)
    case_ids = [c[0] for c in CASE_DEFS]

    def log(msg: str) -> None:
        if args.verbose or not args.quiet:
            print(msg, flush=True)

    log("———————— K11 平台冒烟 · 扩展（文档 48–58 行）————————")
    log(f"CDP：{cdp}  目标：{target_url}")
    log(f"分类切换阈值：{args.switch_ms} ms")
    log("")

    console_bucket: list[str] = []

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(cdp)
        if not browser.contexts:
            print("[失败] CDP 已连上但无 context", file=sys.stderr)
            return 2
        ctx = browser.contexts[0]
        pages = list(ctx.pages)
        if not pages:
            print("[失败] 无打开标签页", file=sys.stderr)
            return 2

        picked = None
        for pg in reversed(pages):
            try:
                u = pg.url or ""
            except Exception:
                u = ""
            if host and host in u.lower():
                picked = pg
                break
        if picked is None:
            if args.navigate_if_no_tab:
                picked = pages[-1]
                log(f"[nav] goto {target_url}")
                await picked.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
            else:
                print(
                    f"[失败] 无含 {host!r} 的标签页。请打开站点或加 --navigate-if-no-tab",
                    file=sys.stderr,
                )
                return 2
        else:
            await picked.bring_to_front()

        page = picked

        def _on_console(msg: Any) -> None:
            try:
                if msg.type == "error":
                    console_bucket.append(f"{msg.type}: {msg.text[:600]}")
            except Exception:
                pass

        page.on("console", _on_console)

        log(f"当前 URL：{page.url}")
        log(f"标题：{await page.title()}")
        log("")
        log("准备：回到大厅首页（避免停在 /my/index）…")
        await _ensure_on_home_feed(page, target_url, log)
        log("轻微滚屏以便游戏区与列表露出…")
        try:
            await _eval_timeout(page, "() => window.scrollBy(0, 400)", timeout=6.0)
        except (Exception, asyncio.TimeoutError):
            pass
        await page.wait_for_timeout(400)

        results: list[dict[str, Any]] = []
        for i, cid in enumerate(case_ids, start=1):
            tier = next(t for k, t, _ in CASE_DEFS if k == cid)
            title = CASE_TITLE_ZH[cid]
            log(f"【{i}/{len(case_ids)}】[{tier}] {title}（{cid}）")
            if args.verbose:
                try:
                    _t = await page.title()
                    log(f"  [诊断·用例] 执行前 url={page.url!r} title={_t!r}")
                except Exception as e:
                    log(f"  [诊断·用例] 执行前读页面信息失败：{_exc_tail(e, 100)}")
            v, detail = await _run_case(
                cid,
                page,
                console_bucket=console_bucket,
                switch_ms=float(args.switch_ms),
                log=log,
                target_url=target_url,
            )
            if args.verbose:
                try:
                    log(f"  [诊断·用例] 执行后 url={page.url!r} → 本条结论 {v}")
                except Exception as e:
                    log(f"  [诊断·用例] 执行后读 url 失败：{_exc_tail(e, 100)}")
            log(f"  观察说明：{detail}")
            log(f"  结论：{VERDICT_ZH.get(v, v)}（{v}）")
            log("")
            results.append(
                {
                    "case": cid,
                    "tier": tier,
                    "title_zh": title,
                    "verdict": v,
                    "detail": detail,
                }
            )

        mime_hits = _mime_console_failures(console_bucket)
        if mime_hits:
            log("———————— Console · MIME/模块脚本相关（供对照）————————")
            for s in mime_hits[:8]:
                log("  · " + s[:300])
            log("")

        out = {
            "schema": "k11_platform_smoke_extended/v1",
            "ts_utc": datetime.now(timezone.utc).isoformat(),
            "cdp": cdp,
            "target_url": target_url,
            "switch_ms_threshold": args.switch_ms,
            "page_url_final": page.url,
            "page_title_final": await page.title(),
            "console_errors_sample": console_bucket[:30],
            "results": results,
        }
        if args.json_out:
            outp = Path(args.json_out)
            outp.parent.mkdir(parents=True, exist_ok=True)
            outp.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            log(f"JSON：{outp.resolve()}")

        if not args.no_xlsx_report:
            xlsx_p = args.xlsx_report if args.xlsx_report is not None else _default_k11_xlsx_report_path()
            log("")
            write_k11_extended_results_to_xlsx(Path(xlsx_p), results, log=log)

        log("———————— 汇总 ————————")
        for r in results:
            m = "✓" if r["verdict"] == "PASS" else ("○" if r["verdict"] == "SKIP" else "✗")
            log(f"  {m} [{r['tier']}] {r['title_zh']} → {VERDICT_ZH.get(r['verdict'], r['verdict'])}")

        bad = {r["verdict"] for r in results}
        if "FAIL" in bad or "BLOCKED" in bad:
            log("\n最终结果：存在未通过项，退出码 1。")
            return 1
        log("\n最终结果：无 FAIL，退出码 0。")
        return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="K11 文档 48–58 行扩展冒烟（Playwright CDP）")
    ap.add_argument("--target-url", default=DEFAULT_TARGET, help="站点 URL（匹配标签页 host）")
    ap.add_argument("--cdp-http", default="", help="覆盖 KALAROKO_CDP_ENDPOINT")
    ap.add_argument("--navigate-if-no-tab", action="store_true")
    ap.add_argument(
        "--switch-ms",
        type=float,
        default=12000.0,
        help="P1 响应时间：底栏 Party 点击后至 Party 路由或底栏 active 的上限毫秒（默认 12000；不用 networkidle）",
    )
    ap.add_argument("--json-out", type=Path, default=None)
    ap.add_argument(
        "--xlsx-report",
        type=Path,
        default=None,
        help="K11平台测试用例.xlsx 路径；默认 K11_XLSX_REPORT 或 ~/Downloads/K11平台测试用例.xlsx",
    )
    ap.add_argument(
        "--no-xlsx-report",
        action="store_true",
        help="不将结果写入 Excel",
    )
    ap.add_argument("--quiet", action="store_true")
    ap.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="每条用例前后打印 url/title；响应时间用例含底栏 viewport + frame 计数 + 多策略点击明细",
    )
    args = ap.parse_args()
    try:
        return asyncio.run(_async_main(args))
    except Exception as e:
        print(f"[失败] {type(e).__name__}: {e}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
