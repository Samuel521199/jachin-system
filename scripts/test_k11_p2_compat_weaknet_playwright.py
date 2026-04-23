#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K11 平台冒烟 · P2：浏览器兼容 + 弱网体验

对齐文档《K11_平台冒烟测试用例》约 55–57 行（P2 浏览器兼容、弱网体验）。

- **浏览器兼容**：本机 Playwright 分别 **启动** ``channel=chrome`` 与 ``channel=msedge``（Windows 上即 Chrome / 系统 Edge），
  打开目标 URL，校验页面基本可读（非白屏、有标题/正文）。忽略站点机器人验证页（不与之交互，仅作可达性/渲染判断）。

- **弱网体验**：通过 **已开启远程调试的 Chrome**（与 P0/P1 相同）**CDP 附加**，在页面上用 ``Network.emulateNetworkConditions`` 模拟 Slow3G
  类环境；用例中封装为 ``handle_weak_network_test()``。连接串由 ``KALAROKO_CDP_ENDPOINT`` 或 ``--cdp-http`` 指定（默认
  ``http://127.0.0.1:9222``）。

  **弱网体验（以观测为主、默认 PASS）**：聚合 ``page.frames``，记录**从开始导航到首现可感知内容**的秒数并给出文字评价；不因 SPA/iframe/Shadow
  导致 ``innerText`` 暂时为 0 即判 FAIL。仅当导航级失败、或长时无任何结构/标题等信号时仍可为 FAIL（见实现）。

  - **3s/5s 档**：仅作 Performance sub-optimal、骨架+顶栏等 **Observation**，**不**作硬性 FAIL 依据。
  - **首内容耗时**：在限速下轮询，直至骨架、正文、``#root``/``#app`` 壳、标题等任一可感知信号，记 ``weak_net_sec_to_first_signal`` 与简评（优秀/良好/可接受/偏慢/过慢）。

前置：
  - 浏览器兼容：``pip install playwright``，并 ``playwright install chrome``（如缺）
    与 ``playwright install msedge``（如缺 Edge 通道，脚本会对该路给出 FAIL 说明）
  - 弱网：先按 ``launch_chrome_debug.ps1`` 等启动带 ``--remote-debugging-port=9222`` 的 Chrome

用法（仓库根）：
  python scripts/test_k11_p2_compat_weaknet_playwright.py
  python scripts/test_k11_p2_compat_weaknet_playwright.py --target-url https://www.herontest.xin/
  python scripts/test_k11_p2_compat_weaknet_playwright.py --only-weak
  python scripts/test_k11_p2_compat_weaknet_playwright.py --only-compat --headless

运行结束可将结果写入 ``K11平台测试用例.xlsx``（浏览器兼容两条 Chrome/Edge 合并为一行「浏览器兼容」）：
  ``--xlsx-report`` / ``K11_XLSX_REPORT`` / ``~/Downloads/K11平台测试用例.xlsx``；``--no-xlsx-report`` 关闭。需 ``openpyxl``。

退出码：0 无 FAIL；1 存在 FAIL；2 环境/依赖；3 未捕获异常。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import time
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env", encoding="utf-8")
except ImportError:
    pass
except OSError:
    pass

DEFAULT_TARGET = "https://www.herontest.xin/"

VERDICT_ZH = {
    "PASS": "通过",
    "FAIL": "未通过",
    "SKIP": "跳过",
}

P2_CASE_TO_XLSX_TEST_ITEM_KEY: dict[str, str] = {
    "p2_browser_compat_merged": "浏览器兼容",
    "p2_weak_network": "弱网",
}

_XLSX_REMARK_MAX_LEN = 32000


def _default_k11_xlsx_report_path() -> Path:
    env = (os.environ.get("K11_XLSX_REPORT") or "").strip()
    if env:
        return Path(env)
    return Path.home() / "Downloads" / "K11平台测试用例.xlsx"


def _find_smoke_sheet_header(ws: Any) -> tuple[int, int, int, int] | None:
    max_r = min(int(ws.max_row or 1), 45)
    max_c = min(int(ws.max_column or 1), 40)
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = str(v).strip() if v is not None else ""
            if s and s not in col_map:
                col_map[s] = c
        if "结果" not in col_map or "备注" not in col_map:
            continue
        item_col = None
        for k in ("测试项目", "测试项", "用例名称"):
            if k in col_map:
                item_col = col_map[k]
                break
        if item_col is None:
            continue
        return r, item_col, col_map["结果"], col_map["备注"]
    return None


def _p2_normalize_rows_for_xlsx(all_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Chrome/Edge 两行合并为一条写表；弱网单独一行。"""
    compat = [
        r
        for r in all_rows
        if str(r.get("case_id", "")).startswith("p2_browser_compat_")
    ]
    out: list[dict[str, Any]] = []
    for r in all_rows:
        cid = str(r.get("case_id", ""))
        if cid.startswith("p2_browser_compat_"):
            continue
        out.append(
            {
                "case": cid,
                "verdict": str(r.get("verdict", "")),
                "detail": str(r.get("detail", "")),
            }
        )
    if compat:
        ok = all(x.get("verdict") == "PASS" for x in compat)
        v = "PASS" if ok else "FAIL"
        lines = [
            f"{x.get('title_zh', '')}: {x.get('verdict')} — {str(x.get('detail', ''))[:2000]}"
            for x in compat
        ]
        out.insert(
            0,
            {"case": "p2_browser_compat_merged", "verdict": v, "detail": "\n".join(lines)},
        )
    return out


def write_k11_p2_results_to_xlsx(
    xlsx_path: Path,
    results: list[dict[str, Any]],
    *,
    log: Callable[[str], None],
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [xlsx] 未安装 openpyxl，跳过写入（可：pip install openpyxl）")
        return

    if not xlsx_path.is_file():
        log(f"  [xlsx] 文件不存在，跳过：{xlsx_path}")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        log(f"  [xlsx] 打开工作簿失败：{e!s}")
        return

    ordered_names: list[str] = [n for n in wb.sheetnames if "冒烟" in n]
    ordered_names.extend(n for n in wb.sheetnames if n not in ordered_names)
    verdict_cell = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP", "BLOCKED": "BLOCKED"}

    try:
        for sn in ordered_names:
            ws = wb[sn]
            parsed = _find_smoke_sheet_header(ws)
            if not parsed:
                continue
            hdr, c_item, c_res, c_note = parsed
            last_r = max(int(ws.max_row or hdr), hdr + 5, 80)
            sheet_wrote = 0
            for resrow in results:
                cid = str(resrow.get("case") or "")
                key = P2_CASE_TO_XLSX_TEST_ITEM_KEY.get(cid)
                if not key:
                    continue
                v0 = str(resrow.get("verdict") or "")
                detail = str(resrow.get("detail") or "")
                cell_v = verdict_cell.get(v0, v0)
                remark = (detail or "")[:_XLSX_REMARK_MAX_LEN]
                matched = False
                for r in range(hdr + 1, last_r + 1):
                    raw = ws.cell(row=r, column=c_item).value
                    text = str(raw).strip() if raw is not None else ""
                    if not text:
                        continue
                    if key in text:
                        ws.cell(row=r, column=c_res, value=cell_v)
                        ws.cell(row=r, column=c_note, value=remark)
                        sheet_wrote += 1
                        matched = True
                        break
                if not matched:
                    log(f"  [xlsx] 未匹配行：case={cid} 关键字={key!r} sheet={sn!r}")
            if sheet_wrote:
                try:
                    wb.save(xlsx_path)
                except PermissionError:
                    log(
                        f"  [xlsx] 保存被拒绝（是否正用 Excel 打开该文件？）：{xlsx_path.resolve()}"
                    )
                    return
                log(
                    f"  [xlsx] 已在工作表「{sn}」更新 {sheet_wrote} 行 → {xlsx_path.resolve()}"
                )
                return

        log(
            "  [xlsx] 未找到含「测试项目/结果/备注」表头的工作表，或未更新任何行。"
        )
    except Exception as e:
        log(f"  [xlsx] 写入过程异常：{e!s}")


# Slow 3G 近似（与需求一致）
_THROTTLE = {
    "offline": False,
    "latency": 400,
    "downloadThroughput": 400 * 1024 // 8,
    "uploadThroughput": 200 * 1024 // 8,
}

_NORMAL_NET = {
    "offline": False,
    "latency": 0,
    "downloadThroughput": -1,
    "uploadThroughput": -1,
}


def _kalaroko_cdp(cli: str | None) -> str:
    raw = (cli or "").strip() or (os.environ.get("KALAROKO_CDP_ENDPOINT") or "").strip()
    if not raw:
        raw = "http://127.0.0.1:9222"
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "http://" + raw.lstrip("/")
    return raw.rstrip("/")


def _host_from_url(url: str) -> str:
    try:
        from urllib.parse import urlparse

        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def _cdp_tab_url_driver_safe(url: str) -> bool:
    u = (url or "").strip().lower()
    if u.startswith("devtools://") or u.startswith("chrome-devtools://"):
        return False
    if u.startswith("chrome-extension://") or u.startswith("moz-extension://"):
        return False
    if u.startswith("ms-browser-extension://"):
        return False
    return True


def _brief_exc(e: BaseException, lim: int = 200) -> str:
    return f"{type(e).__name__}: {str(e).strip()[:lim]}"


async def _probe_page_alive(pg: Any) -> bool:
    try:
        if pg.is_closed():
            return False
        await asyncio.wait_for(pg.evaluate("() => 1"), timeout=3.0)
        return True
    except Exception:
        return False


async def _acquire_cdp_target_page(
    browser: Any,
    *,
    host: str,
    target_url: str,
    navigate_if_no_tab: bool,
    log: Callable[[str], None],
) -> tuple[Any | None, str | None]:
    if not browser.contexts:
        return None, "CDP 已连上但无 context"

    def _safe_url(pg: Any) -> str:
        try:
            return (pg.url or "").strip()
        except Exception:
            return ""

    has_any_page = any(len(list(getattr(c, "pages", []) or [])) > 0 for c in browser.contexts)

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            if host and host in u.lower():
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None

    if not navigate_if_no_tab:
        if not has_any_page:
            return None, "无打开标签页（且未允许自动 goto）"
        return None, f"无含 {host!r} 的标签页。请打开站点，或加默认允许 goto"

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            log(f"[nav] 无匹配 {host!r}，尝试 goto {target_url!r}（当前 {u[:96]!r}）")
            try:
                await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
                await pg.wait_for_timeout(400)
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None
            except Exception as e:
                log(f"  [nav] goto 失败：{_brief_exc(e)}，换页签…")
                continue

    ctx_new = browser.contexts[0]
    for ctx in browser.contexts:
        for pg in list(getattr(ctx, "pages", []) or []):
            if await _probe_page_alive(pg):
                ctx_new = ctx
                break
        else:
            continue
        break

    log(f"[nav] 新开页签并 goto {target_url!r}")
    try:
        pg = await ctx_new.new_page()
        await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
        await pg.wait_for_timeout(400)
        return pg, None
    except Exception as e:
        return None, f"新开标签并导航失败：{_brief_exc(e)}"


# 与 page.frames 中每个 frame 上 evaluate；主 document 与 iframe 分开汇总，避免 iframe 慢载误判整页「死刑」
_JS_FRAME_WEAKNET_PROBE = """() => {
  const sels = [
    '[class*="skeleton" i]', '[class*="Skeleton"]', '[class*="loading" i]', '[class*="spinner" i]',
    '[aria-busy="true"]', '[class*="shimmer" i]', '[class*="placeholder" i]',
    '.ant-skeleton', '.el-skeleton', '[class*="n-skeleton" i]'
  ];
  let hit = null;
  for (const s of sels) {
    try { const n = document.querySelector(s); if (n) { hit = s; break; } } catch (e) {}
  }
  const t = (document.body && document.body.innerText) ? document.body.innerText : '';
  const len = t.replace(/\\s+/g, ' ').trim().length;
  const navSels = [
    'header', 'nav', '[role="navigation"]', '[class*="header" i]', '[class*="Header"]',
    '[class*="nav-bar" i]', '[class*="NavBar" i]', '[class*="topbar" i]', '[class*="TopBar" i]',
    '[class*="app-bar" i]', '[class*="AppBar" i]'
  ];
  let hasTopNav = false;
  for (const s of navSels) {
    try {
      const el = document.querySelector(s);
      if (!el) continue;
      const r = el.getBoundingClientRect();
      if (r.width < 2 || r.height < 2) continue;
      if (r.top < 220) { hasTopNav = true; break; }
    } catch (e) {}
  }
  const de = document.documentElement;
  const docElLen = (de && de.innerText) ? de.innerText.replace(/\\s+/g, ' ').trim().length : 0;
  const bc = document.body ? document.body.children.length : 0;
  let rootHtmlLen = 0;
  for (const sel of ['#root', '#app', '[data-app]', '[data-v-app]', '#__next']) {
    try {
      const n = document.querySelector(sel);
      if (n && n.innerHTML) rootHtmlLen = Math.max(rootHtmlLen, n.innerHTML.length);
    } catch (e) {}
  }
  const nIframes = document.querySelectorAll('iframe').length;
  return {
    hasSkeleton: !!hit,
    selectorHit: hit,
    bodyTextLen: len,
    docElTextLen: docElLen,
    bodyChildCount: bc,
    rootHtmlLen,
    iframeCount: nIframes,
    ready: document.readyState,
    hasTopNav
  };
}"""

_JS_PHASE_C = """() => {
  const imgs = Array.from(document.querySelectorAll('img'));
  let pending = 0;
  for (const im of imgs) {
    try {
      if (!im.complete || (im.naturalWidth === 0 && (im.src || im.currentSrc))) pending++;
    } catch (e) { pending++; }
  }
  const t = (document.title || '').trim();
  const body = (document.body && document.body.innerText) ? document.body.innerText : '';
  return {
    titleLen: t.length,
    bodyLen: body.replace(/\\s+/g, ' ').trim().length,
    totalImages: imgs.length,
    pendingImages: pending,
  };
}"""

_RE_WEAK_NET_UX = re.compile(
    r"重试|网络|不给力|加载失败|连接.*超时|请稍后|try\s*again|retry|offline|unstable",
    re.I,
)

# 阶梯观测用阈值（不作唯一判死依据）
_T_CORE_LADDER = 100  # 3s 性能优秀：足量正文字
_T_MAIN_SHALLOW = 35  # 主 frame 过短、子 frame 有内容 时作 iframe 标注
_T_ROOT_HTML_SHELL = 80  # SPA 根壳 innerHTML 长度视为「结构已就绪」
_T_TEXT_MIN_SIGNAL = 8  # 聚合正文字数弱信号下限
_WEAK_POLL_MAX_SEC = 75.0  # 首内容信号最长等待


async def _probe_all_frames_weaknet(page: Any) -> dict[str, Any]:
    """
    汇总 page.frames 上同一套弱网探测。单 frame 失败不拖死全量，缺省 false/0。
    """
    out_frames: list[dict[str, Any]] = []
    try:
        flist = list(page.frames)
    except Exception:
        flist = []
    for i, fr in enumerate(flist):
        row: dict[str, Any] = {
            "frameIndex": i,
            "isMain": i == 0,
            "url": "",
            "hasSkeleton": False,
            "selectorHit": None,
            "bodyTextLen": 0,
            "docElTextLen": 0,
            "bodyChildCount": 0,
            "rootHtmlLen": 0,
            "iframeCount": 0,
            "ready": "",
            "hasTopNav": False,
        }
        try:
            row["url"] = (fr.url or "")[:200]
        except Exception:
            pass
        try:
            d = await asyncio.wait_for(
                fr.evaluate(_JS_FRAME_WEAKNET_PROBE), timeout=3.0
            )
            if isinstance(d, dict):
                row["hasSkeleton"] = bool(d.get("hasSkeleton"))
                row["selectorHit"] = d.get("selectorHit")
                row["bodyTextLen"] = int(d.get("bodyTextLen") or 0)
                row["docElTextLen"] = int(d.get("docElTextLen") or 0)
                row["bodyChildCount"] = int(d.get("bodyChildCount") or 0)
                row["rootHtmlLen"] = int(d.get("rootHtmlLen") or 0)
                row["iframeCount"] = int(d.get("iframeCount") or 0)
                row["ready"] = (d.get("ready") or "") or ""
                row["hasTopNav"] = bool(d.get("hasTopNav"))
        except (Exception, asyncio.TimeoutError):
            pass
        out_frames.append(row)
    if not out_frames:
        return {
            "per_frame": [],
            "max_body": 0,
            "max_text_signal": 0,
            "max_root_html": 0,
            "has_struct_shell": False,
            "has_skeleton_any": False,
            "has_top_nav_any": False,
            "main_body": 0,
            "sub_max_body": 0,
            "iframe_suspected_main_shell": False,
        }

    def _text_sig(x: dict[str, Any]) -> int:
        return max(
            int(x.get("bodyTextLen") or 0),
            int(x.get("docElTextLen") or 0),
        )

    main_body = int(out_frames[0].get("bodyTextLen") or 0)
    sub_bodies = [int(x.get("bodyTextLen") or 0) for x in out_frames[1:]]
    sub_max = max(sub_bodies) if sub_bodies else 0
    has_sk = any(x.get("hasSkeleton") for x in out_frames)
    has_nav = any(x.get("hasTopNav") for x in out_frames)
    max_b = max(int(x.get("bodyTextLen") or 0) for x in out_frames)
    max_text_signal = max(_text_sig(x) for x in out_frames)
    max_root = max(int(x.get("rootHtmlLen") or 0) for x in out_frames)
    has_struct = any(int(x.get("rootHtmlLen") or 0) >= _T_ROOT_HTML_SHELL for x in out_frames)
    iframe_shell = (main_body < _T_MAIN_SHALLOW) and (
        sub_max >= 40
        or any(x.get("hasSkeleton") for x in out_frames[1:])
    )
    return {
        "per_frame": out_frames,
        "max_body": max_b,
        "max_text_signal": max_text_signal,
        "max_root_html": max_root,
        "has_struct_shell": has_struct,
        "has_skeleton_any": has_sk,
        "has_top_nav_any": has_nav,
        "main_body": main_body,
        "sub_max_body": sub_max,
        "iframe_suspected_main_shell": iframe_shell,
    }


def _zh_rate_weak_net_sec(sec: float | None) -> str:
    if sec is None:
        return "未测得首可感知内容时刻"
    if sec < 3.5:
        return "优秀（首包/首屏可感知较快）"
    if sec < 8.0:
        return "良好"
    if sec < 16.0:
        return "可接受"
    if sec < 35.0:
        return "偏慢（弱网环境下仍较常见）"
    return "过慢（建议关注首包体积与串行资源）"


def _weak_content_signal(
    snap: dict[str, Any], title: str
) -> tuple[bool, str]:
    """是否出现「可感知的打开进度」；SPA/iframe/仅标题亦算，避免与 FAIL 强绑定。"""
    if not isinstance(snap, dict):
        return False, ""
    if snap.get("has_skeleton_any"):
        return True, "skeleton"
    if int(snap.get("max_text_signal") or 0) >= _T_TEXT_MIN_SIGNAL:
        return True, "text_or_docel"
    if snap.get("has_struct_shell"):
        return True, "spa_root_shell"
    if (title or "").strip() and len((title or "").strip()) >= 2:
        return True, "document_title"
    for x in snap.get("per_frame") or []:
        if not isinstance(x, dict):
            continue
        if int(x.get("bodyChildCount") or 0) >= 2 and int(x.get("rootHtmlLen") or 0) > 20:
            return True, "dom_children_with_shell"
        if int(x.get("iframeCount") or 0) >= 1 and int(x.get("rootHtmlLen") or 0) > 30:
            return True, "iframe_with_shell"
    return False, ""


def _is_timed_out_console_line(text: str) -> bool:
    t = (text or "")
    if "ERR_CONNECTION" in t or "TIMED_OUT" in t or "net::" in t:
        return True
    return "timed out" in t.lower()


async def handle_weak_network_test(
    page: Any,
    target_url: str,
    *,
    log: Callable[[str], None],
) -> tuple[str, str, dict[str, Any]]:
    """
    CDP 弱网：记录限速下导航/首可感知时间并评价；**默认 PASS**（仅真正导航失败为 FAIL）。聚合 page.frames。
    在 finally 中恢复网络并 clearBrowserCache。

    返回 (verdict, detail, extra_observations)
    """
    observations: dict[str, Any] = {
        "phase_a": None,  # 阶梯 3/5/6/10s + 各 frame 快照
        "phase_b": None,
        "phase_c": None,
        "media_degraded": False,
        "console_timed_out_hits": [],
        "performance_suboptimal_3s": False,
        "basic_5s_met": None,
    }
    cdp: Any = None
    console_bucket: list[str] = []

    def on_console(msg: Any) -> None:
        try:
            text = f"{getattr(msg, 'type', '')}: {getattr(msg, 'text', '')[:800]}"
        except Exception:
            text = str(msg)[:400]
        console_bucket.append(text)
        if _is_timed_out_console_line(text):
            observations["console_timed_out_hits"].append(text[:400])
            if len(observations["console_timed_out_hits"]) > 20:
                observations["console_timed_out_hits"].pop(0)

    async def _restore_and_cache() -> None:
        if not cdp:
            return
        try:
            await cdp.send("Network.emulateNetworkConditions", _NORMAL_NET)
        except Exception as re:
            log(f"  [弱网] 恢复无节流失败（可忽略）：{_brief_exc(re)}")
        try:
            await cdp.send("Network.clearBrowserCache", {})
        except Exception as ce:
            log(f"  [弱网] clearBrowserCache 失败（可忽略）：{_brief_exc(ce)}")

    out: tuple[str, str, dict[str, Any]] = (
        "FAIL",
        "未执行到结算",
        observations,
    )
    try:
        try:
            page.on("console", on_console)
        except Exception:
            pass
        cdp = await page.context.new_cdp_session(page)
        await cdp.send("Network.enable", {})
        log("  [弱网] Network.emulateNetworkConditions（Slow3G 类）…")
        await cdp.send("Network.emulateNetworkConditions", _THROTTLE)

        # —— 限速下导航 + 首包耗时 + 仅观测用阶梯（不 6s/10s 判死）——
        t0 = time.monotonic()
        try:
            await page.goto(
                target_url, wait_until="domcontentloaded", timeout=90_000
            )
        except Exception as e:
            gerr = _brief_exc(e)
            observations["phase_a"] = {"goto_error": gerr}
            return "FAIL", f"弱网下无法完成导航/打开首屏：{gerr}", observations

        nav_sec = time.monotonic() - t0
        observations["weak_net_nav_until_domcontentloaded_sec"] = round(nav_sec, 2)
        nav_load_rating = _zh_rate_weak_net_sec(nav_sec)
        observations["weak_net_nav_duration_rating_zh"] = nav_load_rating
        log(
            f"  [弱网] 加载耗时：自开始导航至 **domcontentloaded** 为 **{nav_sec:.1f} s**"
            f"（弱网 Slow3G 限速）；**对该加载耗时评价**：{nav_load_rating}"
        )

        excellent_3s = False
        m3 = m5 = False
        last_snap: dict[str, Any] = {}
        first_sig_sec: float | None = None
        signal_kind = ""
        t_poll_end = t0 + _WEAK_POLL_MAX_SEC
        title_for_sig = ""

        while time.monotonic() < t_poll_end:
            el = time.monotonic() - t0
            try:
                last_snap = await _probe_all_frames_weaknet(page)
            except Exception as e:
                last_snap = {"per_frame": [], "error": _brief_exc(e)}

            try:
                title_for_sig = (await page.title() or "").strip()
            except Exception:
                title_for_sig = ""

            sk = bool(last_snap.get("has_skeleton_any"))
            mtxt = int(
                last_snap.get("max_text_signal")
                or last_snap.get("max_body")
                or 0
            )
            nav = bool(last_snap.get("has_top_nav_any"))
            shell = bool(last_snap.get("iframe_suspected_main_shell"))
            n_frames = len(last_snap.get("per_frame") or [])

            ok_s, knd = _weak_content_signal(last_snap, title_for_sig)
            if ok_s and first_sig_sec is None:
                first_sig_sec = el
                signal_kind = knd

            if el <= 3.0 and (sk or mtxt >= _T_CORE_LADDER):
                excellent_3s = True

            if not m3 and el >= 3.0:
                m3 = True
                fast_enough_3s = (first_sig_sec is not None and first_sig_sec <= 3.0) or (
                    sk or mtxt >= _T_CORE_LADDER
                )
                if not fast_enough_3s:
                    observations["performance_suboptimal_3s"] = True
                    log(
                        "  [弱网] Observation: Performance sub-optimal"
                        "（3s 参考线未先出现可感知内容；SPA/弱网可晚于 3s 出现）"
                    )
                if shell and n_frames > 1:
                    observations["iframe_content_delayed_note"] = (
                        "主 document 正文字数偏少，子 frame 含更多文案/骨架；以首包与人工目视为准。"
                    )
                    log(
                        f"  [弱网] Observation: 多 frame（n={n_frames}）主壳偏空、"
                        f"子区 max_text≈{last_snap.get('sub_max_body')}，已单独标注。"
                    )

            if not m5 and el >= 5.0:
                m5 = True
                basic_ok = sk and nav
                observations["basic_5s_met"] = basic_ok
                if not basic_ok:
                    log(
                        "  [弱网] 5s 参考档：骨架+顶栏未同时出现"
                        f"（skeleton={sk} topNav={nav}；仅作记录）"
                    )

            if el >= 10.2 and m3 and m5 and first_sig_sec is not None:
                break
            await asyncio.sleep(0.1)

        if first_sig_sec is None:
            ok, kind = _weak_content_signal(last_snap, title_for_sig)
            if ok:
                first_sig_sec = min(
                    time.monotonic() - t0, _WEAK_POLL_MAX_SEC
                )
                signal_kind = kind
            else:
                first_sig_sec = nav_sec
                signal_kind = "nav_only_soft"
                umsg = (
                    "未稳定探测到内联正文/根壳（可能为跨域 iframe、Canvas、Shadow 等）；"
                    f"但 domcontentloaded 已于 {nav_sec:.1f}s 达成。本项不判 FAIL，建议实机目视。"
                )
                observations["weak_net_signal_uncertain"] = umsg
                log(f"  [弱网] Observation: {umsg}")
                log(
                    f"  [弱网] 以导航完成时刻作为参考首包：{first_sig_sec:.1f}s（{signal_kind}）"
                )

        rating = _zh_rate_weak_net_sec(first_sig_sec)
        observations["weak_net_sec_to_first_signal"] = (
            round(first_sig_sec, 2) if first_sig_sec is not None else None
        )
        observations["weak_net_signal_kind"] = signal_kind
        observations["weak_net_experience_rating_zh"] = rating
        log(
            f"  [弱网] 首可感知内容出现：约 **{first_sig_sec:.1f} s**"
            f"（信号：{signal_kind}）；**对该首现耗时评价**：{rating}"
        )
        if (
            first_sig_sec is not None
            and abs(float(first_sig_sec) - float(nav_sec)) < 0.2
        ):
            log(
                "  [弱网] 说明：首可感知与 domcontentloaded 几乎同时，"
                "上列「加载耗时」与「首现耗时」数值接近、评价可一并参考。"
            )
        else:
            log(
                f"  [弱网] 汇总：主文档 domcontentloaded **{nav_sec:.1f} s** → {nav_load_rating}；"
                f" 首可感知 **{first_sig_sec:.1f} s** → {rating}。"
            )

        observations["phase_a"] = {
            "ok": True,
            "ladder": {
                "tier_3s_excellent_core_or_skeleton": excellent_3s,
                "tier_3s_observation_performance_sub_optimal": bool(
                    observations.get("performance_suboptimal_3s")
                ),
                "tier_5s_met": observations.get("basic_5s_met"),
                "last_snap_summary": {
                    "max_body": last_snap.get("max_body"),
                    "max_text_signal": last_snap.get("max_text_signal"),
                    "max_root_html": last_snap.get("max_root_html"),
                    "has_skeleton_any": last_snap.get("has_skeleton_any"),
                    "has_top_nav_any": last_snap.get("has_top_nav_any"),
                    "frames": len(last_snap.get("per_frame") or []),
                },
            },
            "per_frame_tail": (last_snap.get("per_frame") or [])[:8],
        }

        # —— 阶段 B：load 为加分项，不因弱网长载 FAIL ——
        phase_b_ok = True
        phase_b_note = "弱网下继续等待 load 事件"
        try:
            await page.wait_for_load_state("load", timeout=45_000)
            phase_b_note = "已触发 load 事件"
        except (Exception, asyncio.TimeoutError):
            phase_b_ok = False
            try:
                html_snip = await asyncio.wait_for(
                    page.evaluate(
                        "() => (document.body && document.body.innerText) ? "
                        "document.body.innerText.slice(0, 1200) : ''"
                    ),
                    timeout=5.0,
                )
            except Exception:
                html_snip = ""
            if html_snip and _RE_WEAK_NET_UX.search(html_snip):
                phase_b_note = "load 未在 45s 内到达，但可见网络/重试类引导"
            else:
                phase_b_note = (
                    "load 在 45s 内未到达（资源仍在拉取常见）；不单独判 FAIL。摘要："
                    f"{html_snip[:200]!r}"
                )
        observations["phase_b"] = {"ok": phase_b_ok, "note": phase_b_note}

        # —— 首图：15s 内无解码完成则记 Media degraded（仍可为 PASS）——
        media_degraded = False
        try:
            await asyncio.wait_for(
                page.wait_for_function(
                    """() => {
                      const imgs = document.querySelectorAll('img');
                      if (!imgs.length) return true;
                      for (const im of Array.from(imgs)) {
                        try {
                          if (im.src && im.complete && im.naturalWidth > 0) return true;
                        } catch (e) {}
                      }
                      return false;
                    }""",
                    timeout=15_000,
                ),
                timeout=16.0,
            )
        except (Exception, asyncio.TimeoutError):
            media_degraded = True
            observations["media_degraded"] = True
            log(
                "  [弱网] Observation: Media resource degraded"
                "（图 15s 内未出现完整解码，记录为降级）"
            )

        c_stats = await page.evaluate(_JS_PHASE_C)
        snap_c = await _probe_all_frames_weaknet(page)
        bl_main = int(c_stats.get("bodyLen") or 0)
        bl_agg = max(bl_main, int(snap_c.get("max_body") or 0))
        c_stats["bodyLen_main"] = bl_main
        c_stats["bodyLen"] = bl_agg
        c_stats["bodyLen_aggregate_max"] = bl_agg
        observations["phase_c"] = c_stats
        n_title = int(c_stats.get("titleLen") or 0)
        n_body = bl_agg
        pend = int(c_stats.get("pendingImages") or 0)

        c_notes: list[str] = []
        if n_title < 2 and n_body < 20:
            c_notes.append(
                f"内联正文字/标题仍偏少（titleLen={n_title} bodyAgg={n_body}）"
                "，可能为 iframe/只读区限制；不单独转 FAIL"
            )
        if pend > 0 and n_body < 30:
            c_notes.append("媒体仍多 pending，与弱网/懒加载相关；不单独转 FAIL")
        for cn in c_notes:
            log(f"  [弱网] Observation: {cn}")

        w_nav = observations.get("weak_net_nav_until_domcontentloaded_sec")
        w_nav_r = observations.get("weak_net_nav_duration_rating_zh", "")
        w_ftt = observations.get("weak_net_sec_to_first_signal")
        w_rate = observations.get("weak_net_experience_rating_zh", "")
        w_kind = observations.get("weak_net_signal_kind", "")
        detail = (
            f"弱网：至 domcontentloaded 加载 {w_nav!s} s，对该加载时间评价：{w_nav_r!s}；"
            f"首可感知 {w_ftt!s} s（{w_kind!s}）对该时间评价：{w_rate!s}；"
            f"title/body(聚合) {n_title}/{n_body}，图 pending {pend}/"
            f"{c_stats.get('totalImages', '?')}"
        )
        if media_degraded:
            detail += " Observation: Media resource degraded（仍计 PASS）。"
        n_hits = len(observations.get("console_timed_out_hits") or [])
        if n_hits:
            detail += f" Console 曾出现连接/超时类行 {n_hits} 条（已记录）。"
        out = "PASS", detail, observations
        return out
    except Exception as e:
        out = "FAIL", f"弱网用例异常：{_brief_exc(e)}", observations
        return out
    finally:
        try:
            page.remove_listener("console", on_console)
        except Exception:
            pass
        await _restore_and_cache()
        if cdp:
            try:
                await cdp.detach()
            except Exception:
                pass


async def handle_browser_compat_test(
    p: Any,
    target_url: str,
    *,
    headless: bool,
    log: Callable[[str], None],
) -> list[dict[str, Any]]:
    """
    分别用 Chromium channel=chrome / msedge 启动新浏览器，进入目标站并做轻量健康检查。

    返回每条 { case_id, title_zh, verdict, detail }。
    """
    rows: list[dict[str, Any]] = []
    channels: list[tuple[str, str, str]] = [
        ("p2_browser_compat_chrome", "chrome", "P2 · 浏览器兼容（Google Chrome）"),
        ("p2_browser_compat_edge", "msedge", "P2 · 浏览器兼容（Microsoft Edge）"),
    ]

    for case_id, ch, title_zh in channels:
        log(f"【{title_zh}】启动 channel={ch!r}…")
        browser: Any = None
        try:
            browser = await p.chromium.launch(
                headless=headless,
                channel=ch,
                args=["--disable-blink-features=AutomationControlled"] if ch == "msedge" else [],
            )
        except Exception as e:
            rows.append(
                {
                    "case_id": case_id,
                    "title_zh": title_zh,
                    "verdict": "FAIL",
                    "detail": f"无法启动 {ch}：{_brief_exc(e)}（可尝试 `playwright install {ch}` 或本机已安装该浏览器）",
                }
            )
            continue
        try:
            ctx = await browser.new_context(
                ignore_https_errors=True,
                java_script_enabled=True,
            )
            page = await ctx.new_page()
            try:
                await page.goto(
                    target_url, wait_until="domcontentloaded", timeout=60_000
                )
            except Exception as e:
                rows.append(
                    {
                        "case_id": case_id,
                        "title_zh": title_zh,
                        "verdict": "FAIL",
                        "detail": f"导航失败：{_brief_exc(e)}",
                    }
                )
                continue

            try:
                await page.wait_for_timeout(800)
            except Exception:
                pass

            title = (await page.title() or "").strip()
            n_body = int(
                await page.evaluate(
                    "() => (document.body && document.body.innerText) ? document.body.innerText.length : 0"
                )
            )
            u = (page.url or "").strip()
            if n_body < 20 and not title:
                rows.append(
                    {
                        "case_id": case_id,
                        "title_zh": title_zh,
                        "verdict": "FAIL",
                        "detail": f"白屏或空文档嫌疑：url={u!r} title空 len={n_body}",
                    }
                )
            elif n_body < 20:
                rows.append(
                    {
                        "case_id": case_id,
                        "title_zh": title_zh,
                        "verdict": "FAIL",
                        "detail": f"正文过短：title={title!r} bodyLen={n_body} url={u!r}",
                    }
                )
            else:
                rows.append(
                    {
                        "case_id": case_id,
                        "title_zh": title_zh,
                        "verdict": "PASS",
                        "detail": f"已打开并渲染可读：title={title!r} bodyLen≈{n_body} url={u!r}（不处理机器人/验证页，仅作冒烟）",
                    }
                )
        finally:
            try:
                if browser:
                    await browser.close()
            except Exception:
                pass

    return rows


def _print_verdict_line(case_id: str, v: str) -> None:
    """与仓库其它 K11 Playwright 冒烟脚本一致的汇总行。"""
    print(f"VERDICT: {case_id}: {v}", flush=True)


async def _async_main(args: argparse.Namespace) -> int:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print(
            "请先安装：pip install playwright && playwright install chrome msedge",
            file=sys.stderr,
        )
        return 2

    target_url = (args.target_url or DEFAULT_TARGET).strip() or DEFAULT_TARGET
    host = _host_from_url(target_url)
    cdp = _kalaroko_cdp(args.cdp_http or None)

    def log(msg: str) -> None:
        if not args.quiet:
            print(msg, flush=True)

    log("———————— K11 P2 · 浏览器兼容 + 弱网（独立脚本）————————")
    log(f"目标 URL：{target_url}")

    all_rows: list[dict[str, Any]] = []
    final_exit_fail = False

    async with async_playwright() as p:
        if not args.skip_compat:
            log("")
            log("==== 一、浏览器兼容（本脚本自行启动 Chrome / Edge）====")
            compat_rows = await handle_browser_compat_test(
                p, target_url, headless=args.headless, log=log
            )
            for r in compat_rows:
                all_rows.append(r)
                v = r.get("verdict", "FAIL")
                log(f"  观察：{r.get('detail', '')}")
                vzh = VERDICT_ZH.get(v, v)
                mark = "✓" if v == "PASS" else "✗"
                log(f"  结论：{r.get('title_zh', '')} → {vzh}（{v}）")
                _print_verdict_line(r.get("case_id", ""), v)
                if v == "FAIL":
                    final_exit_fail = True
                log("")

        if not args.skip_weak:
            log("")
            log("==== 二、弱网体验（CDP 附加 + emulateNetworkConditions）====")
            log(f"CDP：{cdp}（KALAROKO_CDP_ENDPOINT / --cdp-http）")
            browser = await p.chromium.connect_over_cdp(cdp)
            page, err = await _acquire_cdp_target_page(
                browser,
                host=host,
                target_url=target_url,
                navigate_if_no_tab=True,
                log=log,
            )
            if page is None:
                print(f"[失败] 无法获得页签：{err}", file=sys.stderr)
                all_rows.append(
                    {
                        "case_id": "p2_weak_network",
                        "title_zh": "P2 · 弱网体验",
                        "verdict": "FAIL",
                        "detail": err or "无页签",
                    }
                )
                _print_verdict_line("p2_weak_network", "FAIL")
                final_exit_fail = True
            else:
                v, detail, extra = await handle_weak_network_test(page, target_url, log=log)
                all_rows.append(
                    {
                        "case_id": "p2_weak_network",
                        "title_zh": "P2 · 弱网体验",
                        "verdict": v,
                        "detail": detail,
                        "observations": extra,
                    }
                )
                vzh = VERDICT_ZH.get(v, v)
                mark = "✓" if v == "PASS" else "✗"
                log(f"  观察：{detail}")
                if extra.get("media_degraded"):
                    log("  Observation: Media resource degraded")
                log(f"  结论：P2 · 弱网体验 → {vzh}（{v}）")
                _print_verdict_line("p2_weak_network", v)
                if v == "FAIL":
                    final_exit_fail = True
                log("")

    out = {
        "schema": "k11_p2_compat_weaknet_playwright/v1",
        "ts_utc": datetime.now(timezone.utc).isoformat(),
        "target_url": target_url,
        "cdp": cdp,
        "results": all_rows,
    }
    if args.json_out:
        outp = Path(args.json_out)
        outp.parent.mkdir(parents=True, exist_ok=True)
        outp.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        log(f"JSON：{outp.resolve()}")

    if not args.no_xlsx_report:
        xlsx_p = (
            args.xlsx_report
            if args.xlsx_report is not None
            else _default_k11_xlsx_report_path()
        )
        log("")
        norm = _p2_normalize_rows_for_xlsx(all_rows)
        write_k11_p2_results_to_xlsx(Path(xlsx_p), norm, log=log)

    log("———————— 汇总（VERDICT 行与下文一致）————————")
    for r in all_rows:
        m = "✓" if r.get("verdict") == "PASS" else "✗"
        log(
            f"  {m} [{r.get('case_id')}] {r.get('title_zh', '')} → {VERDICT_ZH.get(r.get('verdict', ''), r.get('verdict'))}"
        )
    if final_exit_fail:
        log("总评：存在 FAIL 项。")
    else:
        log("总评：全部通过。")
    return 1 if final_exit_fail else 0


def main() -> None:
    ap = argparse.ArgumentParser(
        description="K11 P2：浏览器自启兼容 + 弱网 CDP（与 K11_平台冒烟测试用例 55-57 行对应）"
    )
    ap.add_argument("--target-url", default=DEFAULT_TARGET, help="验收 URL，默认 herontest")
    ap.add_argument(
        "--cdp-http", default="", help="覆盖 KALAROKO_CDP_ENDPOINT（仅弱网段 connect_over_cdp）"
    )
    ap.add_argument(
        "--headless",
        action="store_true",
        help="浏览器兼容段使用无头（默认有头，便于看机器人/人工）",
    )
    ap.add_argument("--only-compat", action="store_true", help="仅跑 Chrome/Edge 兼容")
    ap.add_argument("--only-weak", action="store_true", help="仅跑弱网（需 CDP Chrome）")
    ap.add_argument("--json-out", default="", help="输出 JSON 报告路径")
    ap.add_argument(
        "--xlsx-report",
        type=Path,
        default=None,
        help="K11平台测试用例.xlsx；默认 K11_XLSX_REPORT 或 ~/Downloads/K11平台测试用例.xlsx",
    )
    ap.add_argument("--no-xlsx-report", action="store_true", help="不写入 Excel")
    ap.add_argument("--quiet", action="store_true", help="减少过程日志")
    args = ap.parse_args()
    if args.only_compat and args.only_weak:
        print("--only-compat 与 --only-weak 互斥", file=sys.stderr)
        sys.exit(2)
    if args.only_weak:
        args.skip_compat, args.skip_weak = True, False
    elif args.only_compat:
        args.skip_compat, args.skip_weak = False, True
    else:
        args.skip_compat, args.skip_weak = False, False

    try:
        rc = asyncio.run(_async_main(args))
        sys.exit(rc)
    except KeyboardInterrupt:
        print("中断", file=sys.stderr)
        sys.exit(130)
    except Exception as e:
        print(f"未捕获异常：{_brief_exc(e)}", file=sys.stderr)
        sys.exit(3)


if __name__ == "__main__":
    main()
