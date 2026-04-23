#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K11 平台冒烟 · P0（文档《K11_平台冒烟测试用例》约 31–39 行）

与 ``test_k11_p1_skill_herontest_playwright.py`` 相同：**不经过 L3**，本机 Playwright
``connect_over_cdp`` 附加已由 ``launch_chrome_debug.ps1`` 启动的 Chrome。

覆盖（自动化能力范围内）：
  P0 环境访问、首页加载、页面标题、Play Now!、分类 All/1vs1/Party/Live、
  游戏卡片抽检、Console 抽样、本次更新点（SKIP 人工）
  （「关键卡片点击」已不在本脚本执行，可用 ``test_k11_p0_key_card_kalaroko_playwright.py`` 单独跑。）

前置：
  - ``KALAROKO_CDP_ENDPOINT`` 或 ``--cdp-http``（默认 http://127.0.0.1:9222）
  - ``pip install playwright``（connect_over_cdp 仍需驱动）

用法（仓库根）：
  python scripts/test_k11_p0_platform_smoke_playwright.py
  python scripts/test_k11_p0_platform_smoke_playwright.py --target-url https://www.kalaroko.com/
  python scripts/test_k11_p0_platform_smoke_playwright.py --target-url https://www.herontest.xin/
  python scripts/test_k11_p0_platform_smoke_playwright.py --require-existing-tab
  python scripts/test_k11_p0_platform_smoke_playwright.py -v --json-out out/k11_p0.json

默认：无含目标域的页签时，会在**当前窗口最后一个页签**内自动 goto ``--target-url``；
若必须已手动打开站点、禁止自动跳转，请加 ``--require-existing-tab``。

运行结束可将各条「结果 / 备注」写入 ``K11平台测试用例.xlsx``（列名与 ``docs/K11平台测试用例_冒烟测试用例.csv`` 对齐）：
  ``--xlsx-report`` / ``K11_XLSX_REPORT`` / ``~/Downloads/K11平台测试用例.xlsx``；``--no-xlsx-report`` 关闭。需 ``openpyxl``；写入前请关闭 Excel。

退出码：0 无 FAIL；1 存在 FAIL；2 CDP/环境；3 未捕获异常。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import time
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env", encoding="utf-8")
except ImportError:
    pass
except OSError:
    pass

# 与文档示例一致；可用 --target-url 指向 herontest 等
DEFAULT_TARGET = "https://www.kalaroko.com/"

CASE_DEFS: list[tuple[str, str]] = [
    ("p0_env_access", "环境访问"),
    ("p0_home_load", "首页加载"),
    ("p0_page_title", "页面标题"),
    ("p0_play_now", "主按钮可用（Play Now!）"),
    ("p0_category_tabs", "分类切换（All / 1 vs 1 / Party / Live）"),
    ("p0_game_cards", "游戏卡片展示"),
    ("p0_console_clean", "页面无严重报错（Console 抽样）"),
    ("p0_release_checklist", "本次更新点验证"),
]

CASE_TITLE_ZH = {k: v for k, v in CASE_DEFS}

VERDICT_ZH = {
    "PASS": "通过",
    "FAIL": "未通过",
    "SKIP": "跳过",
    "BLOCKED": "阻塞",
}

P0_CASE_TO_XLSX_TEST_ITEM_KEY: dict[str, str] = {
    "p0_env_access": "环境访问",
    "p0_home_load": "首页加载",
    "p0_page_title": "页面标题",
    "p0_play_now": "主按钮可用",
    "p0_category_tabs": "分类切换",
    "p0_game_cards": "游戏卡片展示",
    "p0_console_clean": "无严重报错",
    "p0_release_checklist": "本次更新点",
}

_XLSX_REMARK_MAX_LEN = 32000


def _default_k11_xlsx_report_path() -> Path:
    env = (os.environ.get("K11_XLSX_REPORT") or "").strip()
    if env:
        return Path(env)
    return Path.home() / "Downloads" / "K11平台测试用例.xlsx"


def _find_smoke_sheet_header(ws: Any) -> tuple[int, int, int, int] | None:
    max_r = min(int(ws.max_row or 1), 45)
    max_c = min(int(ws.max_column or 1), 40)
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = str(v).strip() if v is not None else ""
            if s and s not in col_map:
                col_map[s] = c
        if "结果" not in col_map or "备注" not in col_map:
            continue
        item_col = None
        for k in ("测试项目", "测试项", "用例名称"):
            if k in col_map:
                item_col = col_map[k]
                break
        if item_col is None:
            continue
        return r, item_col, col_map["结果"], col_map["备注"]
    return None


def write_k11_p0_results_to_xlsx(
    xlsx_path: Path,
    results: list[dict[str, Any]],
    *,
    log: Callable[[str], None],
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [xlsx] 未安装 openpyxl，跳过写入（可：pip install openpyxl）")
        return

    if not xlsx_path.is_file():
        log(f"  [xlsx] 文件不存在，跳过：{xlsx_path}")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        log(f"  [xlsx] 打开工作簿失败：{e!s}")
        return

    ordered_names: list[str] = [n for n in wb.sheetnames if "冒烟" in n]
    ordered_names.extend(n for n in wb.sheetnames if n not in ordered_names)
    verdict_cell = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP", "BLOCKED": "BLOCKED"}

    try:
        for sn in ordered_names:
            ws = wb[sn]
            parsed = _find_smoke_sheet_header(ws)
            if not parsed:
                continue
            hdr, c_item, c_res, c_note = parsed
            last_r = max(int(ws.max_row or hdr), hdr + 5, 80)
            sheet_wrote = 0
            for resrow in results:
                cid = str(resrow.get("case") or "")
                key = P0_CASE_TO_XLSX_TEST_ITEM_KEY.get(cid)
                if not key:
                    continue
                v = str(resrow.get("verdict") or "")
                detail = str(resrow.get("detail") or "")
                cell_v = verdict_cell.get(v, v)
                remark = (detail or "")[:_XLSX_REMARK_MAX_LEN]
                matched = False
                for r in range(hdr + 1, last_r + 1):
                    raw = ws.cell(row=r, column=c_item).value
                    text = str(raw).strip() if raw is not None else ""
                    if not text:
                        continue
                    if key in text:
                        ws.cell(row=r, column=c_res, value=cell_v)
                        ws.cell(row=r, column=c_note, value=remark)
                        sheet_wrote += 1
                        matched = True
                        break
                if not matched:
                    log(f"  [xlsx] 未匹配行：case={cid} 关键字={key!r} sheet={sn!r}")
            if sheet_wrote:
                try:
                    wb.save(xlsx_path)
                except PermissionError:
                    log(
                        f"  [xlsx] 保存被拒绝（是否正用 Excel 打开该文件？）：{xlsx_path.resolve()}"
                    )
                    return
                log(
                    f"  [xlsx] 已在工作表「{sn}」更新 {sheet_wrote} 行 → {xlsx_path.resolve()}"
                )
                return

        log(
            "  [xlsx] 未找到含「测试项目/结果/备注」表头的工作表，或未更新任何行。"
        )
    except Exception as e:
        log(f"  [xlsx] 写入过程异常：{e!s}")


def _kalaroko_cdp(cli: str | None) -> str:
    raw = (cli or "").strip() or (os.environ.get("KALAROKO_CDP_ENDPOINT") or "").strip()
    if not raw:
        raw = "http://127.0.0.1:9222"
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "http://" + raw.lstrip("/")
    return raw.rstrip("/")


def _host_from_url(url: str) -> str:
    try:
        from urllib.parse import urlparse

        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def _home_feed_url(target: str) -> str:
    """与 P1 一致：站点根路径大厅，便于分类条与卡片区稳定。"""
    from urllib.parse import urlparse, urlunparse

    t = (target or DEFAULT_TARGET).strip() or DEFAULT_TARGET
    p = urlparse(t)
    if p.scheme and p.netloc:
        return urlunparse((p.scheme, p.netloc, "/", "", "", ""))
    return t.rstrip("/") + "/" if t else DEFAULT_TARGET


def _needs_goto_home_feed(current_url: str) -> bool:
    u = (current_url or "").lower()
    if "/my/" in u or "/me/" in u:
        return True
    if re.search(r"/(profile|account|wallet|settings)(/|$)", u):
        return True
    if "party-hubs" in u:
        return True
    if "app_tabbar=no" in u:
        return True
    return False


async def _ensure_on_home_feed(
    page: Any, target_url: str, log: Callable[[str], None] | None
) -> None:
    home = _home_feed_url(target_url)
    try:
        cur = page.url or ""
    except Exception:
        cur = ""
    if not _needs_goto_home_feed(cur):
        if log:
            log(f"  [诊断·P0] 无需回大厅：{cur!r}")
        return
    if log:
        log(f"  [诊断·P0] 回大厅：{cur!r} → goto {home!r}")
    await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(600)


# 分类条：文案有时在子节点内，与底栏相同策略（P1 _JS_CLICK_TAB_FROM_LABEL）
_JS_CLICK_TAB_FROM_LABEL = """(el) => {
  let n = el.parentElement;
  for (let i = 0; i < 12 && n; i++) {
    const tag = (n.tagName || '').toUpperCase();
    const cls = (typeof n.className === 'string') ? n.className : '';
    const role = n.getAttribute && n.getAttribute('role');
    const isLabelOnly = /_item_label_/i.test(cls);
    if (!isLabelOnly && (tag === 'BUTTON' || tag === 'A' || role === 'button' || role === 'tab'
        || /\\bitem\\b/i.test(cls) || (cls.includes('_item_') && !cls.includes('_item_label_')))) {
      n.click();
      return { ok: true, step: i, tag, cls: cls.slice(0, 72) };
    }
    n = n.parentElement;
  }
  if (el.parentElement) {
    el.parentElement.click();
    return { ok: true, fallback: 'parent', tag: el.parentElement.tagName };
  }
  el.click();
  return { ok: true, fallback: 'label-only', tag: el.tagName };
}"""


def _cdp_tab_url_driver_safe(url: str) -> bool:
    """排除 DevTools / 扩展页等：对其 goto 易导致 TargetClosedError 或无效。"""
    u = (url or "").strip().lower()
    if u.startswith("devtools://") or u.startswith("chrome-devtools://"):
        return False
    if u.startswith("chrome-extension://") or u.startswith("moz-extension://"):
        return False
    if u.startswith("ms-browser-extension://"):
        return False
    return True


async def _probe_page_alive(pg: Any) -> bool:
    """参考 kalaroko_capture_page_metrics：避免选中已关闭或不可执行的页签引用。"""
    try:
        if pg.is_closed():
            return False
        await asyncio.wait_for(pg.evaluate("() => 1"), timeout=3.0)
        return True
    except Exception:
        return False


def _brief_exc(e: BaseException, lim: int = 180) -> str:
    return f"{type(e).__name__}: {str(e).strip()[:lim]}"


def _is_benign_console_line(text: str) -> bool:
    t = (text or "").lower()
    if "font-size:0" in t and "nan" in t:
        return True
    if "failed to load resource" in t and "favicon" in t:
        return True
    if "resizeobserver" in t:
        return True
    return False


async def _eval_safe(page: Any, expr: str, *, timeout: float = 12.0) -> Any:
    return await asyncio.wait_for(page.evaluate(expr), timeout=timeout)


async def _acquire_cdp_target_page(
    browser: Any,
    *,
    host: str,
    target_url: str,
    navigate_if_no_tab: bool,
    log: Callable[[str], None],
) -> tuple[Any | None, str | None]:
    """
    扫描全部 browser.contexts（与仅 contexts[0] 相比更贴近真实多窗口场景）：
    优先选用 URL 含目标 host 且可驱动的页签；否则在允许导航时依次尝试
    安全 URL 的存活页签 goto，失败则 new_page 后 goto（避免末页签常为 DevTools）。
    """
    if not browser.contexts:
        return None, "CDP 已连上但无 context"

    def _safe_url(pg: Any) -> str:
        try:
            return (pg.url or "").strip()
        except Exception:
            return ""

    has_any_page = any(len(list(getattr(c, "pages", []) or [])) > 0 for c in browser.contexts)

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            if host and host in u.lower():
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None

    if not navigate_if_no_tab:
        if not has_any_page:
            return None, "无打开标签页（且未允许自动 goto）"
        return (
            None,
            f"无含 {host!r} 的标签页。请打开站点，或去掉 --require-existing-tab 以允许自动 goto",
        )

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            log(f"[nav] 无匹配 {host!r}，尝试在存活页签 goto {target_url!r}（当前 {u[:96]!r}）")
            try:
                await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
                await pg.wait_for_timeout(400)
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None
            except Exception as e:
                log(f"  [nav] 该页签 goto 失败：{_brief_exc(e)}，换候选或新开标签…")
                continue

    ctx_new = browser.contexts[0]
    for ctx in browser.contexts:
        for pg in list(getattr(ctx, "pages", []) or []):
            if await _probe_page_alive(pg):
                ctx_new = ctx
                break
        else:
            continue
        break

    log(f"[nav] 无可用页签可复用，在 context 中新开标签并 goto {target_url!r}")
    try:
        pg = await ctx_new.new_page()
        await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
        await pg.wait_for_timeout(400)
        return pg, None
    except Exception as e:
        return None, f"新开标签并导航失败：{_brief_exc(e)}"


async def _ensure_target_page(
    page: Any,
    target_url: str,
    *,
    log: Callable[[str], None],
    navigate_if_no_tab: bool,
    host: str,
) -> tuple[bool, str]:
    """环境访问：当前标签含目标 host；必要时 goto。"""
    try:
        u = (page.url or "").strip()
    except Exception:
        u = ""
    if host and host in u.lower():
        return True, f"当前页已为目标域：{u!r}"
    if navigate_if_no_tab:
        log(f"  [诊断·P0] goto {target_url!r}")
        await page.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(400)
        return True, f"已 goto {target_url!r}"
    return (
        False,
        f"当前 URL {u!r} 不含 host {host!r}；请加 --navigate-if-no-tab 或先手动打开站点。",
    )


async def _run_p0_home_load(page: Any) -> tuple[str, str]:
    try:
        rs = await _eval_safe(page, "() => document.readyState")
        if rs != "complete":
            try:
                await page.wait_for_load_state("load", timeout=12_000)
            except Exception:
                pass
            rs = await _eval_safe(page, "() => document.readyState")
        n = await _eval_safe(
            page,
            "() => (document.body && document.body.innerText) ? document.body.innerText.length : 0",
        )
        if int(n or 0) < 80:
            return ("FAIL", f"首屏正文过短（innerText 长度 {n}），疑似白屏或壳页。")
        return ("PASS", f"readyState={rs!r}，正文长度≈{n}。")
    except Exception as e:
        return ("FAIL", f"首屏/readyState 检测异常：{_brief_exc(e)}")


async def _run_p0_page_title(page: Any) -> tuple[str, str]:
    try:
        t = (await page.title() or "").strip()
    except Exception as e:
        return ("FAIL", f"读取标题失败：{_brief_exc(e)}")
    if re.search(r"KalaroKo", t, re.I):
        return ("PASS", f"标题匹配 KalaroKo：{t!r}")
    return ("FAIL", f"标题未包含 KalaroKo：{t!r}")


async def _click_locator_robust(loc: Any, page: Any, *, timeout_ms: int = 7000) -> None:
    """Playwright click → P1 式父级 DOM click → 合成 click（应对遮罩/可见性误判）。"""
    await loc.wait_for(state="attached", timeout=min(10_000, timeout_ms + 2000))
    try:
        await loc.scroll_into_view_if_needed(timeout=4000)
    except Exception:
        pass
    try:
        await loc.click(timeout=timeout_ms, force=True)
        return
    except Exception:
        pass
    try:
        await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
        return
    except Exception:
        pass
    await loc.evaluate(
        "e => e.dispatchEvent(new MouseEvent('click', {bubbles:true,cancelable:true,view:window}))"
    )


def _category_tab_name_re(label: str) -> re.Pattern[str]:
    if re.sub(r"\s+", " ", label.strip()).lower() in ("1 vs 1", "1vs1"):
        return re.compile(r"^\s*1\s*v\s*s\s*1\s*$", re.I)
    return re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)


async def _run_p0_play_now(page: Any) -> tuple[str, str]:
    """可见 + 有有效点击区域；不执行 click，避免跳走导致后续分类/卡片失败。"""
    pat = re.compile(r"Play\s*Now", re.I)
    try:
        loc = page.get_by_role("button", name=pat).first
        if await loc.count() < 1:
            loc = page.get_by_text(pat).first
        if await loc.count() < 1:
            return ("FAIL", "未找到「Play Now!」按钮或等价文案。")
        await loc.wait_for(state="visible", timeout=5000)
        await loc.scroll_into_view_if_needed(timeout=4000)
        box = await loc.bounding_box()
        if not box or box.get("width", 0) < 4 or box.get("height", 0) < 4:
            return ("FAIL", "Play Now! 可见但点击区域异常（bounding box）。")
        return ("PASS", "Play Now! 可见且具备有效点击区域（未实际点击以防打断后续用例）。")
    except Exception as e:
        return ("FAIL", f"Play Now 不可用：{_brief_exc(e)}")


async def _first_home_category_tablist(page: Any) -> Any | None:
    """首页中部筛选条：返回该 tablist 容器 Locator（勿用全页 .last，避免点到底栏同名 tab）。"""
    lists = page.locator('[role="tablist"]')
    nl = await lists.count()
    for idx in range(nl):
        tl = lists.nth(idx)
        tabs = tl.locator('[role="tab"]')
        tc = await tabs.count()
        if tc < 4:
            continue
        has_all = await tl.get_by_text(re.compile(r"^All$", re.I)).count()
        if has_all < 1:
            continue
        return tl
    return None


async def _run_p0_category_tabs(page: Any) -> tuple[str, str]:
    labels = ["All", "1 vs 1", "Party", "Live"]
    try:
        tl = await _first_home_category_tablist(page)
        if tl is None:
            return ("FAIL", "未找到含 All 且至少 4 项的 [role=tablist]（分类条）。")
        try:
            await tl.scroll_into_view_if_needed(timeout=4000)
        except Exception:
            pass
        clicked: list[str] = []
        for lab in labels:
            try:
                name_re = _category_tab_name_re(lab)
                raw = tl.get_by_role("tab", name=name_re)
                if await raw.count() < 1:
                    raw = tl.locator('[role="tab"]').filter(has_text=name_re)
                if await raw.count() < 1:
                    return ("FAIL", f"在本条分类 tablist 内未找到「{lab}」。")
                tloc = raw.first
                await _click_locator_robust(tloc, page, timeout_ms=8000)
                clicked.append(lab)
                await page.wait_for_timeout(450)
                n = await _eval_safe(
                    page,
                    "() => (document.body && document.body.innerText) ? document.body.innerText.length : 0",
                    timeout=8.0,
                )
                if int(n or 0) < 40:
                    return ("FAIL", f"点击「{lab}」后正文异常缩短，疑似白屏。")
            except Exception as e:
                return ("FAIL", f"切换「{lab}」失败：{_brief_exc(e)}")
        return ("PASS", "已依次点击（限定大厅分类 tablist）：" + " → ".join(clicked) + "。")
    except Exception as e:
        return ("FAIL", f"分类切换异常：{_brief_exc(e)}")


async def _run_p0_game_cards(page: Any) -> tuple[str, str]:
    try:
        imgs = page.locator("main img, [class*='card'] img, article img, a img")
        n = await imgs.count()
        if n < 2:
            imgs = page.locator("img")
            n = await imgs.count()
        if n < 2:
            return ("FAIL", f"首页可见 img 过少（{n}），卡片区可能未渲染。")
        broken = 0
        checked = 0
        for i in range(min(n, 12)):
            im = imgs.nth(i)
            try:
                await im.wait_for(state="attached", timeout=2000)
                nw = await im.evaluate("e => e.naturalWidth || 0")
                checked += 1
                complete = await im.evaluate("e => e.complete")
                if complete and int(nw or 0) == 0:
                    broken += 1
            except Exception:
                continue
        if broken > max(1, checked // 4):
            return ("FAIL", f"抽检 {checked} 张图，{broken} 张 complete 且 naturalWidth=0（疑似裂图）。")
        return ("PASS", f"卡片区 img 约 {n} 个；抽检 {checked} 张，典型裂图 {broken}。")
    except Exception as e:
        return ("FAIL", f"游戏卡片抽检异常：{_brief_exc(e)}")


async def _p0_lobby_seems_visible(page: Any) -> bool:
    """大厅：Play Now 或分类 tablist 等。"""
    try:
        if await page.get_by_role("button", name=re.compile(r"Play\s*Now", re.I)).count() >= 1:
            return True
        if await page.locator('[role="tablist"]').count() >= 1:
            return True
    except Exception:
        pass
    return False


async def _p0_wait_entered_game_shell(page: Any, *, timeout_ms: int = 55_000) -> tuple[bool, str]:
    """
    进局后 URL/首段正文常不变（壳内加载）；以 KK 浮标、Exit、Guest 抬头、加载百分比等判定已进入游戏。
    """
    deadline = time.monotonic() + timeout_ms / 1000.0
    guest = page.get_by_text(re.compile(r"Guest[_A-Z0-9]+", re.I))
    pct = page.get_by_text(re.compile(r"\d{1,3}\s*%"))
    while time.monotonic() < deadline:
        for hint, sel in (
            ("KK 浮标/菜单柄", 'img[class*="fab_handle_img"], img[src*="logo-ball"]'),
            ("Exit 图标", 'img[alt="Exit"]'),
            (
                "退出资源图",
                'img[src*="exit"][class*="content_item_icon"], img[src*="/assets/exit"]',
            ),
        ):
            loc = page.locator(sel)
            try:
                if await loc.count() >= 1:
                    if await loc.first.is_visible(timeout=400):
                        return True, f"检测到游戏壳：{hint}"
            except Exception:
                pass
        try:
            if await guest.count() >= 1 and await guest.first.is_visible(timeout=200):
                return True, "检测到局内 Guest 文案"
        except Exception:
            pass
        try:
            if await pct.count() >= 1 and await pct.first.is_visible(timeout=200):
                t = await pct.first.inner_text()
                if t and re.search(r"[6-9]\d\s*%|100\s*%", t):
                    return True, f"检测到高进度加载：{t.strip()[:20]}"
        except Exception:
            pass
        await page.wait_for_timeout(420)
    return False, f"{timeout_ms}ms 内未检测到游戏壳（KK/Exit/局内特征）"


async def _p0_exit_game_via_kk_then_exit(page: Any) -> tuple[bool, str]:
    """
    游戏内：先点右下角 KK 浮标（展开），再点 alt=Exit / 退出图（与产品 UI 一致）。
    """

    async def _try_click_exit() -> bool:
        candidates = (
            page.locator('img[alt="Exit"]'),
            page.locator('img[class*="content_item_icon"][src*="exit"]'),
            page.locator('img[src*="/assets/exit"]'),
        )
        for loc in candidates:
            try:
                if await loc.count() < 1:
                    continue
                await loc.first.wait_for(state="visible", timeout=2800)
                await loc.first.click(timeout=6000, force=True)
                return True
            except Exception:
                continue
        return False

    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "Exit 已可见，直接点击退出"

    kk = page.locator('img[class*="fab_handle_img"], img[src*="logo-ball"]').first
    try:
        await kk.wait_for(state="visible", timeout=15_000)
    except Exception as e:
        return False, f"未找到 KK 浮标：{_brief_exc(e)}"
    try:
        await kk.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    async def _tap_kk() -> None:
        try:
            await kk.click(timeout=6000, force=True)
        except Exception:
            await kk.evaluate("e => e.click()")

    await _tap_kk()
    await page.wait_for_timeout(650)

    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "已点 KK 展开菜单并点击 Exit"

    await _tap_kk()
    await page.wait_for_timeout(500)
    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "第二次点 KK 后出现 Exit 并已点击"

    return False, "展开菜单后仍未点到 Exit（img[alt=Exit] / 退出资源图）"


def _filter_console_errors(bucket: list[str]) -> list[str]:
    out: list[str] = []
    for x in bucket:
        if _is_benign_console_line(x):
            continue
        out.append(x)
    return out


async def _run_p0_console_clean(bucket: list[str]) -> tuple[str, str]:
    bad = _filter_console_errors(bucket)
    if bad:
        return ("FAIL", "Console error（过滤后仍剩）：" + " | ".join(bad[:5]))
    return ("PASS", "Console error 抽样无严重项（已过滤部分已知噪声）。")


async def _run_p0_release_checklist() -> tuple[str, str]:
    return ("SKIP", "本次更新点需对照发布说明人工勾选（脚本不内置变更清单）。")


async def _run_case(
    case_id: str,
    page: Any,
    *,
    log: Callable[[str], None],
    target_url: str,
    console_bucket: list[str],
) -> tuple[str, str]:
    if case_id == "p0_env_access":
        return ("PASS", "已在 _async_main 中完成 host/goto 校验。")
    if case_id == "p0_home_load":
        return await _run_p0_home_load(page)
    if case_id == "p0_page_title":
        return await _run_p0_page_title(page)
    if case_id == "p0_play_now":
        return await _run_p0_play_now(page)
    if case_id == "p0_category_tabs":
        return await _run_p0_category_tabs(page)
    if case_id == "p0_game_cards":
        return await _run_p0_game_cards(page)
    if case_id == "p0_console_clean":
        return await _run_p0_console_clean(console_bucket)
    if case_id == "p0_release_checklist":
        return await _run_p0_release_checklist()
    return ("BLOCKED", f"未知用例：{case_id}")


async def _async_main(args: argparse.Namespace) -> int:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("请先安装：pip install playwright && playwright install chromium", file=sys.stderr)
        return 2

    cdp = _kalaroko_cdp(args.cdp_http or None)
    target_url = (args.target_url or DEFAULT_TARGET).strip()
    host = _host_from_url(target_url)

    def log(msg: str) -> None:
        if args.verbose or not args.quiet:
            print(msg, flush=True)

    log("———————— K11 P0 · 平台冒烟（文档 31–39 行）————————")
    log(f"CDP：{cdp}  目标：{target_url}")
    log(f"用例（共 {len(CASE_DEFS)} 条）：{', '.join(c for c, _ in CASE_DEFS)}")
    log("")

    console_bucket: list[str] = []

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(cdp)
        navigate_if_no_tab = not getattr(args, "require_existing_tab", False)
        page, pick_err = await _acquire_cdp_target_page(
            browser,
            host=host,
            target_url=target_url,
            navigate_if_no_tab=navigate_if_no_tab,
            log=log,
        )
        if page is None:
            print(f"[失败] {pick_err or '无法获取目标页签'}", file=sys.stderr)
            return 2

        def _on_console(msg: Any) -> None:
            try:
                if msg.type == "error":
                    console_bucket.append(f"{msg.type}: {msg.text[:600]}")
            except Exception:
                pass

        page.on("console", _on_console)

        ok_env, env_detail = await _ensure_target_page(
            page,
            target_url,
            log=log,
            navigate_if_no_tab=navigate_if_no_tab,
            host=host,
        )
        if not ok_env:
            print(f"[失败] {env_detail}", file=sys.stderr)
            return 2

        log("准备：若在 party-hubs、app_tabbar=no 或个人中心，先回站点首页，保证分类条与卡片区稳定。")
        await _ensure_on_home_feed(page, target_url, log)

        results: list[dict[str, Any]] = []
        for i, (cid, title_zh) in enumerate(CASE_DEFS, start=1):
            log(f"【{i}/{len(CASE_DEFS)}】{title_zh}（{cid}）")
            if cid == "p0_env_access":
                v, detail = "PASS", env_detail
            else:
                v, detail = await _run_case(
                    cid,
                    page,
                    log=log,
                    target_url=target_url,
                    console_bucket=console_bucket,
                )
            vzh = VERDICT_ZH.get(v, v)
            log(f"  观察说明：{detail}")
            log(f"  结论：{vzh}（{v}）")
            log("")
            results.append(
                {
                    "case": cid,
                    "case_title_zh": title_zh,
                    "verdict": v,
                    "verdict_zh": vzh,
                    "detail": detail,
                }
            )

        bad_console = _filter_console_errors(console_bucket)
        if bad_console:
            log("———————— Console error（过滤后抽样，最多 15 条）————————")
            for s in bad_console[:15]:
                log("  · " + s[:400])
            log("")

        out = {
            "schema": "k11_p0_platform_smoke_playwright/v1",
            "ts_utc": datetime.now(timezone.utc).isoformat(),
            "cdp": cdp,
            "target_url": target_url,
            "page_url_final": page.url,
            "page_title_final": await page.title(),
            "console_errors_filtered_sample": bad_console[:30],
            "results": results,
        }
        if args.json_out:
            outp = Path(args.json_out)
            outp.parent.mkdir(parents=True, exist_ok=True)
            outp.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            log(f"JSON：{outp.resolve()}")

        if not args.no_xlsx_report:
            xlsx_p = (
                args.xlsx_report
                if args.xlsx_report is not None
                else _default_k11_xlsx_report_path()
            )
            log("")
            write_k11_p0_results_to_xlsx(Path(xlsx_p), results, log=log)

        log("———————— 汇总 ————————")
        for r in results:
            m = "✓" if r["verdict"] == "PASS" else ("○" if r["verdict"] == "SKIP" else "✗")
            log(f"  {m} [{r['case']}] {r['case_title_zh']} → {r['verdict_zh']}")

        verdicts = {r["verdict"] for r in results}
        if "FAIL" in verdicts or "BLOCKED" in verdicts:
            log("\n最终结果：存在未通过项，退出码 1。")
            return 1
        log("\n最终结果：无 FAIL，退出码 0。")
        return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="K11 P0 平台冒烟（Playwright CDP，文档 31–39 行）")
    ap.add_argument("--target-url", default=DEFAULT_TARGET, help="站点 URL（匹配标签页 host）")
    ap.add_argument("--cdp-http", default="", help="覆盖 KALAROKO_CDP_ENDPOINT")
    ap.add_argument(
        "--require-existing-tab",
        action="store_true",
        help="必须已有含目标域的页签；默认会在末页签自动 goto --target-url",
    )
    ap.add_argument(
        "--navigate-if-no-tab",
        action="store_true",
        help="已弃用：与默认行为相同（保留仅为兼容旧命令行）",
    )
    ap.add_argument("--json-out", type=Path, default=None)
    ap.add_argument(
        "--xlsx-report",
        type=Path,
        default=None,
        help="K11平台测试用例.xlsx；默认 K11_XLSX_REPORT 或 ~/Downloads/K11平台测试用例.xlsx",
    )
    ap.add_argument("--no-xlsx-report", action="store_true", help="不写入 Excel")
    ap.add_argument("--quiet", action="store_true")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()
    try:
        return asyncio.run(_async_main(args))
    except Exception as e:
        print(f"[失败] {type(e).__name__}: {e}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
