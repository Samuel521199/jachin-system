#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K11 P1 · 独立 Playwright 冒烟：可选读取 SKILL_P1_MODULES.md 中的 ### p1_* 顺序；缺省文件时用内置 P1 用例列表。连接已由 launch_chrome_debug.ps1 启动的 Chrome，检测 herontest.xin。

**不经过 L3**：仅用本机 Playwright CDP 附加，适合快速验证「调试 Chrome + 目标站」是否就绪。

前置：
  1) 运行 ``scripts/launch_chrome_debug.ps1``（默认 ``--remote-debugging-port=9222``），可选传入 URL：
       .\\scripts\\launch_chrome_debug.ps1 \"https://www.herontest.xin/\"
  2) 仓库根 ``.env`` 设置 ``KALAROKO_CDP_ENDPOINT=http://127.0.0.1:9222``（端口与上一步一致）。
  3) ``pip install playwright`` 且 ``playwright install chromium``（connect_over_cdp 仍需浏览器驱动）。

用法（仓库根）：
  python scripts/test_k11_p1_skill_herontest_playwright.py
  python scripts/test_k11_p1_skill_herontest_playwright.py -v
        # -v / --verbose：打印 [诊断] tablist 快照、Home/Party 各策略失败原因（定位底栏无 ARIA 等）
  python scripts/test_k11_p1_skill_herontest_playwright.py --target-url https://www.herontest.xin/
  python scripts/test_k11_p1_skill_herontest_playwright.py --cdp-http http://127.0.0.1:9222
  python scripts/test_k11_p1_skill_herontest_playwright.py --navigate-if-no-tab
        # 若无含目标域名的标签页，则对当前 context 执行一次 goto（调试用；正式验收请手工摆页）

运行结束可将结果写入 ``K11平台测试用例.xlsx``：``--xlsx-report`` / ``K11_XLSX_REPORT`` / ``~/Downloads/K11平台测试用例.xlsx``；``--no-xlsx-report`` 关闭。需 ``openpyxl``。

退出码：0 全部 PASS/SKIP；1 存在 FAIL；2 环境或 CDP 失败；3 未捕获异常。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
import time
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env", encoding="utf-8")
except ImportError:
    pass
except OSError:
    pass

DEFAULT_SKILL = ROOT / "skills_repo" / "k11-herontest-browser-qa" / "SKILL_P1_MODULES.md"
DEFAULT_TARGET = "https://www.herontest.xin/"


def _kalaroko_cdp(cli: str | None) -> str:
    raw = (cli or "").strip() or (os.environ.get("KALAROKO_CDP_ENDPOINT") or "").strip()
    if not raw:
        raw = "http://127.0.0.1:9222"
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "http://" + raw.lstrip("/")
    return raw.rstrip("/")


def _strip_frontmatter(md: str) -> str:
    if md.lstrip().startswith("---"):
        parts = md.split("---", 2)
        if len(parts) >= 3:
            return parts[2].lstrip("\n")
    return md


def _host_from_url(url: str) -> str:
    try:
        from urllib.parse import urlparse

        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def _home_feed_url(target: str) -> str:
    """与 --target-url 同站点的根路径（大厅首页），用于离开 /my/index 等个人中心。"""
    from urllib.parse import urlparse, urlunparse

    t = (target or DEFAULT_TARGET).strip() or DEFAULT_TARGET
    p = urlparse(t)
    if p.scheme and p.netloc:
        return urlunparse((p.scheme, p.netloc, "/", "", "", ""))
    return t.rstrip("/") + "/" if t else DEFAULT_TARGET


def _needs_goto_home_feed(current_url: str) -> bool:
    u = (current_url or "").lower()
    if "/my/" in u or "/me/" in u:
        return True
    if re.search(r"/(profile|account|wallet|settings)(/|$)", u):
        return True
    # Party 子站与隐藏底栏参数：不先回 / 则底栏用例、Hottest 区块均不可靠
    if "party-hubs" in u:
        return True
    if "app_tabbar=no" in u:
        return True
    return False


# KalaroKo 底栏：文案在 _item_label_* 内，Playwright 点 label 常超时；从父级 item 做 DOM click
_JS_CLICK_TAB_FROM_LABEL = """(el) => {
  let n = el.parentElement;
  for (let i = 0; i < 12 && n; i++) {
    const tag = (n.tagName || '').toUpperCase();
    const cls = (typeof n.className === 'string') ? n.className : '';
    const role = n.getAttribute && n.getAttribute('role');
    const isLabelOnly = /_item_label_/i.test(cls);
    if (!isLabelOnly && (tag === 'BUTTON' || tag === 'A' || role === 'button' || role === 'tab'
        || /\\bitem\\b/i.test(cls) || (cls.includes('_item_') && !cls.includes('_item_label_')))) {
      n.click();
      return { ok: true, step: i, tag, cls: cls.slice(0, 72) };
    }
    n = n.parentElement;
  }
  if (el.parentElement) {
    el.parentElement.click();
    return { ok: true, fallback: 'parent', tag: el.parentElement.tagName };
  }
  el.click();
  return { ok: true, fallback: 'label-only', tag: el.tagName };
}"""


async def _ensure_on_home_feed(
    page: Any, target_url: str, log: Callable[[str], None] | None
) -> None:
    """Profile 等用例会留在 /my/index；热门 Party 与底栏依赖大厅首页。"""
    home = _home_feed_url(target_url)
    try:
        cur = page.url or ""
    except Exception:
        cur = ""
    if not _needs_goto_home_feed(cur):
        if log:
            log(f"  [诊断] 无需回大厅：当前 {cur!r}")
        return
    if log:
        log(f"  [诊断] 离开个人中心等页：{cur!r} → goto {home!r}")
    await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(700)


def _load_skill(path: Path) -> str:
    text = path.read_text(encoding="utf-8-sig")
    return _strip_frontmatter(text)


def _parse_p1_ids(skill_body: str) -> list[str]:
    return re.findall(r"^###\s+(p1_\w+)\s", skill_body, re.MULTILINE)


# 控制台输出用中文名（key 仍为 SKILL 里的 id）
CASE_TITLE_ZH: dict[str, str] = {
    "p1_customer_service": "客服入口",
    "p1_share_tab": "分享页签",
    "p1_task_tab": "任务页签",
    "p1_profile_tab": "我的/Profile",
    "p1_hottest_parties": "热门 Party 板块",
    "p1_party_status": "Party 状态展示",
}

VERDICT_ZH: dict[str, str] = {
    "PASS": "通过",
    "FAIL": "未通过",
    "SKIP": "跳过（本页无相关文案，不算失败）",
    "BLOCKED": "阻塞",
}

P1_CASE_TO_XLSX_TEST_ITEM_KEY: dict[str, str] = {
    "p1_customer_service": "客服",
    "p1_share_tab": "分享",
    "p1_task_tab": "任务",
    "p1_profile_tab": "我的",
    "p1_hottest_parties": "热门",
    "p1_party_status": "Party 状态",
}

_XLSX_REMARK_MAX_LEN = 32000


def _default_k11_xlsx_report_path() -> Path:
    env = (os.environ.get("K11_XLSX_REPORT") or "").strip()
    if env:
        return Path(env)
    return Path.home() / "Downloads" / "K11平台测试用例.xlsx"


def _find_smoke_sheet_header(ws: Any) -> tuple[int, int, int, int] | None:
    max_r = min(int(ws.max_row or 1), 45)
    max_c = min(int(ws.max_column or 1), 40)
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = str(v).strip() if v is not None else ""
            if s and s not in col_map:
                col_map[s] = c
        if "结果" not in col_map or "备注" not in col_map:
            continue
        item_col = None
        for k in ("测试项目", "测试项", "用例名称"):
            if k in col_map:
                item_col = col_map[k]
                break
        if item_col is None:
            continue
        return r, item_col, col_map["结果"], col_map["备注"]
    return None


def write_k11_p1_results_to_xlsx(
    xlsx_path: Path,
    results: list[dict[str, Any]],
    *,
    log: Callable[[str], None],
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [xlsx] 未安装 openpyxl，跳过写入（可：pip install openpyxl）")
        return

    if not xlsx_path.is_file():
        log(f"  [xlsx] 文件不存在，跳过：{xlsx_path}")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        log(f"  [xlsx] 打开工作簿失败：{e!s}")
        return

    ordered_names: list[str] = [n for n in wb.sheetnames if "冒烟" in n]
    ordered_names.extend(n for n in wb.sheetnames if n not in ordered_names)
    verdict_cell = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP", "BLOCKED": "BLOCKED"}

    try:
        for sn in ordered_names:
            ws = wb[sn]
            parsed = _find_smoke_sheet_header(ws)
            if not parsed:
                continue
            hdr, c_item, c_res, c_note = parsed
            last_r = max(int(ws.max_row or hdr), hdr + 5, 80)
            sheet_wrote = 0
            for resrow in results:
                cid = str(resrow.get("case") or "")
                key = P1_CASE_TO_XLSX_TEST_ITEM_KEY.get(cid)
                if not key:
                    continue
                v = str(resrow.get("verdict") or "")
                detail = str(resrow.get("detail") or "")
                cell_v = verdict_cell.get(v, v)
                remark = (detail or "")[:_XLSX_REMARK_MAX_LEN]
                matched = False
                for r in range(hdr + 1, last_r + 1):
                    raw = ws.cell(row=r, column=c_item).value
                    text = str(raw).strip() if raw is not None else ""
                    if not text:
                        continue
                    if key in text:
                        ws.cell(row=r, column=c_res, value=cell_v)
                        ws.cell(row=r, column=c_note, value=remark)
                        sheet_wrote += 1
                        matched = True
                        break
                if not matched:
                    log(f"  [xlsx] 未匹配行：case={cid} 关键字={key!r} sheet={sn!r}")
            if sheet_wrote:
                try:
                    wb.save(xlsx_path)
                except PermissionError:
                    log(
                        f"  [xlsx] 保存被拒绝（是否正用 Excel 打开该文件？）：{xlsx_path.resolve()}"
                    )
                    return
                log(
                    f"  [xlsx] 已在工作表「{sn}」更新 {sheet_wrote} 行 → {xlsx_path.resolve()}"
                )
                return

        log(
            "  [xlsx] 未找到含「测试项目/结果/备注」表头的工作表，或未更新任何行。"
        )
    except Exception as e:
        log(f"  [xlsx] 写入过程异常：{e!s}")


def _brief_exc(e: BaseException, limit: int = 200) -> str:
    return f"{type(e).__name__}: {str(e).strip()[:limit]}"


async def _try_click_bottom_label_js(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    tag: str = "底栏·JS",
) -> tuple[bool, str]:
    """KalaroKo 底栏：对精确文案 .last 做 JS 父级 click（与 Playwright 点 div 标签层分离）。"""
    def lg(m: str) -> None:
        if log:
            log(f"  [诊断·{tag}] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = (fr.url or "")[:120]
        except Exception:
            pass
        try:
            if "about:blank" in (fr.url or "").lower() and fi > 0:
                continue
        except Exception:
            pass
        try:
            n = await fr.get_by_text(label, exact=True).count()
            if n < 1:
                continue
            loc = fr.get_by_text(label, exact=True).last
            await loc.wait_for(state="attached", timeout=4000)
            try:
                await loc.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            info = await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
            lg(f"frame[{fi}]「{label}」.last JS → {info} | {furl!r}")
            await page.wait_for_timeout(450)
            return True, f"frame[{fi}] JS 底栏「{label}」"
        except Exception as e:
            lg(f"frame[{fi}]「{label}」：{_brief_exc(e, 160)}")
    return False, ""


async def _log_tablist_snapshot(page: Any, log: Callable[[str], None], *, tag: str = "") -> None:
    """打印当前页 URL、frame 数、每个 tablist 的 tab 数量与前几项文案（便于区分底栏与筛选条）。"""
    prefix = f"  [诊断{tag}] " if tag else "  [诊断] "
    try:
        log(f"{prefix}page.url = {page.url}")
    except Exception:
        log(f"{prefix}page.url = （无法读取）")
    try:
        frames = page.frames
        log(f"{prefix}frames 数量 = {len(frames)}（含主文档与 iframe）")
    except Exception as e:
        log(f"{prefix}frames = （{_brief_exc(e)}）")
    try:
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        log(f'{prefix}[role="tablist"] 数量 = {nl}')
        for idx in range(nl):
            tl = lists.nth(idx)
            tabs = tl.locator('[role="tab"]')
            tc = await tabs.count()
            parts: list[str] = []
            for j in range(min(tc, 8)):
                try:
                    raw = (await tabs.nth(j).inner_text() or "").strip()
                    one = " ".join(raw.split())[:48]
                    if len(raw) > 48:
                        one += "…"
                    parts.append(f"[{j}]「{one}」")
                except Exception as e:
                    parts.append(f"[{j}]（读文案失败 {_brief_exc(e, 80)}）")
            log(f"{prefix}  tablist[{idx}] → {tc} 个 tab: " + ("；".join(parts) if parts else "（无子 tab）"))
    except Exception as e:
        log(f"{prefix}枚举 tablist 失败：{_brief_exc(e)}")

    for role_name, label in (
        ("tab", "Home"),
        ("tab", "Party"),
    ):
        try:
            n = await page.get_by_role(role_name, name=re.compile(rf"^{re.escape(label)}$", re.I)).count()
            log(f'{prefix}get_by_role({role_name}, name=^{label}$) 匹配数 = {n}')
        except Exception as e:
            log(f"{prefix}get_by_role 统计 {label} 失败：{_brief_exc(e, 120)}")

    try:
        n_link = await page.get_by_role("link", name=re.compile(r"Party\s*Hubs", re.I)).count()
        log(f"{prefix}get_by_role(link, Party Hubs…) 匹配数 = {n_link}")
    except Exception as e:
        log(f"{prefix}统计 Party Hubs 链接失败：{_brief_exc(e, 120)}")


async def _visible_any(
    page: Any, patterns: list[str], *, timeout_ms: float = 2500
) -> tuple[bool, str]:
    """任一 Playwright get_by_text(re) 可见即 True。"""
    for pat in patterns:
        try:
            loc = page.get_by_text(re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=int(timeout_ms))
            return True, f"页面上有可见文字，匹配规则 /{pat}/i"
        except Exception:
            continue
    return False, (
        "页面上未找到与下列文案匹配的可见元素："
        + "、".join(patterns[:8])
        + "（若站点用语不同，需在脚本里补充关键词）"
    )


# 纯图标客服（无文案）：img 文件名 / class / title 等常见命名（KalaroKo 右上角耳机按钮等）
_ICON_MARKERS_SRC_ALT = (
    "headset",
    "Headset",
    "headphone",
    "earphone",
    "livechat",
    "LiveChat",
    "live-chat",
    "kefu",
    "customer-service",
    "CustomerService",
    "customer_service",
    "help-center",
    "HelpCenter",
)
_ICON_CLASS_MARKERS = (
    "headset",
    "Headset",
    "headphone",
    "customer-service",
    "CustomerService",
    "customer_service",
    "livechat",
    "LiveChat",
    "kefu",
    "Kefu",
    "kf-",
)


def _css_pure_icon_customer_service() -> str:
    parts: list[str] = []
    for m in _ICON_MARKERS_SRC_ALT:
        parts.append(f"img[src*='{m}']")
        parts.append(f"img[srcset*='{m}']")
        parts.append(f"img[alt*='{m}']")
        parts.append(f"button:has(img[src*='{m}'])")
        parts.append(f"[role='button']:has(img[src*='{m}'])")
        parts.append(f"div[role='button']:has(img[src*='{m}'])")
        parts.append(f"a:has(img[src*='{m}'])")
    for m in _ICON_CLASS_MARKERS:
        parts.append(f"button[class*='{m}']")
        parts.append(f"[role='button'][class*='{m}']")
        parts.append(f"a[class*='{m}']")
        parts.append(f"button:has(img[class*='{m}'])")
    parts.append("img[alt*='客服']")
    parts.extend(
        [
            "button[title*='客服']",
            "[role='button'][title*='客服']",
            "button[title*='Headset']",
            "button[title*='headset']",
            "[role='button'][title*='Headset']",
            "[role='button'][title*='headset']",
            "button[title*='Customer']",
            "button[title*='customer']",
            "[role='button'][title*='Customer']",
            "[role='button'][title*='customer']",
            "button[title*='Service']",
            "button[title*='service']",
            "[role='button'][title*='Service']",
            "[role='button'][title*='service']",
        ]
    )
    return ", ".join(parts)


async def _customer_service_on_frame(frame: Any, *, in_subframe: bool) -> tuple[bool, str]:
    """在单个 Page 或 Frame 内检测客服入口。"""
    suffix = "（检测于页面内嵌 iframe）" if in_subframe else ""

    text_patterns = [
        r"Customer\s*Service",
        r"客服",
        r"Support",
        r"Help\s*Center",
        r"Contact(\s+Us)?",
        r"Live\s*Chat",
        r"在线客服",
        r"联系客服",
        r"联系我们",
        r"Feedback",
        r"帮助",
        r"Help",
        r"技术支持",
        r"售后服务",
        r"Assist",
        r"Inquiry",
    ]
    ok, msg = await _visible_any(frame, text_patterns, timeout_ms=1800)
    if ok:
        return True, f"可见文案命中：{msg}{suffix}"

    name_pat = re.compile(
        r"service|support|help|客服|contact|chat|feedback|assist|inquiry|联系我们",
        re.I,
    )
    for role in ("button", "link"):
        try:
            loc = frame.get_by_role(role, name=name_pat).first
            await loc.wait_for(state="visible", timeout=1200)
            return True, f"找到可见的「{role}」，名称含客服相关关键词{suffix}"
        except Exception:
            pass

    try:
        loc = frame.get_by_alt_text(
            re.compile(
                r"headset|headphone|customer\s*service|support|help|客服|在线|联系|售后",
                re.I,
            )
        ).first
        await loc.wait_for(state="visible", timeout=1200)
        return True, f"通过 img 的 alt 文案命中疑似客服图标{suffix}"
    except Exception:
        pass

    attr_sel = (
        "a[href*='support'], a[href*='help'], a[href*='chat'], a[href*='customer'], "
        "a[href*='service'], a[href*='feedback'], a[href*='contact'], "
        "[role='button'][aria-label*='service'], [role='button'][aria-label*='support'], "
        "[role='button'][aria-label*='help'], [role='button'][aria-label*='chat'], "
        "[role='button'][aria-label*='客服'], [aria-label*='客服'], "
        "[data-testid*='support'], [data-testid*='service'], [data-testid*='help'], "
        "[data-testid*='chat'], [data-testid*='customer']"
    )
    try:
        loc = frame.locator(attr_sel).first
        await loc.wait_for(state="visible", timeout=1800)
        return True, f"通过链接 href 或 aria-label / data-testid 命中疑似客服入口{suffix}"
    except Exception:
        pass

    try:
        flo = frame.locator(
            "[class*='float'][class*='service'], [class*='float'][class*='support'], "
            "[id*='chat-widget'], [id*='customer-service']"
        ).first
        await flo.wait_for(state="visible", timeout=600)
        return True, f"命中浮动区/挂件选择器（疑似在线客服）{suffix}"
    except Exception:
        pass

    try:
        loc = frame.locator(_css_pure_icon_customer_service()).first
        await loc.wait_for(state="visible", timeout=2200)
        return True, f"命中纯图标客服入口（img src/alt、button:has(img) 或 class/title 特征）{suffix}"
    except Exception:
        pass

    try:
        hs = frame.locator("button:has(svg), [role='button']:has(svg)").filter(
            has=frame.locator(
                "[class*='headset'], [class*='Headset'], [class*='headphone'], "
                "[class*='customer'], [class*='service'], [class*='support'], [class*='kefu']"
            )
        )
        loc = hs.first
        await loc.wait_for(state="visible", timeout=1000)
        return True, f"命中带 SVG 的按钮且子节点 class 含客服相关特征{suffix}"
    except Exception:
        pass

    return False, ""


async def _customer_service_detect(page: Any) -> tuple[bool, str]:
    """
    客服入口：主文档 + 各 iframe 依次检测（文案 → 角色 → img alt →
    href/aria-label/data-testid → 纯图标/img 包裹按钮 → SVG 按钮）。
    """
    frames = list(page.frames)
    main = page.main_frame
    order = [main] + [f for f in frames if f != main]
    for fr in order:
        sub = fr != main
        ok, msg = await _customer_service_on_frame(fr, in_subframe=sub)
        if ok:
            return True, msg
    return (
        False,
        "主页面与所有 iframe 内均未找到客服入口（文案、无障碍名称、href/aria、"
        "纯图标 img/src/class、SVG 按钮等已轮询）。若仍失败，请在开发者工具中查看该节点 DOM 后发我补充选择器。",
    )


# —— P1 客服：仅主文档顶栏/视口右上打开，全 frame 关 Garden（含 about:blank 子帧）——

_P1_CS_IMG_CLICK_JS = """(el) => {
  const t = el.closest('button, a, [role="button"], [onclick]') || el.parentElement;
  if (t && t !== el) { t.click(); return { via: 'closest' }; }
  el.click();
  return { via: 'self' };
}"""


async def _p1_click_customer_service_top_right(
    page: Any, log: Callable[[str], None] | None
) -> tuple[bool, str]:
    """首页主文档：header 内耳机图 → 视口右上几何命中 → 全文 alt 兜底（不点 iframe 内）。"""
    fr = page.main_frame
    alt_re = re.compile(
        r"headset|headphone|customer|service|support|help|客服|在线|联系|售后",
        re.I,
    )
    try:
        hdr = fr.locator("header").first
        if await hdr.count() > 0:
            im = hdr.get_by_alt_text(alt_re).first
            if await im.count() > 0:
                await im.wait_for(state="attached", timeout=3200)
                try:
                    await im.scroll_into_view_if_needed()
                except Exception:
                    pass
                try:
                    await im.click(timeout=4000, force=True)
                except Exception:
                    await im.evaluate(_P1_CS_IMG_CLICK_JS)
                return True, "已点击 header 内客服图标（img alt）"
    except Exception as e:
        if log:
            log(f"  [诊断·客服] header+alt：{_brief_exc(e, 100)}")

    try:
        clicked = await fr.evaluate("""() => {
          const W = window.innerWidth, H = window.innerHeight;
          const imgs = [];
          document.querySelectorAll('header img').forEach(i => imgs.push(i));
          document.querySelectorAll('[class*="Header"] img, [class*="header"] img').forEach(i => {
            if (!imgs.includes(i)) imgs.push(i);
          });
          for (const img of imgs) {
            const r = img.getBoundingClientRect();
            if (r.width < 6 || r.height < 6) continue;
            const cx = r.left + r.width / 2, cy = r.top + r.height / 2;
            if (cx > W * 0.46 && cy < H * 0.45) {
              const t = img.closest('button, a, [role="button"]');
              (t || img).click();
              return true;
            }
          }
          return false;
        }""")
        if clicked:
            return True, "已点击视口右上 header 区域图标（几何命中）"
    except Exception as e:
        if log:
            log(f"  [诊断·客服] 几何命中：{_brief_exc(e, 100)}")

    try:
        loc = fr.get_by_alt_text(alt_re).first
        if await loc.count() > 0:
            await loc.wait_for(state="attached", timeout=2800)
            try:
                await loc.click(timeout=4000, force=True)
            except Exception:
                await loc.evaluate(_P1_CS_IMG_CLICK_JS)
            return True, "已点击主文档客服图标（全文 alt 兜底）"
    except Exception:
        pass
    return False, ""


async def _p1_frame_has_zendesk_widget(page: Any) -> bool:
    """任意 frame 出现 Garden 收起图标或典型聊天 UI 即视为已加载。"""
    for fr in page.frames:
        try:
            if await fr.locator('svg[data-garden-id="buttons.icon"]').count() > 0:
                return True
        except Exception:
            pass
        for pat in (
            r"Type\s*a\s*message",
            r"Kalaro\s*Bot",
            r"Privacy\s*Notice",
            r"Zendesk",
        ):
            try:
                if await fr.get_by_text(re.compile(pat, re.I)).count() > 0:
                    return True
            except Exception:
                pass
    return False


async def _p1_click_zendesk_garden_close(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """
    点击 Zendesk Garden 向下收起：svg[data-garden-id="buttons.icon"]。
    不跳过 about:blank（Messenger 子帧常见）；每轮遍历当前所有 frame。
    """
    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:140]
        except Exception:
            pass
        try:
            btn = fr.locator(
                'button:has(svg[data-garden-id="buttons.icon"]), '
                '[role="button"]:has(svg[data-garden-id="buttons.icon"])'
            ).last
            if await btn.count() > 0:
                await btn.click(timeout=2800, force=True)
                if log:
                    log(f"  [诊断·客服] 已点 Garden 收起 button frame[{fi}] {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception:
            pass
        try:
            svg = fr.locator('svg[data-garden-id="buttons.icon"]').last
            if await svg.count() > 0:
                await svg.evaluate(
                    """(el) => {
                      const b = el.closest('button, [role="button"]');
                      (b || el.parentElement || el).click();
                    }"""
                )
                if log:
                    log(f"  [诊断·客服] 已点 Garden svg→父级 frame[{fi}] {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception:
            pass
    return False


async def _p1_try_zendesk_header_last_in_chat_frames(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """在已出现聊天特征的 frame 内点 header 最后一颗按钮（⋮/收起旁）。"""
    for fi, fr in enumerate(page.frames):
        try:
            in_chat = False
            if await fr.locator('svg[data-garden-id="buttons.icon"]').count() > 0:
                in_chat = True
            if not in_chat and await fr.get_by_text(
                re.compile(r"Type\s*a\s*message", re.I)
            ).count() > 0:
                in_chat = True
            if not in_chat and await fr.get_by_text(
                re.compile(r"Kalaro\s*Bot", re.I)
            ).count() > 0:
                in_chat = True
            if not in_chat:
                continue
            hdr = fr.locator("[class*='header'], header").first
            if await hdr.count() < 1:
                continue
            btns = hdr.locator("button, [role='button']")
            bn = await btns.count()
            if bn < 1:
                continue
            await btns.nth(bn - 1).click(timeout=2600, force=True)
            if log:
                log(f"  [诊断·客服] 已点聊天窗顶栏最后按钮 frame[{fi}]")
            await page.wait_for_timeout(350)
            return True
        except Exception:
            continue
    return False


async def _try_spa_header_back(page: Any) -> bool:
    """SPA 顶栏返回（图2 红框）：主文档 header 内首颗按钮或 history.back。"""
    fr = page.main_frame
    locators: list[Any] = [
        fr.locator("header").locator("button, [role='button'], a").first,
        fr.locator("[class*='Header']").locator("button, [role='button']").first,
        fr.locator("[class*='header']").locator("button, [role='button']").first,
        fr.locator("[class*='navbar']").locator("button, [role='button']").first,
        fr.locator("[class*='NavBar']").locator("button, [role='button']").first,
        fr.get_by_role("button", name=re.compile(r"back", re.I)).first,
        fr.locator("[aria-label*='返回'], [aria-label*='back' i]").first,
    ]
    for loc in locators:
        try:
            if await loc.count() < 1:
                continue
            await loc.wait_for(state="attached", timeout=1800)
            await loc.click(timeout=4500, force=True)
            await page.wait_for_timeout(450)
            return True
        except Exception:
            continue
    try:
        ok = await fr.evaluate("""() => {
          const sels = ['header button', '[class*="Header"] button', '[class*="header"] button',
            '[class*="navbar"] button', '[class*="NavBar"] button', '[class*="title-bar"] button'];
          for (const s of sels) {
            const el = document.querySelector(s);
            if (el) { el.click(); return true; }
          }
          return false;
        }""")
        await page.wait_for_timeout(450)
        if ok:
            return True
    except Exception:
        pass
    try:
        await page.go_back(wait_until="domcontentloaded", timeout=15_000)
        await page.wait_for_timeout(450)
        return True
    except Exception:
        return False


async def _leave_party_hubs_to_home(
    page: Any,
    target_url: str,
    log: Callable[[str], None] | None,
) -> None:
    """离开 Party Hubs 子页，恢复底栏（图2 顶栏返回，失败则 goto /）。"""
    u = (page.url or "").lower()
    if "party-hubs" not in u:
        return
    if log:
        log("  [诊断] 当前在 Party Hubs，顶栏返回大厅…")
    await _try_spa_header_back(page)
    u2 = (page.url or "").lower()
    if "party-hubs" in u2:
        if log:
            log(f"  [诊断] 仍停留在 party-hubs（{page.url!r}），goto 站点根路径")
        home = _home_feed_url(target_url)
        await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(500)


# Party Hubs 子页：不滚底，仅用顶栏与列表区文案做轻量校验
_P1_PARTY_HUBS_VERIFY_PATTERNS = [
    r"Follow",
    r"All",
    r"Preparing",
    r"准备",
    r"In\s*Game",
    r"Random\s*Match",
    r"Create\s*a\s*Party",
    r"Party",
    r"Guest",
]


async def _try_party_hubs_top_bar_back(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """
    Party Hubs 顶栏左侧返回：KalaroKo 为 img._top_bar_item_icon_*（内嵌左箭头 SVG）。
    """
    def lg(m: str) -> None:
        if log:
            log("  [诊断·HottestParty] " + m)

    fr = page.main_frame
    sels = (
        'img[class*="_top_bar_item_icon_"]',
        'img[class*="top_bar_item_icon"]',
        r'img[src*="15.0005"]',
        r'img[src*="M15.0005"]',
    )
    for sel in sels:
        try:
            loc = fr.locator(sel).first
            if await loc.count() < 1:
                continue
            await loc.wait_for(state="attached", timeout=2200)
            try:
                await loc.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            try:
                await loc.click(timeout=3500, force=True)
            except Exception:
                await loc.evaluate(
                    """(el) => {
                      const p = el.closest('button, a, [role="button"]');
                      (p || el.parentElement || el).click();
                    }"""
                )
            lg(f"已点击顶栏返回图标（{sel!r}）")
            await page.wait_for_timeout(450)
            return True
        except Exception as e:
            lg(f"顶栏返回 {sel!r}：{_brief_exc(e, 100)}")
    return False


async def _p1_leave_party_hubs_after_hottest(
    page: Any,
    target_url: str,
    log: Callable[[str], None] | None,
) -> None:
    """先点顶栏 img 返回，仍在 party-hubs 则走通用 back/goto。"""
    if "party-hubs" not in (page.url or "").lower():
        return
    await _try_party_hubs_top_bar_back(page, log)
    if "party-hubs" in (page.url or "").lower():
        await _leave_party_hubs_to_home(page, target_url, log)


async def _customer_service_click_open_close(
    page: Any, log: Callable[[str], None] | None
) -> tuple[str, str]:
    """
    P1 客服：主文档顶栏/右上打开 → 轮询全部 frame（含 about:blank）点
    svg[data-garden-id="buttons.icon"] 收起 → 兜底顶栏最后键 → Esc。
    """
    ok, cm = await _p1_click_customer_service_top_right(page, log)
    if not ok:
        det = await _customer_service_detect(page)
        return (
            "FAIL",
            "未能点击首页客服入口。"
            + (f"（页面上曾可见：{det[1]}）" if det[0] else f" {det[1]}"),
        )
    if log:
        log(f"  [诊断·客服] {cm}")
    await page.wait_for_timeout(500)

    saw_widget = False
    closed_ok = False
    deadline = time.monotonic() + 16.0
    while time.monotonic() < deadline:
        saw_widget = saw_widget or await _p1_frame_has_zendesk_widget(page)
        if await _p1_click_zendesk_garden_close(page, log):
            closed_ok = True
            break
        await asyncio.sleep(0.16)

    if not closed_ok:
        if log:
            log("  [诊断·客服] 轮询内未点到 Garden，尝试聊天窗 header 最后按钮…")
        closed_ok = await _p1_try_zendesk_header_last_in_chat_frames(page, log)

    if not closed_ok:
        for _ in range(7):
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(160)
            except Exception:
                break
        if saw_widget:
            closed_ok = True

    if closed_ok:
        load_note = "已检测到聊天窗特征" if saw_widget else "关闭动作已执行"
        return (
            "PASS",
            f"{cm}；{load_note}；已通过 Garden svg / 顶栏按钮 / Esc 收起。继续后续用例。",
        )
    return (
        "FAIL",
        f"{cm}；未检测到聊天窗或未点中 "
        f'svg[data-garden-id="buttons.icon"]（当前 frame 数 {len(page.frames)}）。',
    )


async def _try_click_tab(
    page: Any, patterns: list[str], *, timeout_ms: float = 4000
) -> tuple[bool, str]:
    for pat in patterns:
        try:
            loc = page.get_by_role("tab", name=re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=2000)
            await loc.click(timeout=int(timeout_ms))
            await page.wait_for_timeout(800)
            return True, f"已点击「页签」角色，名称匹配 /{pat}/i"
        except Exception:
            pass
        try:
            loc = page.locator("a,button,[role='tab']").filter(has_text=re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=2000)
            await loc.click(timeout=int(timeout_ms))
            await page.wait_for_timeout(800)
            return True, f"已点击导航控件，文案匹配 /{pat}/i"
        except Exception:
            pass
    return False, "未点到对应页签，尝试过：" + "、".join(patterns)


async def _try_click_visible_text(
    page: Any, patterns: list[str], *, timeout_ms: float = 2000, settle_ms: float = 700
) -> tuple[bool, str]:
    """对可见文案执行 scroll + click（用于分区标题、非标准 tab 等）。"""
    for pat in patterns:
        try:
            loc = page.get_by_text(re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=int(timeout_ms))
            await loc.scroll_into_view_if_needed()
            await loc.click(timeout=4500)
            await page.wait_for_timeout(int(settle_ms))
            return True, f"已点击匹配 /{pat}/i 的可见文案（scroll+click）"
        except Exception:
            continue
    return False, "未点击到任何匹配文案：" + "、".join(patterns[:8])


async def _try_click_exact_text_last_in_frames(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    diag_tag: str,
) -> tuple[bool, str]:
    """
    无 ARIA tablist 的 H5 底栏：各 frame 内精确匹配文案，取 .last（DOM 中固定底栏往往在后）。
    """
    def lg(m: str) -> None:
        if log:
            log(f"  [诊断·{diag_tag}] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = fr.url or ""
        except Exception:
            pass
        try:
            n = await fr.get_by_text(label, exact=True).count()
        except Exception as e:
            lg(f"frame[{fi}] {furl!r} 统计「{label}」失败：{_brief_exc(e, 120)}")
            continue
        if n < 1:
            continue
        try:
            lg(f"frame[{fi}] {furl!r} 精确「{label}」×{n} → .last + JS 父级 click")
            loc = fr.get_by_text(label, exact=True).last
            await loc.wait_for(state="attached", timeout=3200)
            try:
                await loc.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
            await page.wait_for_timeout(700)
            return True, f"frame[{fi}] 精确「{label}」.last（JS）"
        except Exception as e:
            lg(f"frame[{fi}] 点击失败：{_brief_exc(e, 200)}")
    return False, ""


async def _try_click_party_in_scoped_bottom_bar(
    page: Any, *, log: Callable[[str], None] | None
) -> tuple[bool, str]:
    """底栏容器同时含 Home + Profile（或 Task）时，在其中点「Party」，避免点到筛选条。"""
    def lg(m: str) -> None:
        if log:
            log("  [诊断·Party底栏·容器] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = fr.url or ""
        except Exception:
            pass
        for shell_sel in (
            "nav",
            "footer",
            "[class*='tabbar' i]",
            "[class*='TabBar']",
            "[class*='tab-bar' i]",
            "[class*='bottomNav' i]",
            "[class*='BottomNav' i]",
        ):
            try:
                shell = fr.locator(shell_sel).filter(
                    has=fr.get_by_text("Home", exact=True)
                )
                n = await shell.count()
                if n < 1:
                    continue
                bar = None
                for anchor in ("Profile", "Task", "Share"):
                    cand = shell.filter(has=fr.get_by_text(anchor, exact=True)).last
                    if await cand.count() > 0:
                        bar = cand
                        break
                if bar is None:
                    bar = shell.last
                p = bar.get_by_text("Party", exact=True).first
                await p.wait_for(state="attached", timeout=2800)
                try:
                    await p.scroll_into_view_if_needed(timeout=2800)
                except Exception:
                    pass
                await p.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                await page.wait_for_timeout(800)
                return True, f"frame[{fi}] {shell_sel} 容器内「Party」（{furl!r}）"
            except Exception as e:
                lg(f"frame[{fi}] {shell_sel}：{_brief_exc(e, 160)}")
                continue
    return False, ""


async def _best_bottom_tabs(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[Any, int] | None:
    """
    取「主导航」tablist：优先 tab 数量最多的那一组；并列时取 DOM 中下标最大者（一般为底部固定栏）。
    用于区分顶/中部的筛选 tablist（项数往往较少）与底栏 Home/Party/Share…。
    """
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·主导航] " + msg)

    try:
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        if nl == 0:
            lg('页面中无 [role="tablist"]（底栏可能用 div+click 实现，无 ARIA tablist）')
            return None
        candidates: list[tuple[int, Any, int]] = []
        for idx in range(nl):
            tl = lists.nth(idx)
            nh = await tl.get_by_text("Home", exact=True).count()
            nshare = await tl.get_by_text("Share", exact=True).count()
            if nh < 1 or nshare < 1:
                lg(
                    f"tablist[{idx}] 跳过（非底栏：须同时含精确「Home」与「Share」；"
                    f"当前 Home={nh} Share={nshare}）"
                )
                continue
            tabs = tl.locator('[role="tab"]')
            tc = await tabs.count()
            if tc >= 2:
                candidates.append((idx, tabs, tc))
            lg(f"tablist[{idx}] 内 [role=tab] 数量 = {tc}（已确认含 Home+Share）")
        if not candidates:
            lg("无「同时含 Home+Share」的 tablist（中部筛选条已排除）；改用 JS 底栏或 nav 容器策略")
            return None
        max_tc = max(c[2] for c in candidates)
        best = max((c for c in candidates if c[2] == max_tc), key=lambda c: c[0])
        lg(f"选用 tab 数最多的一组：共 {best[2]} 项（并列时取 DOM 下标较大者 ≈ 底栏）")
        return (best[1], best[2])
    except Exception as e:
        lg(f"_best_bottom_tabs 异常：{_brief_exc(e)}")
        return None


async def _try_click_bottom_home_nav(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """底栏第一项 Home（须先回首页再滚到底才见 Hottest Parties）。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·Home] " + msg)

    reasons: list[str] = []
    lg("策略 0：各 frame 精确「Home」.last + JS 父级点击（KalaroKo）…")
    ok0, m0 = await _try_click_bottom_label_js(page, "Home", log=log, tag="Home")
    if ok0:
        return True, m0

    r = await _best_bottom_tabs(page, log=log)
    if r:
        tabs, tc = r
        try:
            lg(f"策略 A：点击主导航第 1 个 tab（共 {tc} 项）…")
            t0 = tabs.nth(0)
            await t0.wait_for(state="visible", timeout=3200)
            await t0.scroll_into_view_if_needed()
            await t0.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, f"已点击主导航 tablist 第 1 项（共 {tc} 项，假定 Home）"
        except Exception as e:
            reasons.append(f"策略A失败：{_brief_exc(e)}")
            lg(reasons[-1])
    else:
        reasons.append("无可用主导航 tablist（见上方 [诊断·主导航]）")

    pat = re.compile(r"^Home$", re.I)
    try:
        lg("策略 B：get_by_role(tab, name=^Home$)…")
        tabs = page.get_by_role("tab", name=pat)
        n = await tabs.count()
        lg(f"匹配数 = {n}")
        if n > 0:
            t = tabs.nth(n - 1)
            await t.wait_for(state="visible", timeout=2500)
            await t.scroll_into_view_if_needed()
            await t.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, f"已点击「Home」页签（role=tab，第 {n} 个匹配取末项）"
    except Exception as e:
        reasons.append(f"策略B失败：{_brief_exc(e)}")
        lg(reasons[-1])

    try:
        lg("策略 C：a/button/[role=tab] 文案 ^Home$…")
        row = page.locator("a,button,[role='tab']").filter(has_text=pat)
        n = await row.count()
        lg(f"匹配数 = {n}")
        if n > 0:
            b = row.nth(n - 1)
            await b.wait_for(state="visible", timeout=2500)
            await b.scroll_into_view_if_needed()
            await b.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, "已点击底部导航「Home」（a/button/tab，取末项）"
    except Exception as e:
        reasons.append(f"策略C失败：{_brief_exc(e)}")
        lg(reasons[-1])

    lg("策略 D：各 frame 精确「Home」.last（无 tablist 的底栏）…")
    ok_d, msg_d = await _try_click_exact_text_last_in_frames(
        page, "Home", log=log, diag_tag="Home·文案"
    )
    if ok_d:
        return True, msg_d + "（精确文案 .last）"

    summary = "；".join(reasons) if reasons else "未知原因"
    lg(f"全部策略失败，摘要：{summary}")
    return False, summary


async def _scroll_home_to_bottom(page: Any) -> None:
    """首页「Hottest Parties」在底部，先滚到底再点 Party Hubs / 底栏。"""
    try:
        await page.evaluate("window.scrollTo(0, 0)")
        await page.wait_for_timeout(150)
    except Exception:
        pass
    try:
        h = await page.evaluate(
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0, "
            "document.scrollingElement?.scrollHeight||0)"
        )
        if h is not None:
            await page.evaluate("y => window.scrollTo(0, y)", max(0, int(h)))
        await page.wait_for_timeout(400)
    except Exception:
        pass
    try:
        await page.keyboard.press("End")
        await page.wait_for_timeout(350)
    except Exception:
        pass


async def _try_click_party_hubs_in_hottest_row(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """「Hottest Parties」标题行与「Party Hubs >」同一块区域（图1 红框）：先滚入视口再点 Hub。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·PartyHubs] " + msg)

    try:
        row = page.locator("div, section, article, header").filter(
            has=page.get_by_text(re.compile(r"Hottest\s*Parties", re.I))
        ).filter(has=page.get_by_text(re.compile(r"Party\s*Hubs", re.I)))
        n = await row.count()
        if n < 1:
            lg("未找到同时含 Hottest Parties 与 Party Hubs 的容器")
            return False, ""
        box = row.last
        await box.scroll_into_view_if_needed(timeout=2800)
        await page.wait_for_timeout(180)
        hub = box.get_by_text(re.compile(r"Party\s*Hubs", re.I)).last
        await hub.wait_for(state="attached", timeout=2200)
        await hub.scroll_into_view_if_needed(timeout=2500)
        try:
            await hub.click(timeout=3500, force=True)
        except Exception:
            await hub.evaluate(
                """(el) => {
                  let n = el;
                  for (let i = 0; i < 10 && n; i++) {
                    const t = (n.tagName || '').toUpperCase();
                    if (t === 'A' || t === 'BUTTON' || n.getAttribute('role') === 'button') {
                      n.click(); return;
                    }
                    n = n.parentElement;
                  }
                  el.click();
                }"""
            )
        await page.wait_for_timeout(800)
        return True, "已点击 Hottest Parties 区块内 Party Hubs"
    except Exception as e:
        lg(f"Hottest 同行：{_brief_exc(e, 180)}")
        return False, ""


async def _try_click_party_hubs_link(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """右侧「Party Hubs >」多为 link；优先于整块标题点击。失败策略须短超时，避免 4×4×3s 级联拖分钟。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·PartyHubs] " + msg)

    ok_row, msg_row = await _try_click_party_hubs_in_hottest_row(page, log=log)
    if ok_row:
        return True, msg_row

    # 可见链路与纯文案分支分开超时：无匹配则 count 为 0 立即跳过，不空等 3s
    t_vis = 1100
    t_att = 1800
    t_clk = 3200

    for pat in (
        r"Party\s*Hubs(?:\s*[>›])?",
        r"Hottest\s*Parties",
    ):
        try:
            gl = page.get_by_role("link", name=re.compile(pat, re.I))
            if await gl.count() < 1:
                lg(f"get_by_role(link) /{pat}/ count=0，跳过")
            else:
                loc = gl.first
                await loc.wait_for(state="visible", timeout=t_vis)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await loc.click(timeout=t_clk)
                await page.wait_for_timeout(450)
                return True, f"已点击链接（get_by_role link），匹配 /{pat}/i"
        except Exception as e:
            lg(f"get_by_role(link) /{pat}/ ：{_brief_exc(e, 140)}")
        try:
            al = page.locator("a").filter(has_text=re.compile(pat, re.I))
            if await al.count() < 1:
                lg(f"locator(a)+text /{pat}/ count=0，跳过")
            else:
                loc = al.first
                await loc.wait_for(state="visible", timeout=t_vis)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await loc.click(timeout=t_clk)
                await page.wait_for_timeout(450)
                return True, f"已点击 <a>，文案匹配 /{pat}/i"
        except Exception as e:
            lg(f"locator(a)+text /{pat}/ ：{_brief_exc(e, 140)}")
        try:
            tl = page.get_by_text(re.compile(pat, re.I))
            if await tl.count() < 1:
                lg(f"get_by_text /{pat}/ count=0，跳过")
            else:
                loc = tl.last
                await loc.wait_for(state="attached", timeout=t_att)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await page.wait_for_timeout(120)
                try:
                    await loc.click(timeout=t_clk, force=True)
                except Exception:
                    await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                await page.wait_for_timeout(450)
                return True, f"已点击文案 .last（force/JS），匹配 /{pat}/i"
        except Exception as e:
            lg(f"get_by_text .last /{pat}/ ：{_brief_exc(e, 140)}")
    lg("所有 Party Hubs / Hottest Parties 链接模式均未成功")
    return False, ""


async def _try_click_bottom_party_nav(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """
    底部 TabBar 第二项「Party」。必须用主导航 tablist 的几何顺序，
    禁止优先 get_by_role(name=Party)：否则会点到内容区筛选里的「Party」。
    """
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·Party底栏] " + msg)

    reasons: list[str] = []
    lg("策略 0：各 frame 精确「Party」.last + JS（底栏在 DOM 中通常后于筛选条）…")
    ok0, m0 = await _try_click_bottom_label_js(page, "Party", log=log, tag="Party底栏")
    if ok0:
        return True, m0

    r = await _best_bottom_tabs(page, log=log)
    if r:
        tabs, tc = r
        if tc >= 2:
            try:
                lg(f"策略 A：主导航 tablist 第 2 项（共 {tc} 项）…")
                t1 = tabs.nth(1)
                await t1.wait_for(state="visible", timeout=3200)
                await t1.scroll_into_view_if_needed()
                await t1.click(timeout=4500)
                await page.wait_for_timeout(900)
                return True, f"已点击主导航 tablist 第 2 项（共 {tc} 项，底部 Party）"
            except Exception as e:
                reasons.append(f"策略A失败：{_brief_exc(e)}")
                lg(reasons[-1])
        else:
            reasons.append(f"主导航仅 {tc} 项，无法取第 2 项作为 Party")
            lg(reasons[-1])
    else:
        reasons.append("无主导航 tablist（与 Home 相同根因）")

    try:
        lg("策略 B：最后一个 [role=tablist] 的第 2 项…")
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        lg(f'[role="tablist"] 数量 = {nl}')
        if nl > 0:
            tablist = lists.nth(nl - 1)
            tabs2 = tablist.locator('[role="tab"]')
            tc = await tabs2.count()
            lg(f"最后一个 tablist 内 tab 数 = {tc}")
            if tc >= 2:
                t = tabs2.nth(1)
                await t.wait_for(state="visible", timeout=3200)
                await t.scroll_into_view_if_needed()
                await t.click(timeout=4500)
                await page.wait_for_timeout(900)
                return True, "已点击最后一个 tablist 的第 2 项（兜底，假定 Home, Party, …）"
    except Exception as e:
        reasons.append(f"策略B失败：{_brief_exc(e)}")
        lg(reasons[-1])

    pat_exact = re.compile(r"^Party$", re.I)
    try:
        lg("策略 C：get_by_role(tab, Party) 仅当匹配数>1 时取末项…")
        tabs = page.get_by_role("tab", name=pat_exact)
        n = await tabs.count()
        lg(f"匹配数 = {n}（≤1 则跳过以免点到筛选条唯一 Party）")
        if n > 1:
            t = tabs.nth(n - 1)
            await t.wait_for(state="visible", timeout=3200)
            await t.scroll_into_view_if_needed()
            await t.click(timeout=4500)
            await page.wait_for_timeout(900)
            return True, f"已点击「Party」页签（仅当匹配数>1 时用末项，共 {n} 个）"
    except Exception as e:
        reasons.append(f"策略C失败：{_brief_exc(e)}")
        lg(reasons[-1])

    try:
        lg("策略 D：a/button/[role=tab] 含 Party 且不含 Hubs/Hottest，匹配数>1 取末项…")
        base = page.locator("a,button,[role='tab']")
        row = base.filter(has_text=re.compile(r"Party", re.I)).filter(
            has_not_text=re.compile(r"Hubs|Hottest", re.I)
        )
        n = await row.count()
        lg(f"匹配数 = {n}")
        if n > 1:
            b = row.nth(n - 1)
            await b.wait_for(state="visible", timeout=3200)
            await b.scroll_into_view_if_needed()
            await b.click(timeout=4500)
            await page.wait_for_timeout(900)
            return True, "已点击「Party」（排除 Hubs/Hottest 且匹配数>1 时取末项）"
    except Exception as e:
        reasons.append(f"策略D失败：{_brief_exc(e)}")
        lg(reasons[-1])

    lg("策略 E：nav/footer 容器（含 Home+Profile/Task）内精确「Party」…")
    ok_e, msg_e = await _try_click_party_in_scoped_bottom_bar(page, log=log)
    if ok_e:
        return True, msg_e

    lg("策略 F：各 frame 精确「Party」.last（无 tablist 的底栏）…")
    ok_f, msg_f = await _try_click_exact_text_last_in_frames(
        page, "Party", log=log, diag_tag="Party·文案"
    )
    if ok_f:
        return True, msg_f + "（精确文案 .last）"

    summary = "；".join(reasons) if reasons else "未知原因"
    lg(f"全部策略失败，摘要：{summary}")
    return False, summary


async def _must_click_then_verify(
    page: Any,
    *,
    case_zh: str,
    tab_patterns: list[str],
    text_click_patterns: list[str],
    verify_patterns: list[str],
    verify_timeout_ms: float = 2800,
    bottom_js_label: str | None = None,
    log: Callable[[str], None] | None = None,
) -> tuple[str, str]:
    """
    先点击（底栏 JS → 页签 → 文案），再校验展开/切换后的可见内容。
    禁止在未点击成功时仅凭「页面上有文案」判 PASS。
    """
    clicked = False
    cm = ""
    if bottom_js_label:
        okb, mb = await _try_click_bottom_label_js(
            page, bottom_js_label, log=log, tag=case_zh
        )
        if okb:
            clicked, cm = True, mb
    if not clicked and tab_patterns:
        clicked, cm = await _try_click_tab(page, tab_patterns)
    if not clicked:
        c2, m2 = await _try_click_visible_text(page, text_click_patterns)
        if c2:
            clicked, cm = True, m2
    if not clicked:
        return (
            "FAIL",
            f"「{case_zh}」未能完成点击（页签与文案点击均失败），不允许仅凭背景可见文案判通过。详情：{cm}",
        )
    ok, vm = await _visible_any(page, verify_patterns, timeout_ms=verify_timeout_ms)
    if not ok:
        return (
            "FAIL",
            f"「{case_zh}」已操作（{cm}），但切换后未见预期内容：{vm}",
        )
    return ("PASS", f"「{case_zh}」{cm}；切换/展开后确认：{vm}")


async def _run_case(
    case_id: str,
    page: Any,
    *,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    """返回 (VERDICT, 中文/可读说明)。"""
    if case_id == "p1_customer_service":
        return await _customer_service_click_open_close(page, log)
    if case_id == "p1_share_tab":
        return await _must_click_then_verify(
            page,
            case_zh="分享页签",
            tab_patterns=[r"^Share$", r"分享", r"Share"],
            text_click_patterns=[r"^Share$", r"\bShare\b", r"分享"],
            bottom_js_label="Share",
            log=log,
            verify_patterns=[
                r"Share",
                r"分享",
                r"invite",
                r"邀请",
                r"refer",
                r"friend",
                r"复制",
                r"Copy\s*link",
                r"Link",
                r"链接",
            ],
        )
    if case_id == "p1_task_tab":
        return await _must_click_then_verify(
            page,
            case_zh="任务页签",
            tab_patterns=[r"^Task$", r"任务", r"Tasks"],
            text_click_patterns=[r"^Task$", r"\bTask\b", r"任务"],
            bottom_js_label="Task",
            log=log,
            verify_patterns=[
                r"\bTask\b",
                r"任务",
                r"Quest",
                r"Daily",
                r"每日",
                r"Reward",
                r"奖励",
                r"Mission",
                r"Complete",
                r"进度",
            ],
        )
    if case_id == "p1_profile_tab":
        return await _must_click_then_verify(
            page,
            case_zh="我的/Profile",
            tab_patterns=[r"Profile", r"^Me$", r"我的", r"Account"],
            text_click_patterns=[r"Profile", r"^Me$", r"我的", r"Account"],
            bottom_js_label="Profile",
            log=log,
            verify_patterns=[
                r"Profile",
                r"我的",
                r"Wallet",
                r"钱包",
                r"Balance",
                r"余额",
                r"Setting",
                r"设置",
                r"VIP",
                r"Account",
                r"账户",
                r"Member",
                r"Logout",
                r"退出",
                r"Avatar",
                r"头像",
            ],
        )
    if case_id == "p1_hottest_parties":
        """
        首页仅滚底一次 → 点 Party Hubs → 子页不滚底，校验顶栏/列表常态文案
        → 点顶栏 img 返回 → 继续后续用例。
        """
        await _ensure_on_home_feed(page, target_url, log)
        log("  [诊断·HottestParty] —— 首页：底栏 Home + 滚至 Hottest / Party Hubs ——")
        await _log_tablist_snapshot(page, log, tag="·Home前")
        try:
            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(180)
        except Exception:
            pass
        home_ok, home_msg = await _try_click_bottom_home_nav(page, log=log)
        if home_ok:
            log(f"  [诊断·HottestParty] 底栏 Home：{home_msg}")
        else:
            log(f"  [诊断·HottestParty] 底栏 Home 未点到：{home_msg}")
        await page.wait_for_timeout(320)
        log("  [诊断·HottestParty] 仅在首页滚底（scrollHeight + End），以露出区块…")
        await _scroll_home_to_bottom(page)

        clicked, cm = await _try_click_party_hubs_link(page, log=log)
        if not clicked:
            c2, m2 = await _try_click_visible_text(
                page,
                [
                    r"Party\s*Hubs",
                    r"Hottest\s*Parties",
                    r"热门\s*Party",
                    r"Party\s*Hub",
                    r"热门",
                ],
                timeout_ms=2800,
            )
            if c2:
                clicked, cm = True, m2
            else:
                log(f"  [诊断·HottestParty] 文案点击未成功：{m2}")
        if not clicked and "party-hubs" in (page.url or "").lower():
            log("  [诊断·HottestParty] URL 已在 party-hubs，继续校验子页…")
            clicked, cm = True, "已进入 Party Hubs（URL）"
        if not clicked:
            await _log_tablist_snapshot(page, log, tag="·首页点 Hubs 失败后")
            await _p1_leave_party_hubs_after_hottest(page, target_url, log)
            hint = (
                f" Home：{home_msg}。"
                if home_ok
                else f" Home 未点到：{home_msg}。"
            )
            return (
                "FAIL",
                "「热门 Party 板块」未点到 Party Hubs。"
                + hint
                + " 详见 [诊断·PartyHubs] 与 tablist 快照。",
            )

        await page.wait_for_timeout(550)
        ok, vm = await _visible_any(
            page,
            list(_P1_PARTY_HUBS_VERIFY_PATTERNS),
            timeout_ms=2800,
        )
        if not ok:
            await _p1_leave_party_hubs_after_hottest(page, target_url, log)
            return (
                "FAIL",
                f"「热门 Party 板块」已进入（{cm}），但 Party Hubs 页未见典型展示：{vm}",
            )

        log("  [诊断·HottestParty] 子页展示正常，点击顶栏返回图标离开…")
        await _p1_leave_party_hubs_after_hottest(page, target_url, log)

        h_part = f"{home_msg}；" if home_ok else "（Home 未点到仍进入 Hubs）"
        return (
            "PASS",
            f"「热门 Party 板块」{h_part}{cm}；子页确认：{vm}；已顶栏返回并回大厅。",
        )
    if case_id == "p1_party_status":
        await _ensure_on_home_feed(page, target_url, log)
        log("  [诊断] —— 快照：Party 用例开始（滚底前）——")
        await _log_tablist_snapshot(page, log, tag="·Party前")
        await _scroll_home_to_bottom(page)
        log("  [诊断] 已滚底，开始点击底栏 Party…")
        clicked, cm = await _try_click_bottom_party_nav(page, log=log)
        if not clicked:
            log("  [诊断] —— 快照：Party 底栏点击失败后 ——")
            await _log_tablist_snapshot(page, log, tag="·Party失败后")
            return (
                "FAIL",
                "「Party 状态展示」须点击底栏「Party」（脚本优先 JS 父级点击，其次含 Home+Share 的 tablist / nav）。"
                f"失败摘要：{cm}。"
                "详见上方 [诊断·Party底栏] 与 tablist 快照。",
            )
        ok, vm = await _visible_any(
            page,
            [
                r"Preparing",
                r"准备",
                r"players?",
                r"人数",
                r"Live",
                r"Party",
                r"Playing",
                r"进行中",
                r"Open",
                r"开局",
            ],
            timeout_ms=3200,
        )
        if not ok:
            return (
                "FAIL",
                f"「Party 状态展示」已操作（{cm}），但未见状态类文案：{vm}",
            )
        return ("PASS", f"「Party 状态展示」{cm}；确认：{vm}")
    return ("BLOCKED", f"脚本未实现该用例：{case_id}")


async def _async_main(args: argparse.Namespace) -> int:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("请先安装依赖：pip install playwright，再执行 playwright install chromium", file=sys.stderr)
        return 2

    cdp = _kalaroko_cdp(args.cdp_http or None)
    target_url = (args.target_url or DEFAULT_TARGET).strip()
    host = _host_from_url(target_url)
    skill_path = Path(args.skill_md)
    if not skill_path.is_absolute():
        skill_path = ROOT / skill_path
    skill_md_loaded = skill_path.is_file()
    if skill_md_loaded:
        skill_body = _load_skill(skill_path).replace("{{CONTEXT_URL}}", target_url.rstrip("/"))
        case_ids = _parse_p1_ids(skill_body)
    else:
        skill_body = ""
        case_ids = _parse_p1_ids(skill_body)
    if not case_ids:
        case_ids = [
            "p1_customer_service",
            "p1_share_tab",
            "p1_task_tab",
            "p1_profile_tab",
            "p1_hottest_parties",
            "p1_party_status",
        ]

    def log(msg: str) -> None:
        if args.verbose or not args.quiet:
            print(msg, flush=True)

    log("———————— K11 P1 · Playwright 直连验收 ————————")
    log(f"调试端口（CDP）：{cdp}")
    log(f"目标站点：{target_url}（匹配主机名：{host or '（任意）'}）")
    if skill_md_loaded:
        log(f"SKILL 文档：{skill_path.resolve()}")
    else:
        log(
            f"SKILL 文档：未加载（路径不存在，使用内置用例顺序）：{skill_path.resolve()}"
        )
    log(f"将执行用例（共 {len(case_ids)} 条）：{', '.join(case_ids)}")
    log("")

    console_bucket: list[str] = []

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(cdp)
        if not browser.contexts:
            print("[失败] 已连上 CDP，但浏览器里没有可用的浏览上下文（context）", file=sys.stderr)
            return 2
        ctx = browser.contexts[0]
        pages = list(ctx.pages)
        if not pages:
            print("[失败] 当前没有打开任何标签页", file=sys.stderr)
            return 2

        picked = None
        for pg in reversed(pages):
            try:
                u = pg.url or ""
            except Exception:
                u = ""
            if host and host in u.lower():
                picked = pg
                break
        if picked is None:
            if args.navigate_if_no_tab:
                picked = pages[-1]
                log(f"[nav] 未找到含 {host!r} 的标签页，对当前页 goto {target_url}")
                await picked.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
            else:
                print(
                    f"[失败] 没有 URL 包含「{host}」的标签页。请先在调试 Chrome 里打开：{target_url}\n"
                    f"  或加参数：--navigate-if-no-tab（会强制在当前页跳转一次）",
                    file=sys.stderr,
                )
                return 2
        else:
            await picked.bring_to_front()

        page = picked

        def _on_console(msg: Any) -> None:
            try:
                t = msg.type
                text = msg.text
                if t == "error":
                    console_bucket.append(f"{t}: {text[:500]}")
            except Exception:
                pass

        page.on("console", _on_console)

        url0 = page.url
        title0 = await page.title()
        log(f"当前标签页 URL：{url0}")
        log(f"当前页面标题：{title0}")
        log("")
        log("准备：若在 party-hubs、app_tabbar=no 或个人中心，先回站点首页，保证底栏与 Hottest 区块可测。")
        await _ensure_on_home_feed(page, target_url, log)
        log("")

        results: list[dict[str, Any]] = []
        for i, cid in enumerate(case_ids, start=1):
            title_zh = CASE_TITLE_ZH.get(cid, cid)
            log(f"【{i}/{len(case_ids)}】{title_zh}（{cid}）")
            v, detail = await _run_case(cid, page, log=log, target_url=target_url)
            vzh = VERDICT_ZH.get(v, v)
            log(f"  观察说明：{detail}")
            log(f"  结论：{vzh}（{v}）")
            log("")
            results.append(
                {
                    "case": cid,
                    "case_title_zh": CASE_TITLE_ZH.get(cid, cid),
                    "verdict": v,
                    "verdict_zh": VERDICT_ZH.get(v, v),
                    "detail": detail,
                }
            )

        severe = [x for x in console_bucket[:20]]
        if severe:
            log("———————— 浏览器 Console 错误（抽样，最多 20 条）————————")
            for s in severe:
                log("  · " + s)
            log("")

        out = {
            "schema": "k11_p1_playwright_smoke/v1",
            "ts_utc": datetime.now(timezone.utc).isoformat(),
            "cdp": cdp,
            "target_url": target_url,
            "skill_md": str(skill_path.resolve()),
            "skill_md_loaded": skill_md_loaded,
            "skill_body_chars": len(skill_body),
            "skill_body_preview": skill_body[:1200] + ("…" if len(skill_body) > 1200 else ""),
            "page_url_final": page.url,
            "page_title_final": await page.title(),
            "console_errors_sample": severe,
            "results": results,
        }

        if args.json_out:
            outp = Path(args.json_out)
            outp.parent.mkdir(parents=True, exist_ok=True)
            outp.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            log(f"结构化报告已写入：{outp.resolve()}")

        if not args.no_xlsx_report:
            xlsx_p = (
                args.xlsx_report
                if args.xlsx_report is not None
                else _default_k11_xlsx_report_path()
            )
            log("")
            write_k11_p1_results_to_xlsx(Path(xlsx_p), results, log=log)

        log("———————— 汇总 ————————")
        for r in results:
            cid = r["case"]
            v = r["verdict"]
            mark = "✓" if v == "PASS" else ("○" if v == "SKIP" else "✗")
            log(f"  {mark} {CASE_TITLE_ZH.get(cid, cid)} → {VERDICT_ZH.get(v, v)}")
        bad = {r["verdict"] for r in results}
        if "FAIL" in bad or "BLOCKED" in bad:
            log("\n最终结果：存在「未通过」或「阻塞」项，退出码 1。")
            return 1
        if "SKIP" in bad and "FAIL" not in bad and "BLOCKED" not in bad:
            log("\n最终结果：均为「通过」或「跳过」，退出码 0。")
        else:
            log("\n最终结果：全部「通过」，退出码 0。")
        return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description="可选读取 SKILL Markdown 中的 ### p1_* 用例顺序；文件不存在时用内置 P1 列表。Playwright CDP 在已打开的 herontest.xin 上跑启发式验收",
    )
    ap.add_argument(
        "--skill-md",
        type=Path,
        default=DEFAULT_SKILL,
        help="SKILL Markdown 路径（可选；不存在则跳过解析，使用内置 6 条 P1）",
    )
    ap.add_argument(
        "--target-url",
        default=DEFAULT_TARGET,
        help="期望测试的站点（用于匹配标签页 host；默认 herontest.xin）",
    )
    ap.add_argument("--cdp-http", default="", help="覆盖 KALAROKO_CDP_ENDPOINT")
    ap.add_argument(
        "--navigate-if-no-tab",
        action="store_true",
        help="若无含目标 host 的标签页，则对当前页执行一次 goto（调试用）",
    )
    ap.add_argument("--json-out", type=Path, default=None, help="写入结构化结果 JSON")
    ap.add_argument(
        "--xlsx-report",
        type=Path,
        default=None,
        help="K11平台测试用例.xlsx；默认 K11_XLSX_REPORT 或 ~/Downloads/K11平台测试用例.xlsx",
    )
    ap.add_argument("--no-xlsx-report", action="store_true", help="不写入 Excel")
    ap.add_argument("--quiet", action="store_true", help="仅输出每条 VERDICT 与失败信息")
    ap.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="始终打印 [诊断] 明细（tablist 快照、Home/Party 各策略异常）；可与 --quiet 同用以只看诊断+结论",
    )
    args = ap.parse_args()
    try:
        return asyncio.run(_async_main(args))
    except Exception as e:
        print(f"[失败] 未捕获异常：{type(e).__name__}: {e}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
