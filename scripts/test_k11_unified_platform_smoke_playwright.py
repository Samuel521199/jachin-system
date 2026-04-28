#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
K11 平台冒烟 · 统合版（单次 CDP 会话顺序执行：P0 八条 + P1 六模块 + 扩展八条 + 弱网一条）

实现由下列脚本**抄写合并**（本文件不 import、不以 subprocess 调用它们）：
  ``test_k11_p0_platform_smoke_playwright.py``、``test_k11_p1_skill_herontest_playwright.py``、
  ``test_k11_platform_smoke_extended_playwright.py``、``test_k11_p2_compat_weaknet_playwright.py``（弱网段）。

前置：与上述脚本相同（``KALAROKO_CDP_ENDPOINT``、Playwright、调试 Chrome）。

用法（仓库根）：
  python scripts/test_k11_unified_platform_smoke_playwright.py
  python scripts/test_k11_unified_platform_smoke_playwright.py --target-url https://www.herontest.xin/
  python scripts/test_k11_unified_platform_smoke_playwright.py -v --json-out out/k11_unified.json

**飞书/Lark 报告**（与 ``K11平台测试用例.xlsx`` 同表头的 Wiki 内嵌**电子表格或多维表**；**默认只同步飞书、不写本机 xlsx**）在 ``.env`` 可配：

- 应用与会话（发卡片/写表）**仅** ``K11_SMOKE_LARK_APP_ID``、``K11_SMOKE_LARK_APP_SECRET``、``K11_SMOKE_LARK_NOTIFY_CHAT_ID``（见 ``scripts/k11_lark_smoke_report.py``）
- ``K11_SMOKE_LARK_WIKI_URL``（可省略；与脚本内 ``K11_DEFAULT_LARK_WIKI_URL`` 同链时即同步到该表）
- 可选：``K11_SMOKE_LARK_TABLE_ID`` / ``K11_SMOKE_LARK_SHEET_ID``（子表 id）
- 加 ``--no-lark-report`` 可不发飞书、不同步表格；加 ``--write-local-xlsx`` 才写入 ``~/Downloads/K11平台测试用例.xlsx``（需 ``openpyxl``）。
- 群通知卡片样式在 ``scripts/k11_lark_smoke_report.send_k11_smoke_lark_notification``：原生 table 三列（测试项目 / 结果 / 备注），失败时降级 lark_md/纯文本。
- 主流程结束后（除非 ``--skip-browser-compat``）会子进程执行 P2「仅兼容」段（``--only-compat``），将「浏览器兼容」并入结果；随后（除非 ``--skip-game-open-smoke``）在同一 CDP 页签上跑 ``test_k11_game_open_smoke`` 的 herontest 五款游戏开门探活，**追加行**到同一 ``results``，与前面用例一并写入飞书表并打在**同一张** Lark 消息卡片表格中。
  随 L3 侧车 / ``l3_node.exe`` 跑统合时**禁止** ``l3_node.exe 某.py``（引导器不会当解释器），须用 ``--jachin-k11-p2-compat-subprocess`` 子命令（与 ``l3_node/http_server`` 一致）。

行为对齐：P0/P1/扩展/弱网各段逻辑与对应单脚本一致（含 P0 Play Now 默认**不点击**）。
若需对 Play Now 做真实点击，请加 ``--p0-play-now-really-click``（点击后自动回大厅）。
"""
from __future__ import annotations

import argparse
import asyncio
import importlib.util
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

_env_root = (os.environ.get("JACHIN_APP_ROOT") or "").strip()
ROOT = Path(_env_root).resolve() if _env_root else Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env", encoding="utf-8")
except ImportError:
    pass
except OSError:
    pass

DEFAULT_TARGET = "https://www.herontest.xin/"

# 与 ``scripts/k11_lark_smoke_report`` 中默认知识表一致；环境变量/CLI 可覆盖
K11_DEFAULT_LARK_WIKI_URL = (
    "https://ssgkm409t6q5.sg.larksuite.com/wiki/ZyWlwhdW1iNQuykvy7qlw93sgTe"
)

# 共 23 条：P0×8 + P1×6 + 扩展×8 + 弱网×1
UNIFIED_CASE_DEFS: list[tuple[str, str, str]] = [
    ("p0_env_access", "P0", "环境访问"),
    ("p0_home_load", "P0", "首页加载"),
    ("p0_page_title", "P0", "页面标题"),
    ("p0_play_now", "P0", "主按钮可用（Play Now!）"),
    ("p0_category_tabs", "P0", "分类切换（All / 1 vs 1 / Party / Live）"),
    ("p0_game_cards", "P0", "游戏卡片展示"),
    ("p0_console_clean", "P0", "页面无严重报错（Console 抽样）"),
    ("p0_release_checklist", "P0", "本次更新点验证"),
    ("p1_customer_service", "P1", "客服入口"),
    ("p1_share_tab", "P1", "分享页签"),
    ("p1_task_tab", "P1", "任务页签"),
    ("p1_profile_tab", "P1", "我的/Profile"),
    ("p1_hottest_parties", "P1", "热门 Party 板块"),
    ("p1_party_status", "P1", "Party 状态展示"),
    ("ext_game_list", "P1", "列表完整性（主要游戏名可见）"),
    ("ext_images", "P1", "图片资源（img 裂图抽样）"),
    ("ext_no_more_data", "P1", "无数据提示（No More Data）"),
    ("ext_response_time", "P1", "响应时间（底栏 Party 切换耗时）"),
    ("ext_copy_light", "P2", "文案检查（轻量：Unicode 替换符）"),
    ("ext_layout_light", "P2", "样式检查（轻量：横向溢出提示）"),
    ("ext_scroll_light", "P2", "滚动加载（轻量：滚底后高度与稳定性）"),
    ("ext_static_console", "P1", "静态资源/模块加载（Console MIME/模块脚本抽样）"),
    ("p2_weak_network", "P2", "弱网体验（Slow3G 类 CDP 限速）"),
]

CASE_TITLE_ZH: dict[str, str] = {k: v for k, _, v in UNIFIED_CASE_DEFS}

VERDICT_ZH: dict[str, str] = {
    "PASS": "通过",
    "FAIL": "未通过",
    "SKIP": "跳过",
    "BLOCKED": "阻塞",
}

def _resolve_p2_compat_script_path() -> Path:
    """
    与 ``l3_node.paths.k11_p2_compat_weaknet_script_path`` 一致：开发用仓库 ``scripts/``，
    PyInstaller 用 ``_MEIPASS/scripts/``，便携用 ``JACHIN_APP_ROOT/scripts/``。
    """
    fname = "test_k11_p2_compat_weaknet_playwright.py"
    u = Path(__file__).resolve().parent / fname
    if u.is_file():
        return u
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        p = Path(sys._MEIPASS) / "scripts" / fname
        if p.is_file():
            return p
    portable = ROOT / "scripts" / fname
    if portable.is_file():
        return portable
    return ROOT / "scripts" / fname


def _k11_p2_compat_subprocess_cmd(passthrough: list[str], p2_script: Path) -> list[str]:
    """
    frozen：``l3_node.exe --jachin-k11-p2-compat-subprocess ...``；
    开发：``python -m l3_node --jachin-k11-p2-compat-subprocess ...``。
    与 ``l3_node.http_server._k11_smoke_subprocess_cmd`` 一致，避免 ``exe script.py`` 失败。
    若当前环境无法 ``import l3_node``（裸 ``python scripts/...``），回退为 ``python P2脚本.py ...``。
    """
    sent = "--jachin-k11-p2-compat-subprocess"
    if getattr(sys, "frozen", False):
        return [sys.executable, sent, *passthrough]
    try:
        import l3_node  # noqa: F401
    except ImportError:
        return [sys.executable, str(p2_script), *passthrough]
    return [sys.executable, "-m", "l3_node", sent, *passthrough]


def _resolve_game_open_smoke_script_path() -> Path:
    """与 ``l3_node.paths.k11_game_open_smoke_script_path`` 对齐：frozen / 便携 / 仓库 scripts。"""
    fname = "test_k11_game_open_smoke.py"
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        p = Path(sys._MEIPASS) / "scripts" / fname
        if p.is_file():
            return p
    portable = ROOT / "scripts" / fname
    if portable.is_file():
        return portable
    return Path(__file__).resolve().parent / fname


def _load_game_open_smoke_module(log: Callable[[str], None] | None = None) -> Any:
    """
    动态加载游戏开门脚本。须在 ``exec_module`` 前 ``sys.modules[spec.name] = mod``，
    否则 Python 3.10+ 下 ``@dataclass`` 等会报 ``NoneType has no attribute '__dict__'`` 并静默失败。
    """
    p = _resolve_game_open_smoke_script_path()
    if not p.is_file():
        if log:
            log(f"  [game_open] 脚本不存在：{p}")
        return None
    spec = importlib.util.spec_from_file_location(
        "k11_game_open_smoke_unified_embed", str(p.resolve())
    )
    if spec is None or spec.loader is None:
        if log:
            log("  [game_open] importlib spec/loader 创建失败")
        return None
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    try:
        spec.loader.exec_module(mod)
    except Exception as e:
        if log:
            log(f"  [game_open] 加载脚本失败：{_brief_exc(e, 480)}")
        return None
    return mod


def _resolve_k11_lark_smoke_report_path() -> Path:
    """
    开发：仓库根 ``scripts/k11_lark_smoke_report.py``；
    打包/侧车：常与 ``本脚本`` 同目录（_MEI/.../scripts/），或 ``JACHIN_APP_ROOT/scripts/``。
    """
    u = Path(__file__).resolve().parent / "k11_lark_smoke_report.py"
    if u.is_file():
        return u
    return ROOT / "scripts" / "k11_lark_smoke_report.py"


UNIFIED_CASE_TO_XLSX_TEST_ITEM_KEY: dict[str, str] = {
    "p0_env_access": "环境访问",
    "p0_home_load": "首页加载",
    "p0_page_title": "页面标题",
    "p0_play_now": "主按钮可用",
    "p0_category_tabs": "分类切换",
    "p0_game_cards": "游戏卡片展示",
    "p0_console_clean": "无严重报错",
    "p0_release_checklist": "本次更新点",
    "p1_customer_service": "客服",
    "p1_share_tab": "分享",
    "p1_task_tab": "任务",
    "p1_profile_tab": "我的",
    "p1_hottest_parties": "热门",
    "p1_party_status": "Party 状态",
    "ext_game_list": "列表完整性",
    "ext_images": "图片资源",
    "ext_no_more_data": "无数据提示",
    "ext_response_time": "响应时间",
    "ext_copy_light": "文案检查",
    "ext_layout_light": "样式检查",
    "ext_scroll_light": "滚动加载",
    "ext_static_console": "静态资源",
    "p2_weak_network": "弱网",
    # 子进程 ``test_k11_p2_compat_weaknet_playwright.py --only-compat`` 合并行，与 P2 脚本一致
    "p2_browser_compat_merged": "浏览器兼容",
    # ``test_k11_game_open_smoke`` 并入统合后的行（与游戏标题一致，便于 xlsx/多维表关键词匹配）
    "game_open_bingo_showdown": "Bingo Showdown",
    "game_open_infinity_9_ball": "Infinity 9 Ball",
    "game_open_color_blitz_social": "Color Blitz Social",
    "game_open_royal_pusoy": "Royal Pusoy",
    "game_open_drama_crush": "Drama Crush",
}

_XLSX_REMARK_MAX_LEN = 32000


def _default_k11_xlsx_report_path() -> Path:
    env = (os.environ.get("K11_XLSX_REPORT") or "").strip()
    if env:
        return Path(env)
    return Path.home() / "Downloads" / "K11平台测试用例.xlsx"


def _run_p2_only_compat_subprocess(
    *,
    target_url: str,
    project_root: Path,
    log: Callable[[str], None],
    headless: bool,
    quiet_p2: bool,
) -> tuple[list[dict[str, Any]], int, str | None]:
    """
    子进程执行 ``scripts/test_k11_p2_compat_weaknet_playwright.py --only-compat``，
    经 P2 内 ``_p2_all_rows_to_lark_results`` 转为与统合 ``results`` 同结构的行（通常 1 条：浏览器兼容）。
    返回 (追加行列表, 子进程 exit code, 非 None 时为一行式错误原因)。
    """
    p2_script = _resolve_p2_compat_script_path()
    if not p2_script.is_file():
        return (
            [
                {
                    "case": "p2_browser_compat_merged",
                    "tier": "P2",
                    "case_title_zh": "浏览器兼容",
                    "verdict": "BLOCKED",
                    "verdict_zh": "阻塞",
                    "detail": (
                        f"未找到脚本：{p2_script}（请把 test_k11_p2_compat_weaknet_playwright.py 放入 "
                        "与统合脚本同目录的 scripts/ 或 JACHIN_APP_ROOT/scripts/；"
                        "勿用 l3_node.exe 直接执行 .py，应随 L3 子命令加载。）"
                    ),
                }
            ],
            2,
            "missing p2 script",
        )

    fd, jpath = tempfile.mkstemp(suffix="_k11_compat.json", text=True)
    os.close(fd)
    json_path = Path(jpath)
    try:
        passthrough: list[str] = [
            "--only-compat",
            "--no-lark-report",
            "--json-out",
            str(json_path),
            "--target-url",
            (target_url or "").strip() or "https://www.herontest.xin/",
        ]
        if headless:
            passthrough.append("--headless")
        if quiet_p2:
            passthrough.append("--quiet")
        cmd = _k11_p2_compat_subprocess_cmd(passthrough, p2_script)
        log(
            "  [compat] 启动子进程：l3_node --jachin-k11-p2-compat-subprocess "
            "（P2 --only-compat，自启 Chrome/Edge；与侧车 frozen 兼容）"
        )
        cp = subprocess.run(
            cmd,
            cwd=str(project_root),
            env=os.environ.copy(),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        rc = int(cp.returncode or 0)
        err_tail = ""
        out = (cp.stdout or "") + (cp.stderr or "")
        if out.strip():
            tail = out.strip()[-1200:]
            err_tail = tail
            if rc != 0 or (not json_path.is_file()):
                log(f"  [compat] 子进程输出（尾部）：{tail}")
        if not json_path.is_file():
            extra = f" 子进程输出尾部：{err_tail[:600]}" if err_tail else ""
            return (
                [
                    {
                        "case": "p2_browser_compat_merged",
                        "tier": "P2",
                        "case_title_zh": "浏览器兼容",
                        "verdict": "BLOCKED",
                        "verdict_zh": "阻塞",
                        "detail": (
                            f"子进程结束 exit={rc} 但未生成 JSON 报告文件。"
                            + (extra if extra.strip() else "")
                        ),
                    }
                ],
                rc,
                "missing json",
            )
        raw = json_path.read_text(encoding="utf-8")
        data = json.loads(raw)
        all_rows = data.get("results")
        if not isinstance(all_rows, list):
            return (
                [
                    {
                        "case": "p2_browser_compat_merged",
                        "tier": "P2",
                        "case_title_zh": "浏览器兼容",
                        "verdict": "BLOCKED",
                        "verdict_zh": "阻塞",
                        "detail": "子进程 JSON 中无有效 results 数组。",
                    }
                ],
                rc,
                "bad json",
            )
        _spec = importlib.util.spec_from_file_location(
            "k11_p2_compat_unified_merge", p2_script
        )
        if not _spec or not _spec.loader:
            return (
                [
                    {
                        "case": "p2_browser_compat_merged",
                        "tier": "P2",
                        "case_title_zh": "浏览器兼容",
                        "verdict": "BLOCKED",
                        "verdict_zh": "阻塞",
                        "detail": "无法加载 P2 脚本以归并结果。",
                    }
                ],
                rc,
                "import",
            )
        p2m = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(p2m)
        merged: list[dict[str, Any]] = p2m._p2_all_rows_to_lark_results(  # type: ignore[attr-defined]
            all_rows
        )
        if not merged and all_rows:
            return (
                [
                    {
                        "case": "p2_browser_compat_merged",
                        "tier": "P2",
                        "case_title_zh": "浏览器兼容",
                        "verdict": "BLOCKED",
                        "verdict_zh": "阻塞",
                        "detail": "P2 归并结果为空。",
                    }
                ],
                rc,
                "empty merge",
            )
        if not merged:
            log("  [compat] 子进程未返回可合并行，跳过追加。")
            return [], rc, None
        return merged, rc, None
    except (json.JSONDecodeError, OSError) as e:
        return (
            [
                {
                    "case": "p2_browser_compat_merged",
                    "tier": "P2",
                    "case_title_zh": "浏览器兼容",
                    "verdict": "BLOCKED",
                    "verdict_zh": "阻塞",
                    "detail": f"读取/解析子进程报告失败：{e!s}",
                }
            ],
            2,
            str(e),
        )
    except Exception as e:
        return (
            [
                {
                    "case": "p2_browser_compat_merged",
                    "tier": "P2",
                    "case_title_zh": "浏览器兼容",
                    "verdict": "BLOCKED",
                    "verdict_zh": "阻塞",
                    "detail": f"浏览器兼容子进程链异常：{e!s}",
                }
            ],
            2,
            str(e),
        )
    finally:
        try:
            json_path.unlink(missing_ok=True)
        except OSError:
            pass


def _find_smoke_sheet_header(ws: Any) -> tuple[int, int, int, int] | None:
    max_r = min(int(ws.max_row or 1), 45)
    max_c = min(int(ws.max_column or 1), 40)
    for r in range(1, max_r + 1):
        col_map: dict[str, str] = {}
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            s = str(v).strip() if v is not None else ""
            if s and s not in col_map:
                col_map[s] = c
        if "结果" not in col_map or "备注" not in col_map:
            continue
        item_col = None
        for k in ("测试项目", "测试项", "用例名称"):
            if k in col_map:
                item_col = col_map[k]
                break
        if item_col is None:
            continue
        return r, item_col, col_map["结果"], col_map["备注"]
    return None


def write_k11_unified_results_to_xlsx(
    xlsx_path: Path,
    results: list[dict[str, Any]],
    *,
    log: Callable[[str], None],
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("  [xlsx] 未安装 openpyxl，跳过写入（可：pip install openpyxl）")
        return

    if not xlsx_path.is_file():
        log(f"  [xlsx] 文件不存在，跳过：{xlsx_path}")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        log(f"  [xlsx] 打开工作簿失败：{e!s}")
        return

    ordered_names: list[str] = [n for n in wb.sheetnames if "冒烟" in n]
    ordered_names.extend(n for n in wb.sheetnames if n not in ordered_names)
    verdict_cell = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP", "BLOCKED": "BLOCKED"}

    try:
        for sn in ordered_names:
            ws = wb[sn]
            parsed = _find_smoke_sheet_header(ws)
            if not parsed:
                continue
            hdr, c_item, c_res, c_note = parsed
            last_r = max(int(ws.max_row or hdr), hdr + 5, 80)
            sheet_wrote = 0
            for resrow in results:
                cid = str(resrow.get("case") or "")
                key = UNIFIED_CASE_TO_XLSX_TEST_ITEM_KEY.get(cid)
                if not key:
                    continue
                v = str(resrow.get("verdict") or "")
                detail = str(resrow.get("detail") or "")
                cell_v = verdict_cell.get(v, v)
                remark = (detail or "")[:_XLSX_REMARK_MAX_LEN]
                matched = False
                for r in range(hdr + 1, last_r + 1):
                    raw = ws.cell(row=r, column=c_item).value
                    text = str(raw).strip() if raw is not None else ""
                    if not text:
                        continue
                    if key in text:
                        ws.cell(row=r, column=c_res, value=cell_v)
                        ws.cell(row=r, column=c_note, value=remark)
                        sheet_wrote += 1
                        matched = True
                        break
                if not matched:
                    log(f"  [xlsx] 未匹配行：case={cid} 关键字={key!r} sheet={sn!r}")
            if sheet_wrote:
                try:
                    wb.save(xlsx_path)
                except PermissionError:
                    log(
                        f"  [xlsx] 保存被拒绝（是否正用 Excel 打开该文件？）：{xlsx_path.resolve()}"
                    )
                    return
                log(
                    f"  [xlsx] 已在工作表「{sn}」更新 {sheet_wrote} 行 → {xlsx_path.resolve()}"
                )
                return

        log(
            "  [xlsx] 未找到含「测试项目/结果/备注」表头的工作表，或未更新任何行。"
        )
    except Exception as e:
        log(f"  [xlsx] 写入过程异常：{e!s}")


BOTTOM_SHELL_SELECTORS = (
    "nav",
    "footer",
    "[class*='tabbar' i]",
    "[class*='TabBar']",
    "[class*='tab-bar' i]",
    "[class*='bottomNav' i]",
    "[class*='BottomNav' i]",
    "[class*='footer' i]",
)

def _kalaroko_cdp(cli: str | None) -> str:
    raw = (cli or "").strip() or (os.environ.get("KALAROKO_CDP_ENDPOINT") or "").strip()
    if not raw:
        raw = "http://127.0.0.1:9222"
    if not raw.startswith("http://") and not raw.startswith("https://"):
        raw = "http://" + raw.lstrip("/")
    return raw.rstrip("/")


def _host_from_url(url: str) -> str:
    try:
        from urllib.parse import urlparse

        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def _home_feed_url(target: str) -> str:
    """与 P1 一致：站点根路径大厅，便于分类条与卡片区稳定。"""
    from urllib.parse import urlparse, urlunparse

    t = (target or DEFAULT_TARGET).strip() or DEFAULT_TARGET
    p = urlparse(t)
    if p.scheme and p.netloc:
        return urlunparse((p.scheme, p.netloc, "/", "", "", ""))
    return t.rstrip("/") + "/" if t else DEFAULT_TARGET


def _needs_goto_home_feed(current_url: str) -> bool:
    u = (current_url or "").lower()
    if "/my/" in u or "/me/" in u:
        return True
    if re.search(r"/(profile|account|wallet|settings)(/|$)", u):
        return True
    if "party-hubs" in u or "/party/hub" in u:
        return True
    if "app_tabbar=no" in u:
        return True
    return False


_KALAROKO_NOTIF_BANNER_RE = re.compile(
    r"Turn on notifications|exclusive bonus|"
    r"open notifications|claim .{0,32} bonus|"
    r"打开通知|开启通知|领取.*奖",
    re.I,
)


async def _dismiss_kalaroko_notification_prompt(
    page: Any,
    *,
    log: Callable[[str], None] | None = None,
) -> None:
    """
    站內偶现底部浮层：「Turn on notifications to claim your exclusive bonus」+ Cancel/Agree。
    点 Cancel 或 Esc 关闭，避免遮挡底栏与后续 Locator 超时。
    """
    try:
        hint = page.get_by_text(_KALAROKO_NOTIF_BANNER_RE)
        if await hint.count() < 1:
            return
    except Exception:
        return
    cancel_name = re.compile(r"^Cancel$", re.I)
    frames: list[Any] = []
    try:
        frames = list(page.frames)
    except Exception:
        try:
            frames = [page.main_frame]
        except Exception:
            return
    for fr in frames:
        try:
            by_role = fr.get_by_role("button", name=cancel_name)
            n = await by_role.count()
            if n >= 1:
                await by_role.last.click(timeout=3000, force=True)
                if log:
                    log("  [关弹窗] 已点 Cancel（站內通知/奖励推广）。")
                await page.wait_for_timeout(250)
                return
        except Exception:
            pass
        try:
            by_t = fr.get_by_text("Cancel", exact=True)
            if await by_t.count() >= 1:
                await by_t.last.click(timeout=3000, force=True)
                if log:
                    log("  [关弹窗] 已点 Cancel（站內通知，原文案节点）。")
                await page.wait_for_timeout(250)
                return
        except Exception:
            pass
    try:
        await page.keyboard.press("Escape")
        if log:
            log("  [关弹窗] 已发 Esc 尝试关闭通知推广层。")
        await page.wait_for_timeout(200)
    except Exception:
        pass


async def _ensure_on_home_feed(
    page: Any, target_url: str, log: Callable[[str], None] | None
) -> None:
    home = _home_feed_url(target_url)
    try:
        cur = page.url or ""
    except Exception:
        cur = ""
    if not _needs_goto_home_feed(cur):
        if log:
            log(f"  [诊断·P0] 无需回大厅：{cur!r}")
        await _dismiss_kalaroko_notification_prompt(page, log=log)
        return
    if log:
        log(f"  [诊断·P0] 回大厅：{cur!r} → goto {home!r}")
    await _robust_goto_kalaroko_home(
        page, home, log=log, tag="回大厅", settle_ms=600
    )
    await _dismiss_kalaroko_notification_prompt(page, log=log)


async def _robust_goto_kalaroko_home(
    page: Any,
    url: str,
    *,
    log: Callable[[str], None] | None,
    tag: str = "goto",
    settle_ms: int = 600,
) -> None:
    """
    单页/SPA 下 ``domcontentloaded`` 可能长期不触发或 CDP 偶发卡死；采用多段策略 + 多轮重试，避免一条 goto 拖死整轮统合（exit 3）。
    """
    last: BaseException | None = None
    for round_i in range(3):
        # 1) 首选 domcontentloaded（与历史一致，略放宽到 90s）
        for phase, tmo, extra in (
            ("domcontentloaded", 90_000, None),
            ("commit", 22_000, "dcl"),  # 仅收到响应头后再等 DCL/短暂 settle
        ):
            try:
                if extra == "dcl":
                    await page.goto(url, wait_until="commit", timeout=tmo)
                    try:
                        await page.wait_for_load_state("domcontentloaded", timeout=75_000)
                    except Exception:
                        # 部分 SPA 主文档不典型触发 DCL，短等让主线程推进
                        await page.wait_for_timeout(1800)
                else:
                    await page.goto(url, wait_until=phase, timeout=tmo)
                if log:
                    log(f"  [nav] {tag} 成功：wait_until={phase!r}（第 {round_i + 1} 轮）")
                if settle_ms > 0:
                    await page.wait_for_timeout(settle_ms)
                return
            except Exception as e:
                last = e
                if log:
                    log(
                        f"  [nav] {tag} 第 {round_i + 1} 轮 {phase!r} 失败："
                        f"{_brief_exc(e, 160)}"
                    )
        if round_i < 2:
            if log:
                log(f"  [nav] {tag} 第 {round_i + 1} 轮整轮重试，等待 {1.0 + round_i * 0.5:.1f}s…")
            await page.wait_for_timeout(int(1000 + 500 * round_i))
    if last is not None:
        raise last
    raise RuntimeError(f"{tag}：未知导航失败")


# 分类条：文案有时在子节点内，与底栏相同策略（P1 _JS_CLICK_TAB_FROM_LABEL）
_JS_CLICK_TAB_FROM_LABEL = """(el) => {
  let n = el.parentElement;
  for (let i = 0; i < 12 && n; i++) {
    const tag = (n.tagName || '').toUpperCase();
    const cls = (typeof n.className === 'string') ? n.className : '';
    const role = n.getAttribute && n.getAttribute('role');
    const isLabelOnly = /_item_label_/i.test(cls);
    if (!isLabelOnly && (tag === 'BUTTON' || tag === 'A' || role === 'button' || role === 'tab'
        || /\\bitem\\b/i.test(cls) || (cls.includes('_item_') && !cls.includes('_item_label_')))) {
      n.click();
      return { ok: true, step: i, tag, cls: cls.slice(0, 72) };
    }
    n = n.parentElement;
  }
  if (el.parentElement) {
    el.parentElement.click();
    return { ok: true, fallback: 'parent', tag: el.parentElement.tagName };
  }
  el.click();
  return { ok: true, fallback: 'label-only', tag: el.tagName };
}"""


def _cdp_tab_url_driver_safe(url: str) -> bool:
    """排除 DevTools / 扩展页等：对其 goto 易导致 TargetClosedError 或无效。"""
    u = (url or "").strip().lower()
    if u.startswith("devtools://") or u.startswith("chrome-devtools://"):
        return False
    if u.startswith("chrome-extension://") or u.startswith("moz-extension://"):
        return False
    if u.startswith("ms-browser-extension://"):
        return False
    return True


async def _probe_page_alive(pg: Any) -> bool:
    """参考 kalaroko_capture_page_metrics：避免选中已关闭或不可执行的页签引用。"""
    try:
        if pg.is_closed():
            return False
        await asyncio.wait_for(pg.evaluate("() => 1"), timeout=3.0)
        return True
    except Exception:
        return False


def _brief_exc(e: BaseException, lim: int = 200) -> str:
    return f"{type(e).__name__}: {str(e).strip()[:lim]}"


def _is_benign_console_line(text: str) -> bool:
    t = (text or "").lower()
    if "font-size:0" in t and "nan" in t:
        return True
    if "failed to load resource" in t and "favicon" in t:
        return True
    if "resizeobserver" in t:
        return True
    return False


async def _eval_safe(page: Any, expr: str, *, timeout: float = 12.0) -> Any:
    return await asyncio.wait_for(page.evaluate(expr), timeout=timeout)


async def _eval_timeout(
    page: Any,
    expression: str,
    arg: Any = None,
    *,
    timeout: float = 15.0,
) -> Any:
    """避免页面主线程卡死导致 evaluate 永久挂起；超时由 asyncio 控制。"""
    if arg is None:
        return await asyncio.wait_for(page.evaluate(expression), timeout=timeout)
    return await asyncio.wait_for(page.evaluate(expression, arg), timeout=timeout)


def _exc_tail(e: BaseException, lim: int = 200) -> str:
    return _brief_exc(e, lim)

async def _acquire_cdp_target_page(
    browser: Any,
    *,
    host: str,
    target_url: str,
    navigate_if_no_tab: bool,
    log: Callable[[str], None],
) -> tuple[Any | None, str | None]:
    """
    扫描全部 browser.contexts（与仅 contexts[0] 相比更贴近真实多窗口场景）：
    优先选用 URL 含目标 host 且可驱动的页签；否则在允许导航时依次尝试
    安全 URL 的存活页签 goto，失败则 new_page 后 goto（避免末页签常为 DevTools）。
    """
    if not browser.contexts:
        return None, "CDP 已连上但无 context"

    def _safe_url(pg: Any) -> str:
        try:
            return (pg.url or "").strip()
        except Exception:
            return ""

    has_any_page = any(len(list(getattr(c, "pages", []) or [])) > 0 for c in browser.contexts)

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            if host and host in u.lower():
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None

    if not navigate_if_no_tab:
        if not has_any_page:
            return None, "无打开标签页（且未允许自动 goto）"
        return (
            None,
            f"无含 {host!r} 的标签页。请打开站点，或去掉 --require-existing-tab 以允许自动 goto",
        )

    for ctx in browser.contexts:
        for pg in reversed(list(getattr(ctx, "pages", []) or [])):
            u = _safe_url(pg)
            if not _cdp_tab_url_driver_safe(u):
                continue
            if not await _probe_page_alive(pg):
                continue
            log(f"[nav] 无匹配 {host!r}，尝试在存活页签 goto {target_url!r}（当前 {u[:96]!r}）")
            try:
                await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
                await pg.wait_for_timeout(400)
                try:
                    await pg.bring_to_front()
                except Exception:
                    pass
                return pg, None
            except Exception as e:
                log(f"  [nav] 该页签 goto 失败：{_brief_exc(e)}，换候选或新开标签…")
                continue

    ctx_new = browser.contexts[0]
    for ctx in browser.contexts:
        for pg in list(getattr(ctx, "pages", []) or []):
            if await _probe_page_alive(pg):
                ctx_new = ctx
                break
        else:
            continue
        break

    log(f"[nav] 无可用页签可复用，在 context 中新开标签并 goto {target_url!r}")
    try:
        pg = await ctx_new.new_page()
        await pg.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
        await pg.wait_for_timeout(400)
        return pg, None
    except Exception as e:
        return None, f"新开标签并导航失败：{_brief_exc(e)}"


async def _ensure_target_page(
    page: Any,
    target_url: str,
    *,
    log: Callable[[str], None],
    navigate_if_no_tab: bool,
    host: str,
) -> tuple[bool, str]:
    """环境访问：当前标签含目标 host；必要时 goto。"""
    try:
        u = (page.url or "").strip()
    except Exception:
        u = ""
    if host and host in u.lower():
        return True, f"当前页已为目标域：{u!r}"
    if navigate_if_no_tab:
        log(f"  [诊断·P0] goto {target_url!r}")
        await page.goto(target_url, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(400)
        return True, f"已 goto {target_url!r}"
    return (
        False,
        f"当前 URL {u!r} 不含 host {host!r}；请加 --navigate-if-no-tab 或先手动打开站点。",
    )


async def _run_p0_home_load(page: Any) -> tuple[str, str]:
    try:
        rs = await _eval_safe(page, "() => document.readyState")
        if rs != "complete":
            try:
                await page.wait_for_load_state("load", timeout=12_000)
            except Exception:
                pass
            rs = await _eval_safe(page, "() => document.readyState")
        n = await _eval_safe(
            page,
            "() => (document.body && document.body.innerText) ? document.body.innerText.length : 0",
        )
        if int(n or 0) < 80:
            return ("FAIL", f"首屏正文过短（innerText 长度 {n}），疑似白屏或壳页。")
        return ("PASS", f"readyState={rs!r}，正文长度≈{n}。")
    except Exception as e:
        return ("FAIL", f"首屏/readyState 检测异常：{_brief_exc(e)}")


async def _run_p0_page_title(page: Any) -> tuple[str, str]:
    try:
        t = (await page.title() or "").strip()
    except Exception as e:
        return ("FAIL", f"读取标题失败：{_brief_exc(e)}")
    if re.search(r"KalaroKo", t, re.I):
        return ("PASS", f"标题匹配 KalaroKo：{t!r}")
    return ("FAIL", f"标题未包含 KalaroKo：{t!r}")


async def _click_locator_robust(loc: Any, page: Any, *, timeout_ms: int = 7000) -> None:
    """Playwright click → P1 式父级 DOM click → 合成 click（应对遮罩/可见性误判）。"""
    await loc.wait_for(state="attached", timeout=min(10_000, timeout_ms + 2000))
    try:
        await loc.scroll_into_view_if_needed(timeout=4000)
    except Exception:
        pass
    try:
        await loc.click(timeout=timeout_ms, force=True)
        return
    except Exception:
        pass
    try:
        await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
        return
    except Exception:
        pass
    await loc.evaluate(
        "e => e.dispatchEvent(new MouseEvent('click', {bubbles:true,cancelable:true,view:window}))"
    )


def _category_tab_name_re(label: str) -> re.Pattern[str]:
    if re.sub(r"\s+", " ", label.strip()).lower() in ("1 vs 1", "1vs1"):
        return re.compile(r"^\s*1\s*v\s*s\s*1\s*$", re.I)
    return re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)


async def _run_p0_play_now(
    page: Any,
    *,
    really_click: bool = False,
    target_url: str = "",
    log: Callable[[str], None] | None = None,
) -> tuple[str, str]:
    """
    与 ``test_k11_p0_platform_smoke_playwright.py`` 默认一致：只校验可见与热区，**不**点按钮（避免进局打断后续）。
    若 ``really_click=True``：在通过校验后执行与 P0 相同的稳健点击，再 ``goto`` 回大厅，便于验收「可点」。
    """
    pat = re.compile(r"Play\s*Now", re.I)
    try:
        loc = page.get_by_role("button", name=pat).first
        if await loc.count() < 1:
            loc = page.get_by_text(pat).first
        if await loc.count() < 1:
            return ("FAIL", "未找到「Play Now!」按钮或等价文案。")
        await loc.wait_for(state="visible", timeout=5000)
        await loc.scroll_into_view_if_needed(timeout=4000)
        box = await loc.bounding_box()
        if not box or box.get("width", 0) < 4 or box.get("height", 0) < 4:
            return ("FAIL", "Play Now! 可见但点击区域异常（bounding box）。")
        if not really_click:
            return ("PASS", "Play Now! 可见且具备有效点击区域（未实际点击以防打断后续用例）。")
        if log:
            log("  [诊断·P0·PlayNow] --p0-play-now-really-click：执行点击并回大厅…")
        await _click_locator_robust(loc, page, timeout_ms=8000)
        await page.wait_for_timeout(900)
        home = _home_feed_url(target_url) if target_url else DEFAULT_TARGET
        await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(500)
        return ("PASS", "已真实点击 Play Now!；已 goto 回站点大厅（与统合脚本后续用例衔接）。")
    except Exception as e:
        return ("FAIL", f"Play Now 不可用：{_brief_exc(e)}")


async def _first_home_category_tablist(page: Any) -> Any | None:
    """首页中部筛选条：返回该 tablist 容器 Locator（勿用全页 .last，避免点到底栏同名 tab）。"""
    lists = page.locator('[role="tablist"]')
    nl = await lists.count()
    for idx in range(nl):
        tl = lists.nth(idx)
        tabs = tl.locator('[role="tab"]')
        tc = await tabs.count()
        if tc < 4:
            continue
        has_all = await tl.get_by_text(re.compile(r"^All$", re.I)).count()
        if has_all < 1:
            continue
        return tl
    return None


async def _run_p0_category_tabs(page: Any) -> tuple[str, str]:
    labels = ["All", "1 vs 1", "Party", "Live"]
    try:
        tl = await _first_home_category_tablist(page)
        if tl is None:
            return ("FAIL", "未找到含 All 且至少 4 项的 [role=tablist]（分类条）。")
        try:
            await tl.scroll_into_view_if_needed(timeout=4000)
        except Exception:
            pass
        clicked: list[str] = []
        for lab in labels:
            try:
                name_re = _category_tab_name_re(lab)
                raw = tl.get_by_role("tab", name=name_re)
                if await raw.count() < 1:
                    raw = tl.locator('[role="tab"]').filter(has_text=name_re)
                if await raw.count() < 1:
                    return ("FAIL", f"在本条分类 tablist 内未找到「{lab}」。")
                tloc = raw.first
                await _click_locator_robust(tloc, page, timeout_ms=8000)
                clicked.append(lab)
                await page.wait_for_timeout(450)
                n = await _eval_safe(
                    page,
                    "() => (document.body && document.body.innerText) ? document.body.innerText.length : 0",
                    timeout=8.0,
                )
                if int(n or 0) < 40:
                    return ("FAIL", f"点击「{lab}」后正文异常缩短，疑似白屏。")
            except Exception as e:
                return ("FAIL", f"切换「{lab}」失败：{_brief_exc(e)}")
        return ("PASS", "已依次点击（限定大厅分类 tablist）：" + " → ".join(clicked) + "。")
    except Exception as e:
        return ("FAIL", f"分类切换异常：{_brief_exc(e)}")


async def _run_p0_game_cards(page: Any) -> tuple[str, str]:
    try:
        imgs = page.locator("main img, [class*='card'] img, article img, a img")
        n = await imgs.count()
        if n < 2:
            imgs = page.locator("img")
            n = await imgs.count()
        if n < 2:
            return ("FAIL", f"首页可见 img 过少（{n}），卡片区可能未渲染。")
        broken = 0
        checked = 0
        for i in range(min(n, 12)):
            im = imgs.nth(i)
            try:
                await im.wait_for(state="attached", timeout=2000)
                nw = await im.evaluate("e => e.naturalWidth || 0")
                checked += 1
                complete = await im.evaluate("e => e.complete")
                if complete and int(nw or 0) == 0:
                    broken += 1
            except Exception:
                continue
        if broken > max(1, checked // 4):
            return ("FAIL", f"抽检 {checked} 张图，{broken} 张 complete 且 naturalWidth=0（疑似裂图）。")
        return ("PASS", f"卡片区 img 约 {n} 个；抽检 {checked} 张，典型裂图 {broken}。")
    except Exception as e:
        return ("FAIL", f"游戏卡片抽检异常：{_brief_exc(e)}")


async def _p0_lobby_seems_visible(page: Any) -> bool:
    """大厅：Play Now 或分类 tablist 等。"""
    try:
        if await page.get_by_role("button", name=re.compile(r"Play\s*Now", re.I)).count() >= 1:
            return True
        if await page.locator('[role="tablist"]').count() >= 1:
            return True
    except Exception:
        pass
    return False


async def _p0_wait_entered_game_shell(page: Any, *, timeout_ms: int = 55_000) -> tuple[bool, str]:
    """
    进局后 URL/首段正文常不变（壳内加载）；以 KK 浮标、Exit、Guest 抬头、加载百分比等判定已进入游戏。
    """
    deadline = time.monotonic() + timeout_ms / 1000.0
    guest = page.get_by_text(re.compile(r"Guest[_A-Z0-9]+", re.I))
    pct = page.get_by_text(re.compile(r"\d{1,3}\s*%"))
    while time.monotonic() < deadline:
        for hint, sel in (
            ("KK 浮标/菜单柄", 'img[class*="fab_handle_img"], img[src*="logo-ball"]'),
            ("Exit 图标", 'img[alt="Exit"]'),
            (
                "退出资源图",
                'img[src*="exit"][class*="content_item_icon"], img[src*="/assets/exit"]',
            ),
        ):
            loc = page.locator(sel)
            try:
                if await loc.count() >= 1:
                    if await loc.first.is_visible(timeout=400):
                        return True, f"检测到游戏壳：{hint}"
            except Exception:
                pass
        try:
            if await guest.count() >= 1 and await guest.first.is_visible(timeout=200):
                return True, "检测到局内 Guest 文案"
        except Exception:
            pass
        try:
            if await pct.count() >= 1 and await pct.first.is_visible(timeout=200):
                t = await pct.first.inner_text()
                if t and re.search(r"[6-9]\d\s*%|100\s*%", t):
                    return True, f"检测到高进度加载：{t.strip()[:20]}"
        except Exception:
            pass
        await page.wait_for_timeout(420)
    return False, f"{timeout_ms}ms 内未检测到游戏壳（KK/Exit/局内特征）"


async def _p0_exit_game_via_kk_then_exit(page: Any) -> tuple[bool, str]:
    """
    游戏内：先点右下角 KK 浮标（展开），再点 alt=Exit / 退出图（与产品 UI 一致）。
    """

    async def _try_click_exit() -> bool:
        candidates = (
            page.locator('img[alt="Exit"]'),
            page.locator('img[class*="content_item_icon"][src*="exit"]'),
            page.locator('img[src*="/assets/exit"]'),
        )
        for loc in candidates:
            try:
                if await loc.count() < 1:
                    continue
                await loc.first.wait_for(state="visible", timeout=2800)
                await loc.first.click(timeout=6000, force=True)
                return True
            except Exception:
                continue
        return False

    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "Exit 已可见，直接点击退出"

    kk = page.locator('img[class*="fab_handle_img"], img[src*="logo-ball"]').first
    try:
        await kk.wait_for(state="visible", timeout=15_000)
    except Exception as e:
        return False, f"未找到 KK 浮标：{_brief_exc(e)}"
    try:
        await kk.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    async def _tap_kk() -> None:
        try:
            await kk.click(timeout=6000, force=True)
        except Exception:
            await kk.evaluate("e => e.click()")

    await _tap_kk()
    await page.wait_for_timeout(650)

    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "已点 KK 展开菜单并点击 Exit"

    await _tap_kk()
    await page.wait_for_timeout(500)
    if await _try_click_exit():
        await page.wait_for_timeout(700)
        return True, "第二次点 KK 后出现 Exit 并已点击"

    return False, "展开菜单后仍未点到 Exit（img[alt=Exit] / 退出资源图）"


def _filter_console_errors(bucket: list[str]) -> list[str]:
    out: list[str] = []
    for x in bucket:
        if _is_benign_console_line(x):
            continue
        out.append(x)
    return out


async def _run_p0_console_clean(bucket: list[str]) -> tuple[str, str]:
    bad = _filter_console_errors(bucket)
    if bad:
        return ("FAIL", "Console error（过滤后仍剩）：" + " | ".join(bad[:5]))
    return ("PASS", "Console error 抽样无严重项（已过滤部分已知噪声）。")


async def _run_p0_release_checklist() -> tuple[str, str]:
    return ("SKIP", "本次更新点需对照发布说明人工勾选（脚本不内置变更清单）。")


async def _run_p0_case(
    case_id: str,
    page: Any,
    *,
    log: Callable[[str], None],
    target_url: str,
    console_bucket: list[str],
    p0_play_now_really_click: bool = False,
) -> tuple[str, str]:
    if case_id == "p0_env_access":
        return ("PASS", "已在 _async_main 中完成 host/goto 校验。")
    if case_id == "p0_home_load":
        return await _run_p0_home_load(page)
    if case_id == "p0_page_title":
        return await _run_p0_page_title(page)
    if case_id == "p0_play_now":
        return await _run_p0_play_now(
            page,
            really_click=p0_play_now_really_click,
            target_url=target_url,
            log=log,
        )
    if case_id == "p0_category_tabs":
        return await _run_p0_category_tabs(page)
    if case_id == "p0_game_cards":
        return await _run_p0_game_cards(page)
    if case_id == "p0_console_clean":
        return await _run_p0_console_clean(console_bucket)
    if case_id == "p0_release_checklist":
        return await _run_p0_release_checklist()
    return ("BLOCKED", f"未知用例：{case_id}")

async def _try_click_bottom_label_js(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    tag: str = "底栏·JS",
) -> tuple[bool, str]:
    """KalaroKo 底栏：对精确文案 .last 做 JS 父级 click（与 Playwright 点 div 标签层分离）。"""
    def lg(m: str) -> None:
        if log:
            log(f"  [诊断·{tag}] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = (fr.url or "")[:120]
        except Exception:
            pass
        try:
            if "about:blank" in (fr.url or "").lower() and fi > 0:
                continue
        except Exception:
            pass
        try:
            n = await fr.get_by_text(label, exact=True).count()
            if n < 1:
                continue
            loc = fr.get_by_text(label, exact=True).last
            await loc.wait_for(state="attached", timeout=4000)
            try:
                await loc.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            info = await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
            lg(f"frame[{fi}]「{label}」.last JS → {info} | {furl!r}")
            await page.wait_for_timeout(450)
            return True, f"frame[{fi}] JS 底栏「{label}」"
        except Exception as e:
            lg(f"frame[{fi}]「{label}」：{_brief_exc(e, 160)}")
    return False, ""


async def _log_tablist_snapshot(page: Any, log: Callable[[str], None], *, tag: str = "") -> None:
    """打印当前页 URL、frame 数、每个 tablist 的 tab 数量与前几项文案（便于区分底栏与筛选条）。"""
    prefix = f"  [诊断{tag}] " if tag else "  [诊断] "
    try:
        log(f"{prefix}page.url = {page.url}")
    except Exception:
        log(f"{prefix}page.url = （无法读取）")
    try:
        frames = page.frames
        log(f"{prefix}frames 数量 = {len(frames)}（含主文档与 iframe）")
    except Exception as e:
        log(f"{prefix}frames = （{_brief_exc(e)}）")
    try:
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        log(f'{prefix}[role="tablist"] 数量 = {nl}')
        for idx in range(nl):
            tl = lists.nth(idx)
            tabs = tl.locator('[role="tab"]')
            tc = await tabs.count()
            parts: list[str] = []
            for j in range(min(tc, 8)):
                try:
                    raw = (await tabs.nth(j).inner_text() or "").strip()
                    one = " ".join(raw.split())[:48]
                    if len(raw) > 48:
                        one += "…"
                    parts.append(f"[{j}]「{one}」")
                except Exception as e:
                    parts.append(f"[{j}]（读文案失败 {_brief_exc(e, 80)}）")
            log(f"{prefix}  tablist[{idx}] → {tc} 个 tab: " + ("；".join(parts) if parts else "（无子 tab）"))
    except Exception as e:
        log(f"{prefix}枚举 tablist 失败：{_brief_exc(e)}")

    for role_name, label in (
        ("tab", "Home"),
        ("tab", "Party"),
    ):
        try:
            n = await page.get_by_role(role_name, name=re.compile(rf"^{re.escape(label)}$", re.I)).count()
            log(f'{prefix}get_by_role({role_name}, name=^{label}$) 匹配数 = {n}')
        except Exception as e:
            log(f"{prefix}get_by_role 统计 {label} 失败：{_brief_exc(e, 120)}")

    try:
        n_link = await page.get_by_role("link", name=re.compile(r"Party\s*Hubs", re.I)).count()
        log(f"{prefix}get_by_role(link, Party Hubs…) 匹配数 = {n_link}")
    except Exception as e:
        log(f"{prefix}统计 Party Hubs 链接失败：{_brief_exc(e, 120)}")


async def _visible_any(
    page: Any, patterns: list[str], *, timeout_ms: float = 2500
) -> tuple[bool, str]:
    """任一 Playwright get_by_text(re) 可见即 True。"""
    for pat in patterns:
        try:
            loc = page.get_by_text(re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=int(timeout_ms))
            return True, f"页面上有可见文字，匹配规则 /{pat}/i"
        except Exception:
            continue
    return False, (
        "页面上未找到与下列文案匹配的可见元素："
        + "、".join(patterns[:8])
        + "（若站点用语不同，需在脚本里补充关键词）"
    )


# 纯图标客服（无文案）：img 文件名 / class / title 等常见命名（KalaroKo 右上角耳机按钮等）
_ICON_MARKERS_SRC_ALT = (
    "headset",
    "Headset",
    "headphone",
    "earphone",
    "livechat",
    "LiveChat",
    "live-chat",
    "kefu",
    "customer-service",
    "CustomerService",
    "customer_service",
    "help-center",
    "HelpCenter",
)
_ICON_CLASS_MARKERS = (
    "headset",
    "Headset",
    "headphone",
    "customer-service",
    "CustomerService",
    "customer_service",
    "livechat",
    "LiveChat",
    "kefu",
    "Kefu",
    "kf-",
)


def _css_pure_icon_customer_service() -> str:
    parts: list[str] = []
    for m in _ICON_MARKERS_SRC_ALT:
        parts.append(f"img[src*='{m}']")
        parts.append(f"img[srcset*='{m}']")
        parts.append(f"img[alt*='{m}']")
        parts.append(f"button:has(img[src*='{m}'])")
        parts.append(f"[role='button']:has(img[src*='{m}'])")
        parts.append(f"div[role='button']:has(img[src*='{m}'])")
        parts.append(f"a:has(img[src*='{m}'])")
    for m in _ICON_CLASS_MARKERS:
        parts.append(f"button[class*='{m}']")
        parts.append(f"[role='button'][class*='{m}']")
        parts.append(f"a[class*='{m}']")
        parts.append(f"button:has(img[class*='{m}'])")
    parts.append("img[alt*='客服']")
    parts.extend(
        [
            "button[title*='客服']",
            "[role='button'][title*='客服']",
            "button[title*='Headset']",
            "button[title*='headset']",
            "[role='button'][title*='Headset']",
            "[role='button'][title*='headset']",
            "button[title*='Customer']",
            "button[title*='customer']",
            "[role='button'][title*='Customer']",
            "[role='button'][title*='customer']",
            "button[title*='Service']",
            "button[title*='service']",
            "[role='button'][title*='Service']",
            "[role='button'][title*='service']",
        ]
    )
    return ", ".join(parts)


async def _customer_service_on_frame(frame: Any, *, in_subframe: bool) -> tuple[bool, str]:
    """在单个 Page 或 Frame 内检测客服入口。"""
    suffix = "（检测于页面内嵌 iframe）" if in_subframe else ""

    text_patterns = [
        r"Customer\s*Service",
        r"客服",
        r"Support",
        r"Help\s*Center",
        r"Contact(\s+Us)?",
        r"Live\s*Chat",
        r"在线客服",
        r"联系客服",
        r"联系我们",
        r"Feedback",
        r"帮助",
        r"Help",
        r"技术支持",
        r"售后服务",
        r"Assist",
        r"Inquiry",
    ]
    ok, msg = await _visible_any(frame, text_patterns, timeout_ms=1800)
    if ok:
        return True, f"可见文案命中：{msg}{suffix}"

    name_pat = re.compile(
        r"service|support|help|客服|contact|chat|feedback|assist|inquiry|联系我们",
        re.I,
    )
    for role in ("button", "link"):
        try:
            loc = frame.get_by_role(role, name=name_pat).first
            await loc.wait_for(state="visible", timeout=1200)
            return True, f"找到可见的「{role}」，名称含客服相关关键词{suffix}"
        except Exception:
            pass

    try:
        loc = frame.get_by_alt_text(
            re.compile(
                r"headset|headphone|customer\s*service|support|help|客服|在线|联系|售后",
                re.I,
            )
        ).first
        await loc.wait_for(state="visible", timeout=1200)
        return True, f"通过 img 的 alt 文案命中疑似客服图标{suffix}"
    except Exception:
        pass

    attr_sel = (
        "a[href*='support'], a[href*='help'], a[href*='chat'], a[href*='customer'], "
        "a[href*='service'], a[href*='feedback'], a[href*='contact'], "
        "[role='button'][aria-label*='service'], [role='button'][aria-label*='support'], "
        "[role='button'][aria-label*='help'], [role='button'][aria-label*='chat'], "
        "[role='button'][aria-label*='客服'], [aria-label*='客服'], "
        "[data-testid*='support'], [data-testid*='service'], [data-testid*='help'], "
        "[data-testid*='chat'], [data-testid*='customer']"
    )
    try:
        loc = frame.locator(attr_sel).first
        await loc.wait_for(state="visible", timeout=1800)
        return True, f"通过链接 href 或 aria-label / data-testid 命中疑似客服入口{suffix}"
    except Exception:
        pass

    try:
        flo = frame.locator(
            "[class*='float'][class*='service'], [class*='float'][class*='support'], "
            "[id*='chat-widget'], [id*='customer-service']"
        ).first
        await flo.wait_for(state="visible", timeout=600)
        return True, f"命中浮动区/挂件选择器（疑似在线客服）{suffix}"
    except Exception:
        pass

    try:
        loc = frame.locator(_css_pure_icon_customer_service()).first
        await loc.wait_for(state="visible", timeout=2200)
        return True, f"命中纯图标客服入口（img src/alt、button:has(img) 或 class/title 特征）{suffix}"
    except Exception:
        pass

    try:
        hs = frame.locator("button:has(svg), [role='button']:has(svg)").filter(
            has=frame.locator(
                "[class*='headset'], [class*='Headset'], [class*='headphone'], "
                "[class*='customer'], [class*='service'], [class*='support'], [class*='kefu']"
            )
        )
        loc = hs.first
        await loc.wait_for(state="visible", timeout=1000)
        return True, f"命中带 SVG 的按钮且子节点 class 含客服相关特征{suffix}"
    except Exception:
        pass

    return False, ""


async def _customer_service_detect(page: Any) -> tuple[bool, str]:
    """
    客服入口：主文档 + 各 iframe 依次检测（文案 → 角色 → img alt →
    href/aria-label/data-testid → 纯图标/img 包裹按钮 → SVG 按钮）。
    """
    frames = list(page.frames)
    main = page.main_frame
    order = [main] + [f for f in frames if f != main]
    for fr in order:
        sub = fr != main
        ok, msg = await _customer_service_on_frame(fr, in_subframe=sub)
        if ok:
            return True, msg
    return (
        False,
        "主页面与所有 iframe 内均未找到客服入口（文案、无障碍名称、href/aria、"
        "纯图标 img/src/class、SVG 按钮等已轮询）。若仍失败，请在开发者工具中查看该节点 DOM 后发我补充选择器。",
    )


# —— P1 客服：仅主文档顶栏/视口右上打开，全 frame 关 Garden（含 about:blank 子帧）——

_P1_CS_IMG_CLICK_JS = """(el) => {
  const t = el.closest('button, a, [role="button"], [onclick]') || el.parentElement;
  if (t && t !== el) { t.click(); return { via: 'closest' }; }
  el.click();
  return { via: 'self' };
}"""


async def _p1_click_customer_service_top_right(
    page: Any, log: Callable[[str], None] | None
) -> tuple[bool, str]:
    """首页主文档：header 内耳机图 → 视口右上几何命中 → 全文 alt 兜底（不点 iframe 内）。"""
    fr = page.main_frame
    alt_re = re.compile(
        r"headset|headphone|customer|service|support|help|客服|在线|联系|售后",
        re.I,
    )
    try:
        hdr = fr.locator("header").first
        if await hdr.count() > 0:
            im = hdr.get_by_alt_text(alt_re).first
            if await im.count() > 0:
                await im.wait_for(state="attached", timeout=3200)
                try:
                    await im.scroll_into_view_if_needed()
                except Exception:
                    pass
                try:
                    await im.click(timeout=4000, force=True)
                except Exception:
                    await im.evaluate(_P1_CS_IMG_CLICK_JS)
                return True, "已点击 header 内客服图标（img alt）"
    except Exception as e:
        if log:
            log(f"  [诊断·客服] header+alt：{_brief_exc(e, 100)}")

    try:
        clicked = await fr.evaluate("""() => {
          const W = window.innerWidth, H = window.innerHeight;
          const imgs = [];
          document.querySelectorAll('header img').forEach(i => imgs.push(i));
          document.querySelectorAll('[class*="Header"] img, [class*="header"] img').forEach(i => {
            if (!imgs.includes(i)) imgs.push(i);
          });
          for (const img of imgs) {
            const r = img.getBoundingClientRect();
            if (r.width < 6 || r.height < 6) continue;
            const cx = r.left + r.width / 2, cy = r.top + r.height / 2;
            if (cx > W * 0.46 && cy < H * 0.45) {
              const t = img.closest('button, a, [role="button"]');
              (t || img).click();
              return true;
            }
          }
          return false;
        }""")
        if clicked:
            return True, "已点击视口右上 header 区域图标（几何命中）"
    except Exception as e:
        if log:
            log(f"  [诊断·客服] 几何命中：{_brief_exc(e, 100)}")

    try:
        loc = fr.get_by_alt_text(alt_re).first
        if await loc.count() > 0:
            await loc.wait_for(state="attached", timeout=2800)
            try:
                await loc.click(timeout=4000, force=True)
            except Exception:
                await loc.evaluate(_P1_CS_IMG_CLICK_JS)
            return True, "已点击主文档客服图标（全文 alt 兜底）"
    except Exception:
        pass
    return False, ""


async def _p1_frame_has_zendesk_widget(page: Any) -> bool:
    """任意 frame 出现 Garden 收起图标或典型聊天 UI 即视为已加载。"""
    for fr in page.frames:
        try:
            if await fr.locator('svg[data-garden-id="buttons.icon"]').count() > 0:
                return True
        except Exception:
            pass
        for pat in (
            r"Type\s*a\s*message",
            r"Kalaro\s*Bot",
            r"Privacy\s*Notice",
            r"Zendesk",
        ):
            try:
                if await fr.get_by_text(re.compile(pat, re.I)).count() > 0:
                    return True
            except Exception:
                pass
    return False


async def _p1_click_zendesk_garden_close(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """
    点击 Zendesk Garden 向下收起：svg[data-garden-id="buttons.icon"]。
    不跳过 about:blank（Messenger 子帧常见）；每轮遍历当前所有 frame。
    """
    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:140]
        except Exception:
            pass
        try:
            btn = fr.locator(
                'button:has(svg[data-garden-id="buttons.icon"]), '
                '[role="button"]:has(svg[data-garden-id="buttons.icon"])'
            ).last
            if await btn.count() > 0:
                await btn.click(timeout=2800, force=True)
                if log:
                    log(f"  [诊断·客服] 已点 Garden 收起 button frame[{fi}] {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception:
            pass
        try:
            svg = fr.locator('svg[data-garden-id="buttons.icon"]').last
            if await svg.count() > 0:
                await svg.evaluate(
                    """(el) => {
                      const b = el.closest('button, [role="button"]');
                      (b || el.parentElement || el).click();
                    }"""
                )
                if log:
                    log(f"  [诊断·客服] 已点 Garden svg→父级 frame[{fi}] {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception:
            pass
    return False


async def _p1_try_zendesk_header_last_in_chat_frames(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """在已出现聊天特征的 frame 内点 header 最后一颗按钮（⋮/收起旁）。"""
    for fi, fr in enumerate(page.frames):
        try:
            in_chat = False
            if await fr.locator('svg[data-garden-id="buttons.icon"]').count() > 0:
                in_chat = True
            if not in_chat and await fr.get_by_text(
                re.compile(r"Type\s*a\s*message", re.I)
            ).count() > 0:
                in_chat = True
            if not in_chat and await fr.get_by_text(
                re.compile(r"Kalaro\s*Bot", re.I)
            ).count() > 0:
                in_chat = True
            if not in_chat:
                continue
            hdr = fr.locator("[class*='header'], header").first
            if await hdr.count() < 1:
                continue
            btns = hdr.locator("button, [role='button']")
            bn = await btns.count()
            if bn < 1:
                continue
            await btns.nth(bn - 1).click(timeout=2600, force=True)
            if log:
                log(f"  [诊断·客服] 已点聊天窗顶栏最后按钮 frame[{fi}]")
            await page.wait_for_timeout(350)
            return True
        except Exception:
            continue
    return False


async def _try_spa_header_back(page: Any) -> bool:
    """SPA 顶栏返回（图2 红框）：主文档 header 内首颗按钮或 history.back。"""
    fr = page.main_frame
    locators: list[Any] = [
        fr.locator("header").locator("button, [role='button'], a").first,
        fr.locator("[class*='Header']").locator("button, [role='button']").first,
        fr.locator("[class*='header']").locator("button, [role='button']").first,
        fr.locator("[class*='navbar']").locator("button, [role='button']").first,
        fr.locator("[class*='NavBar']").locator("button, [role='button']").first,
        fr.get_by_role("button", name=re.compile(r"back", re.I)).first,
        fr.locator("[aria-label*='返回'], [aria-label*='back' i]").first,
    ]
    for loc in locators:
        try:
            if await loc.count() < 1:
                continue
            await loc.wait_for(state="attached", timeout=1800)
            await loc.click(timeout=4500, force=True)
            await page.wait_for_timeout(450)
            return True
        except Exception:
            continue
    try:
        ok = await fr.evaluate("""() => {
          const sels = ['header button', '[class*="Header"] button', '[class*="header"] button',
            '[class*="navbar"] button', '[class*="NavBar"] button', '[class*="title-bar"] button'];
          for (const s of sels) {
            const el = document.querySelector(s);
            if (el) { el.click(); return true; }
          }
          return false;
        }""")
        await page.wait_for_timeout(450)
        if ok:
            return True
    except Exception:
        pass
    try:
        await page.go_back(wait_until="domcontentloaded", timeout=15_000)
        await page.wait_for_timeout(450)
        return True
    except Exception:
        return False


async def _leave_party_hubs_to_home(
    page: Any,
    target_url: str,
    log: Callable[[str], None] | None,
) -> None:
    """离开 Party Hubs 子页，恢复底栏（图2 顶栏返回，失败则 goto /）。"""
    u = (page.url or "").lower()
    if "party-hubs" not in u:
        return
    if log:
        log("  [诊断] 当前在 Party Hubs，顶栏返回大厅…")
    await _try_spa_header_back(page)
    u2 = (page.url or "").lower()
    if "party-hubs" in u2:
        if log:
            log(f"  [诊断] 仍停留在 party-hubs（{page.url!r}），goto 站点根路径")
        home = _home_feed_url(target_url)
        await page.goto(home, wait_until="domcontentloaded", timeout=60_000)
        await page.wait_for_timeout(500)


# Party Hubs 子页：不滚底，仅用顶栏与列表区文案做轻量校验
_P1_PARTY_HUBS_VERIFY_PATTERNS = [
    r"Follow",
    r"All",
    r"Preparing",
    r"准备",
    r"In\s*Game",
    r"Random\s*Match",
    r"Create\s*a\s*Party",
    r"Party",
    r"Guest",
]


async def _try_party_hubs_top_bar_back(
    page: Any, log: Callable[[str], None] | None
) -> bool:
    """
    Party Hubs 顶栏左侧返回：KalaroKo 为 img._top_bar_item_icon_*（内嵌左箭头 SVG）。
    """
    def lg(m: str) -> None:
        if log:
            log("  [诊断·HottestParty] " + m)

    fr = page.main_frame
    sels = (
        'img[class*="_top_bar_item_icon_"]',
        'img[class*="top_bar_item_icon"]',
        r'img[src*="15.0005"]',
        r'img[src*="M15.0005"]',
    )
    for sel in sels:
        try:
            loc = fr.locator(sel).first
            if await loc.count() < 1:
                continue
            await loc.wait_for(state="attached", timeout=2200)
            try:
                await loc.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass
            try:
                await loc.click(timeout=3500, force=True)
            except Exception:
                await loc.evaluate(
                    """(el) => {
                      const p = el.closest('button, a, [role="button"]');
                      (p || el.parentElement || el).click();
                    }"""
                )
            lg(f"已点击顶栏返回图标（{sel!r}）")
            await page.wait_for_timeout(450)
            return True
        except Exception as e:
            lg(f"顶栏返回 {sel!r}：{_brief_exc(e, 100)}")
    return False


async def _p1_leave_party_hubs_after_hottest(
    page: Any,
    target_url: str,
    log: Callable[[str], None] | None,
) -> None:
    """先点顶栏 img 返回，仍在 party-hubs 则走通用 back/goto。"""
    if "party-hubs" not in (page.url or "").lower():
        return
    await _try_party_hubs_top_bar_back(page, log)
    if "party-hubs" in (page.url or "").lower():
        await _leave_party_hubs_to_home(page, target_url, log)


async def _customer_service_click_open_close(
    page: Any, log: Callable[[str], None] | None
) -> tuple[str, str]:
    """
    P1 客服：主文档顶栏/右上打开 → 轮询全部 frame（含 about:blank）点
    svg[data-garden-id="buttons.icon"] 收起 → 兜底顶栏最后键 → Esc。
    """
    ok, cm = await _p1_click_customer_service_top_right(page, log)
    if not ok:
        det = await _customer_service_detect(page)
        return (
            "FAIL",
            "未能点击首页客服入口。"
            + (f"（页面上曾可见：{det[1]}）" if det[0] else f" {det[1]}"),
        )
    if log:
        log(f"  [诊断·客服] {cm}")
    await page.wait_for_timeout(500)

    saw_widget = False
    closed_ok = False
    deadline = time.monotonic() + 16.0
    while time.monotonic() < deadline:
        saw_widget = saw_widget or await _p1_frame_has_zendesk_widget(page)
        if await _p1_click_zendesk_garden_close(page, log):
            closed_ok = True
            break
        await asyncio.sleep(0.16)

    if not closed_ok:
        if log:
            log("  [诊断·客服] 轮询内未点到 Garden，尝试聊天窗 header 最后按钮…")
        closed_ok = await _p1_try_zendesk_header_last_in_chat_frames(page, log)

    if not closed_ok:
        for _ in range(7):
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(160)
            except Exception:
                break
        if saw_widget:
            closed_ok = True

    if closed_ok:
        load_note = "已检测到聊天窗特征" if saw_widget else "关闭动作已执行"
        return (
            "PASS",
            f"{cm}；{load_note}；已通过 Garden svg / 顶栏按钮 / Esc 收起。继续后续用例。",
        )
    return (
        "FAIL",
        f"{cm}；未检测到聊天窗或未点中 "
        f'svg[data-garden-id="buttons.icon"]（当前 frame 数 {len(page.frames)}）。',
    )


async def _try_click_tab(
    page: Any, patterns: list[str], *, timeout_ms: float = 4000
) -> tuple[bool, str]:
    for pat in patterns:
        try:
            loc = page.get_by_role("tab", name=re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=2000)
            await loc.click(timeout=int(timeout_ms))
            await page.wait_for_timeout(800)
            return True, f"已点击「页签」角色，名称匹配 /{pat}/i"
        except Exception:
            pass
        try:
            loc = page.locator("a,button,[role='tab']").filter(has_text=re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=2000)
            await loc.click(timeout=int(timeout_ms))
            await page.wait_for_timeout(800)
            return True, f"已点击导航控件，文案匹配 /{pat}/i"
        except Exception:
            pass
    return False, "未点到对应页签，尝试过：" + "、".join(patterns)


async def _try_click_visible_text(
    page: Any, patterns: list[str], *, timeout_ms: float = 2000, settle_ms: float = 700
) -> tuple[bool, str]:
    """对可见文案执行 scroll + click（用于分区标题、非标准 tab 等）。"""
    for pat in patterns:
        try:
            loc = page.get_by_text(re.compile(pat, re.I)).first
            await loc.wait_for(state="visible", timeout=int(timeout_ms))
            await loc.scroll_into_view_if_needed()
            await loc.click(timeout=4500)
            await page.wait_for_timeout(int(settle_ms))
            return True, f"已点击匹配 /{pat}/i 的可见文案（scroll+click）"
        except Exception:
            continue
    return False, "未点击到任何匹配文案：" + "、".join(patterns[:8])


async def _try_click_exact_text_last_in_frames(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    diag_tag: str,
) -> tuple[bool, str]:
    """
    无 ARIA tablist 的 H5 底栏：各 frame 内精确匹配文案，取 .last（DOM 中固定底栏往往在后）。
    """
    def lg(m: str) -> None:
        if log:
            log(f"  [诊断·{diag_tag}] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = fr.url or ""
        except Exception:
            pass
        try:
            n = await fr.get_by_text(label, exact=True).count()
        except Exception as e:
            lg(f"frame[{fi}] {furl!r} 统计「{label}」失败：{_brief_exc(e, 120)}")
            continue
        if n < 1:
            continue
        try:
            lg(f"frame[{fi}] {furl!r} 精确「{label}」×{n} → .last + JS 父级 click")
            loc = fr.get_by_text(label, exact=True).last
            await loc.wait_for(state="attached", timeout=3200)
            try:
                await loc.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
            await page.wait_for_timeout(700)
            return True, f"frame[{fi}] 精确「{label}」.last（JS）"
        except Exception as e:
            lg(f"frame[{fi}] 点击失败：{_brief_exc(e, 200)}")
    return False, ""


async def _try_click_party_in_scoped_bottom_bar(
    page: Any, *, log: Callable[[str], None] | None
) -> tuple[bool, str]:
    """底栏容器同时含 Home + Profile（或 Task）时，在其中点「Party」，避免点到筛选条。"""
    def lg(m: str) -> None:
        if log:
            log("  [诊断·Party底栏·容器] " + m)

    for fi, fr in enumerate(page.frames):
        furl = ""
        try:
            furl = fr.url or ""
        except Exception:
            pass
        for shell_sel in (
            "nav",
            "footer",
            "[class*='tabbar' i]",
            "[class*='TabBar']",
            "[class*='tab-bar' i]",
            "[class*='bottomNav' i]",
            "[class*='BottomNav' i]",
        ):
            try:
                shell = fr.locator(shell_sel).filter(
                    has=fr.get_by_text("Home", exact=True)
                )
                n = await shell.count()
                if n < 1:
                    continue
                bar = None
                for anchor in ("Profile", "Task", "Share"):
                    cand = shell.filter(has=fr.get_by_text(anchor, exact=True)).last
                    if await cand.count() > 0:
                        bar = cand
                        break
                if bar is None:
                    bar = shell.last
                p = bar.get_by_text("Party", exact=True).first
                await p.wait_for(state="attached", timeout=2800)
                try:
                    await p.scroll_into_view_if_needed(timeout=2800)
                except Exception:
                    pass
                await p.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                await page.wait_for_timeout(800)
                return True, f"frame[{fi}] {shell_sel} 容器内「Party」（{furl!r}）"
            except Exception as e:
                lg(f"frame[{fi}] {shell_sel}：{_brief_exc(e, 160)}")
                continue
    return False, ""


async def _best_bottom_tabs(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[Any, int] | None:
    """
    取「主导航」tablist：优先 tab 数量最多的那一组；并列时取 DOM 中下标最大者（一般为底部固定栏）。
    用于区分顶/中部的筛选 tablist（项数往往较少）与底栏 Home/Party/Share…。
    """
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·主导航] " + msg)

    try:
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        if nl == 0:
            lg('页面中无 [role="tablist"]（底栏可能用 div+click 实现，无 ARIA tablist）')
            return None
        candidates: list[tuple[int, Any, int]] = []
        for idx in range(nl):
            tl = lists.nth(idx)
            nh = await tl.get_by_text("Home", exact=True).count()
            nshare = await tl.get_by_text("Share", exact=True).count()
            if nh < 1 or nshare < 1:
                lg(
                    f"tablist[{idx}] 跳过（非底栏：须同时含精确「Home」与「Share」；"
                    f"当前 Home={nh} Share={nshare}）"
                )
                continue
            tabs = tl.locator('[role="tab"]')
            tc = await tabs.count()
            if tc >= 2:
                candidates.append((idx, tabs, tc))
            lg(f"tablist[{idx}] 内 [role=tab] 数量 = {tc}（已确认含 Home+Share）")
        if not candidates:
            lg("无「同时含 Home+Share」的 tablist（中部筛选条已排除）；改用 JS 底栏或 nav 容器策略")
            return None
        max_tc = max(c[2] for c in candidates)
        best = max((c for c in candidates if c[2] == max_tc), key=lambda c: c[0])
        lg(f"选用 tab 数最多的一组：共 {best[2]} 项（并列时取 DOM 下标较大者 ≈ 底栏）")
        return (best[1], best[2])
    except Exception as e:
        lg(f"_best_bottom_tabs 异常：{_brief_exc(e)}")
        return None


async def _try_click_bottom_home_nav(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """底栏第一项 Home（须先回首页再滚到底才见 Hottest Parties）。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·Home] " + msg)

    reasons: list[str] = []
    lg("策略 0：各 frame 精确「Home」.last + JS 父级点击（KalaroKo）…")
    ok0, m0 = await _try_click_bottom_label_js(page, "Home", log=log, tag="Home")
    if ok0:
        return True, m0

    r = await _best_bottom_tabs(page, log=log)
    if r:
        tabs, tc = r
        try:
            lg(f"策略 A：点击主导航第 1 个 tab（共 {tc} 项）…")
            t0 = tabs.nth(0)
            await t0.wait_for(state="visible", timeout=3200)
            await t0.scroll_into_view_if_needed()
            await t0.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, f"已点击主导航 tablist 第 1 项（共 {tc} 项，假定 Home）"
        except Exception as e:
            reasons.append(f"策略A失败：{_brief_exc(e)}")
            lg(reasons[-1])
    else:
        reasons.append("无可用主导航 tablist（见上方 [诊断·主导航]）")

    pat = re.compile(r"^Home$", re.I)
    try:
        lg("策略 B：get_by_role(tab, name=^Home$)…")
        tabs = page.get_by_role("tab", name=pat)
        n = await tabs.count()
        lg(f"匹配数 = {n}")
        if n > 0:
            t = tabs.nth(n - 1)
            await t.wait_for(state="visible", timeout=2500)
            await t.scroll_into_view_if_needed()
            await t.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, f"已点击「Home」页签（role=tab，第 {n} 个匹配取末项）"
    except Exception as e:
        reasons.append(f"策略B失败：{_brief_exc(e)}")
        lg(reasons[-1])

    try:
        lg("策略 C：a/button/[role=tab] 文案 ^Home$…")
        row = page.locator("a,button,[role='tab']").filter(has_text=pat)
        n = await row.count()
        lg(f"匹配数 = {n}")
        if n > 0:
            b = row.nth(n - 1)
            await b.wait_for(state="visible", timeout=2500)
            await b.scroll_into_view_if_needed()
            await b.click(timeout=4500)
            await page.wait_for_timeout(550)
            return True, "已点击底部导航「Home」（a/button/tab，取末项）"
    except Exception as e:
        reasons.append(f"策略C失败：{_brief_exc(e)}")
        lg(reasons[-1])

    lg("策略 D：各 frame 精确「Home」.last（无 tablist 的底栏）…")
    ok_d, msg_d = await _try_click_exact_text_last_in_frames(
        page, "Home", log=log, diag_tag="Home·文案"
    )
    if ok_d:
        return True, msg_d + "（精确文案 .last）"

    summary = "；".join(reasons) if reasons else "未知原因"
    lg(f"全部策略失败，摘要：{summary}")
    return False, summary


async def _scroll_home_to_bottom(page: Any) -> None:
    """首页「Hottest Parties」在底部，先滚到底再点 Party Hubs / 底栏。"""
    try:
        await page.evaluate("window.scrollTo(0, 0)")
        await page.wait_for_timeout(150)
    except Exception:
        pass
    try:
        h = await page.evaluate(
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0, "
            "document.scrollingElement?.scrollHeight||0)"
        )
        if h is not None:
            await page.evaluate("y => window.scrollTo(0, y)", max(0, int(h)))
        await page.wait_for_timeout(400)
    except Exception:
        pass
    try:
        await page.keyboard.press("End")
        await page.wait_for_timeout(350)
    except Exception:
        pass


async def _try_click_party_hubs_in_hottest_row(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """「Hottest Parties」标题行与「Party Hubs >」同一块区域（图1 红框）：先滚入视口再点 Hub。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·PartyHubs] " + msg)

    try:
        row = page.locator("div, section, article, header").filter(
            has=page.get_by_text(re.compile(r"Hottest\s*Parties", re.I))
        ).filter(has=page.get_by_text(re.compile(r"Party\s*Hubs", re.I)))
        n = await row.count()
        if n < 1:
            lg("未找到同时含 Hottest Parties 与 Party Hubs 的容器")
            return False, ""
        box = row.last
        await box.scroll_into_view_if_needed(timeout=2800)
        await page.wait_for_timeout(180)
        hub = box.get_by_text(re.compile(r"Party\s*Hubs", re.I)).last
        await hub.wait_for(state="attached", timeout=2200)
        await hub.scroll_into_view_if_needed(timeout=2500)
        try:
            await hub.click(timeout=3500, force=True)
        except Exception:
            await hub.evaluate(
                """(el) => {
                  let n = el;
                  for (let i = 0; i < 10 && n; i++) {
                    const t = (n.tagName || '').toUpperCase();
                    if (t === 'A' || t === 'BUTTON' || n.getAttribute('role') === 'button') {
                      n.click(); return;
                    }
                    n = n.parentElement;
                  }
                  el.click();
                }"""
            )
        await page.wait_for_timeout(800)
        return True, "已点击 Hottest Parties 区块内 Party Hubs"
    except Exception as e:
        lg(f"Hottest 同行：{_brief_exc(e, 180)}")
        return False, ""


async def _try_click_party_hubs_link(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """右侧「Party Hubs >」多为 link；优先于整块标题点击。失败策略须短超时，避免 4×4×3s 级联拖分钟。"""
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·PartyHubs] " + msg)

    ok_row, msg_row = await _try_click_party_hubs_in_hottest_row(page, log=log)
    if ok_row:
        return True, msg_row

    # 可见链路与纯文案分支分开超时：无匹配则 count 为 0 立即跳过，不空等 3s
    t_vis = 1100
    t_att = 1800
    t_clk = 3200

    for pat in (
        r"Party\s*Hubs(?:\s*[>›])?",
        r"Hottest\s*Parties",
    ):
        try:
            gl = page.get_by_role("link", name=re.compile(pat, re.I))
            if await gl.count() < 1:
                lg(f"get_by_role(link) /{pat}/ count=0，跳过")
            else:
                loc = gl.first
                await loc.wait_for(state="visible", timeout=t_vis)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await loc.click(timeout=t_clk)
                await page.wait_for_timeout(450)
                return True, f"已点击链接（get_by_role link），匹配 /{pat}/i"
        except Exception as e:
            lg(f"get_by_role(link) /{pat}/ ：{_brief_exc(e, 140)}")
        try:
            al = page.locator("a").filter(has_text=re.compile(pat, re.I))
            if await al.count() < 1:
                lg(f"locator(a)+text /{pat}/ count=0，跳过")
            else:
                loc = al.first
                await loc.wait_for(state="visible", timeout=t_vis)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await loc.click(timeout=t_clk)
                await page.wait_for_timeout(450)
                return True, f"已点击 <a>，文案匹配 /{pat}/i"
        except Exception as e:
            lg(f"locator(a)+text /{pat}/ ：{_brief_exc(e, 140)}")
        try:
            tl = page.get_by_text(re.compile(pat, re.I))
            if await tl.count() < 1:
                lg(f"get_by_text /{pat}/ count=0，跳过")
            else:
                loc = tl.last
                await loc.wait_for(state="attached", timeout=t_att)
                await loc.scroll_into_view_if_needed(timeout=t_att)
                await page.wait_for_timeout(120)
                try:
                    await loc.click(timeout=t_clk, force=True)
                except Exception:
                    await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                await page.wait_for_timeout(450)
                return True, f"已点击文案 .last（force/JS），匹配 /{pat}/i"
        except Exception as e:
            lg(f"get_by_text .last /{pat}/ ：{_brief_exc(e, 140)}")
    lg("所有 Party Hubs / Hottest Parties 链接模式均未成功")
    return False, ""


async def _try_click_bottom_party_nav(
    page: Any, *, log: Callable[[str], None] | None = None
) -> tuple[bool, str]:
    """
    底部 TabBar 第二项「Party」。必须用主导航 tablist 的几何顺序，
    禁止优先 get_by_role(name=Party)：否则会点到内容区筛选里的「Party」。
    """
    def lg(msg: str) -> None:
        if log:
            log("  [诊断·Party底栏] " + msg)

    reasons: list[str] = []
    lg("策略 0：各 frame 精确「Party」.last + JS（底栏在 DOM 中通常后于筛选条）…")
    ok0, m0 = await _try_click_bottom_label_js(page, "Party", log=log, tag="Party底栏")
    if ok0:
        return True, m0

    r = await _best_bottom_tabs(page, log=log)
    if r:
        tabs, tc = r
        if tc >= 2:
            try:
                lg(f"策略 A：主导航 tablist 第 2 项（共 {tc} 项）…")
                t1 = tabs.nth(1)
                await t1.wait_for(state="visible", timeout=3200)
                await t1.scroll_into_view_if_needed()
                await t1.click(timeout=4500)
                await page.wait_for_timeout(900)
                return True, f"已点击主导航 tablist 第 2 项（共 {tc} 项，底部 Party）"
            except Exception as e:
                reasons.append(f"策略A失败：{_brief_exc(e)}")
                lg(reasons[-1])
        else:
            reasons.append(f"主导航仅 {tc} 项，无法取第 2 项作为 Party")
            lg(reasons[-1])
    else:
        reasons.append("无主导航 tablist（与 Home 相同根因）")

    try:
        lg("策略 B：最后一个 [role=tablist] 的第 2 项…")
        lists = page.locator('[role="tablist"]')
        nl = await lists.count()
        lg(f'[role="tablist"] 数量 = {nl}')
        if nl > 0:
            tablist = lists.nth(nl - 1)
            tabs2 = tablist.locator('[role="tab"]')
            tc = await tabs2.count()
            lg(f"最后一个 tablist 内 tab 数 = {tc}")
            if tc >= 2:
                t = tabs2.nth(1)
                await t.wait_for(state="visible", timeout=3200)
                await t.scroll_into_view_if_needed()
                await t.click(timeout=4500)
                await page.wait_for_timeout(900)
                return True, "已点击最后一个 tablist 的第 2 项（兜底，假定 Home, Party, …）"
    except Exception as e:
        reasons.append(f"策略B失败：{_brief_exc(e)}")
        lg(reasons[-1])

    pat_exact = re.compile(r"^Party$", re.I)
    try:
        lg("策略 C：get_by_role(tab, Party) 仅当匹配数>1 时取末项…")
        tabs = page.get_by_role("tab", name=pat_exact)
        n = await tabs.count()
        lg(f"匹配数 = {n}（≤1 则跳过以免点到筛选条唯一 Party）")
        if n > 1:
            t = tabs.nth(n - 1)
            await t.wait_for(state="visible", timeout=3200)
            await t.scroll_into_view_if_needed()
            await t.click(timeout=4500)
            await page.wait_for_timeout(900)
            return True, f"已点击「Party」页签（仅当匹配数>1 时用末项，共 {n} 个）"
    except Exception as e:
        reasons.append(f"策略C失败：{_brief_exc(e)}")
        lg(reasons[-1])

    try:
        lg("策略 D：a/button/[role=tab] 含 Party 且不含 Hubs/Hottest，匹配数>1 取末项…")
        base = page.locator("a,button,[role='tab']")
        row = base.filter(has_text=re.compile(r"Party", re.I)).filter(
            has_not_text=re.compile(r"Hubs|Hottest", re.I)
        )
        n = await row.count()
        lg(f"匹配数 = {n}")
        if n > 1:
            b = row.nth(n - 1)
            await b.wait_for(state="visible", timeout=3200)
            await b.scroll_into_view_if_needed()
            await b.click(timeout=4500)
            await page.wait_for_timeout(900)
            return True, "已点击「Party」（排除 Hubs/Hottest 且匹配数>1 时取末项）"
    except Exception as e:
        reasons.append(f"策略D失败：{_brief_exc(e)}")
        lg(reasons[-1])

    lg("策略 E：nav/footer 容器（含 Home+Profile/Task）内精确「Party」…")
    ok_e, msg_e = await _try_click_party_in_scoped_bottom_bar(page, log=log)
    if ok_e:
        return True, msg_e

    lg("策略 F：各 frame 精确「Party」.last（无 tablist 的底栏）…")
    ok_f, msg_f = await _try_click_exact_text_last_in_frames(
        page, "Party", log=log, diag_tag="Party·文案"
    )
    if ok_f:
        return True, msg_f + "（精确文案 .last）"

    summary = "；".join(reasons) if reasons else "未知原因"
    lg(f"全部策略失败，摘要：{summary}")
    return False, summary


async def _must_click_then_verify(
    page: Any,
    *,
    case_zh: str,
    tab_patterns: list[str],
    text_click_patterns: list[str],
    verify_patterns: list[str],
    verify_timeout_ms: float = 2800,
    bottom_js_label: str | None = None,
    log: Callable[[str], None] | None = None,
) -> tuple[str, str]:
    """
    与 ``test_k11_p1_skill_herontest_playwright.py`` 一致：
    先点击（底栏 JS → 页签 → 文案），再用全页 ``_visible_any`` 校验切换后可见内容。
    """
    clicked = False
    cm = ""
    if bottom_js_label:
        okb, mb = await _try_click_bottom_label_js(
            page, bottom_js_label, log=log, tag=case_zh
        )
        if okb:
            clicked, cm = True, mb
    if not clicked and tab_patterns:
        clicked, cm = await _try_click_tab(page, tab_patterns)
    if not clicked:
        c2, m2 = await _try_click_visible_text(page, text_click_patterns)
        if c2:
            clicked, cm = True, m2
    if not clicked:
        return (
            "FAIL",
            f"「{case_zh}」未能完成点击（页签与文案点击均失败），不允许仅凭背景可见文案判通过。详情：{cm}",
        )
    ok, vm = await _visible_any(page, verify_patterns, timeout_ms=verify_timeout_ms)
    if not ok:
        return (
            "FAIL",
            f"「{case_zh}」已操作（{cm}），但切换后未见预期内容：{vm}",
        )
    return ("PASS", f"「{case_zh}」{cm}；切换/展开后确认：{vm}")


async def _run_p1_case(
    case_id: str,
    page: Any,
    *,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    """返回 (VERDICT, 中文/可读说明)。"""
    if case_id == "p1_customer_service":
        return await _customer_service_click_open_close(page, log)
    if case_id == "p1_share_tab":
        return await _must_click_then_verify(
            page,
            case_zh="分享页签",
            tab_patterns=[r"^Share$", r"分享", r"Share"],
            text_click_patterns=[r"^Share$", r"\bShare\b", r"分享"],
            bottom_js_label="Share",
            log=log,
            verify_patterns=[
                r"Share",
                r"分享",
                r"invite",
                r"邀请",
                r"refer",
                r"friend",
                r"复制",
                r"Copy\s*link",
                r"Link",
                r"链接",
            ],
        )
    if case_id == "p1_task_tab":
        return await _must_click_then_verify(
            page,
            case_zh="任务页签",
            tab_patterns=[r"^Task$", r"任务", r"Tasks"],
            text_click_patterns=[r"^Task$", r"\bTask\b", r"任务"],
            bottom_js_label="Task",
            log=log,
            verify_patterns=[
                r"\bTask\b",
                r"任务",
                r"Quest",
                r"Daily",
                r"每日",
                r"Reward",
                r"奖励",
                r"Mission",
                r"Complete",
                r"进度",
            ],
        )
    if case_id == "p1_profile_tab":
        return await _must_click_then_verify(
            page,
            case_zh="我的/Profile",
            tab_patterns=[r"Profile", r"^Me$", r"我的", r"Account"],
            text_click_patterns=[r"Profile", r"^Me$", r"我的", r"Account"],
            bottom_js_label="Profile",
            log=log,
            verify_patterns=[
                r"Profile",
                r"我的",
                r"Wallet",
                r"钱包",
                r"Balance",
                r"余额",
                r"Setting",
                r"设置",
                r"VIP",
                r"Account",
                r"账户",
                r"Member",
                r"Logout",
                r"退出",
                r"Avatar",
                r"头像",
            ],
        )
    if case_id == "p1_hottest_parties":
        """
        首页仅滚底一次 → 点 Party Hubs → 子页不滚底，校验顶栏/列表常态文案
        → 点顶栏 img 返回 → 继续后续用例。

        本用例链式定位曾依赖 Playwright 默认 30s 超时，易长时间无新日志、像“卡死”；
        此处临时压到 10s，单步失败更快走降级策略。
        """
        _prev_tmo = 30_000
        try:
            _prev_tmo = int(page.get_default_timeout())
        except Exception:
            pass
        try:
            page.set_default_timeout(10_000)
            await _ensure_on_home_feed(page, target_url, log)
            log("  [诊断·HottestParty] —— 首页：底栏 Home + 滚至 Hottest / Party Hubs ——")
            await _log_tablist_snapshot(page, log, tag="·Home前")
            try:
                await page.evaluate("window.scrollTo(0, 0)")
                await page.wait_for_timeout(180)
            except Exception:
                pass
            home_ok, home_msg = await _try_click_bottom_home_nav(page, log=log)
            if home_ok:
                log(f"  [诊断·HottestParty] 底栏 Home：{home_msg}")
            else:
                log(f"  [诊断·HottestParty] 底栏 Home 未点到：{home_msg}")
            await page.wait_for_timeout(320)
            log("  [诊断·HottestParty] 仅在首页滚底（scrollHeight + End），以露出区块…")
            await _scroll_home_to_bottom(page)

            clicked, cm = await _try_click_party_hubs_link(page, log=log)
            if not clicked:
                c2, m2 = await _try_click_visible_text(
                    page,
                    [
                        r"Party\s*Hubs",
                        r"Hottest\s*Parties",
                        r"热门\s*Party",
                        r"Party\s*Hub",
                        r"热门",
                    ],
                    timeout_ms=2800,
                )
                if c2:
                    clicked, cm = True, m2
                else:
                    log(f"  [诊断·HottestParty] 文案点击未成功：{m2}")
            if not clicked and "party-hubs" in (page.url or "").lower():
                log("  [诊断·HottestParty] URL 已在 party-hubs，继续校验子页…")
                clicked, cm = True, "已进入 Party Hubs（URL）"
            if not clicked:
                await _log_tablist_snapshot(page, log, tag="·首页点 Hubs 失败后")
                await _p1_leave_party_hubs_after_hottest(page, target_url, log)
                hint = (
                    f" Home：{home_msg}。"
                    if home_ok
                    else f" Home 未点到：{home_msg}。"
                )
                return (
                    "FAIL",
                    "「热门 Party 板块」未点到 Party Hubs。"
                    + hint
                    + " 详见 [诊断·PartyHubs] 与 tablist 快照。",
                )

            await page.wait_for_timeout(550)
            ok, vm = await _visible_any(
                page,
                list(_P1_PARTY_HUBS_VERIFY_PATTERNS),
                timeout_ms=2800,
            )
            if not ok:
                await _p1_leave_party_hubs_after_hottest(page, target_url, log)
                return (
                    "FAIL",
                    f"「热门 Party 板块」已进入（{cm}），但 Party Hubs 页未见典型展示：{vm}",
                )

            log("  [诊断·HottestParty] 子页展示正常，点击顶栏返回图标离开…")
            await _p1_leave_party_hubs_after_hottest(page, target_url, log)

            h_part = f"{home_msg}；" if home_ok else "（Home 未点到仍进入 Hubs）"
            return (
                "PASS",
                f"「热门 Party 板块」{h_part}{cm}；子页确认：{vm}；已顶栏返回并回大厅。",
            )
        finally:
            try:
                page.set_default_timeout(_prev_tmo)
            except Exception:
                pass
    if case_id == "p1_party_status":
        await _ensure_on_home_feed(page, target_url, log)
        log("  [诊断] —— 快照：Party 用例开始（滚底前）——")
        await _log_tablist_snapshot(page, log, tag="·Party前")
        await _scroll_home_to_bottom(page)
        log("  [诊断] 已滚底，开始点击底栏 Party…")
        clicked, cm = await _try_click_bottom_party_nav(page, log=log)
        if not clicked:
            log("  [诊断] —— 快照：Party 底栏点击失败后 ——")
            await _log_tablist_snapshot(page, log, tag="·Party失败后")
            return (
                "FAIL",
                "「Party 状态展示」须点击底栏「Party」（脚本优先 JS 父级点击，其次含 Home+Share 的 tablist / nav）。"
                f"失败摘要：{cm}。"
                "详见上方 [诊断·Party底栏] 与 tablist 快照。",
            )
        ok, vm = await _visible_any(
            page,
            [
                r"Preparing",
                r"准备",
                r"players?",
                r"人数",
                r"Live",
                r"Party",
                r"Playing",
                r"进行中",
                r"Open",
                r"开局",
            ],
            timeout_ms=3200,
        )
        if not ok:
            return (
                "FAIL",
                f"「Party 状态展示」已操作（{cm}），但未见状态类文案：{vm}",
            )
        return ("PASS", f"「Party 状态展示」{cm}；确认：{vm}")
    return ("BLOCKED", f"脚本未实现该用例：{case_id}")


async def _log_bottom_nav_context(page: Any, log: Callable[[str], None], *, tag: str) -> None:
    """打印 viewport、scroll 与各 frame 内 Home/Party 精确匹配数量，便于定位底栏在哪个 frame、是否被判定不可见。"""
    log(f"  [诊断·底栏·{tag}] page.url = {page.url}")
    try:
        vp = await _eval_timeout(
            page,
            "() => ({ w: window.innerWidth, h: window.innerHeight, sy: window.scrollY, "
            "sh: Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0) })",
            timeout=8.0,
        )
        log(
            f"  [诊断·底栏·{tag}] viewport inner={vp.get('w')}×{vp.get('h')} "
            f"scrollY={vp.get('sy')} scrollHeight≈{vp.get('sh')}"
        )
    except Exception as e:
        log(f"  [诊断·底栏·{tag}] 读 viewport 失败：{_exc_tail(e, 120)}")
    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:140]
        except Exception:
            pass
        try:
            nh = await fr.get_by_text("Home", exact=True).count()
            np = await fr.get_by_text("Party", exact=True).count()
            nsh = await fr.get_by_text("Share", exact=True).count()
            nt = await fr.get_by_text("Task", exact=True).count()
            npr = await fr.get_by_text("Profile", exact=True).count()
            log(
                f"  [诊断·底栏·{tag}] frame[{fi}] url≈{fu!r} "
                f"exact: Home={nh} Party={np} Share={nsh} Task={nt} Profile={npr}"
            )
        except Exception as e:
            log(f"  [诊断·底栏·{tag}] frame[{fi}] 计数失败：{_exc_tail(e, 120)}")


async def _click_attached_force(
    loc: Any, *, timeout_ms: int = 8000, force: bool = True
) -> None:
    """不 wait visible（底栏常被判定为 hidden 但仍可 force 点）。"""
    await loc.wait_for(state="attached", timeout=min(6000, timeout_ms))
    try:
        await loc.scroll_into_view_if_needed(timeout=4000)
    except Exception:
        pass
    await loc.click(timeout=timeout_ms, force=force)


async def _scroll_to_bottom(page: Any) -> None:
    try:
        h = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0, "
            "document.scrollingElement?.scrollHeight||0)",
            timeout=12.0,
        )
        if h is not None:
            await _eval_timeout(
                page, "y => window.scrollTo(0, y)", max(0, int(h)), timeout=12.0
            )
        await page.wait_for_timeout(400)
    except (Exception, asyncio.TimeoutError):
        pass
    try:
        await asyncio.wait_for(page.keyboard.press("End"), timeout=5.0)
        await page.wait_for_timeout(350)
    except (Exception, asyncio.TimeoutError):
        pass


async def _scroll_tabbar_into_view(page: Any) -> None:
    """固定底栏常被遮住时，先滚到页面最底部。"""
    try:
        await _eval_timeout(
            page,
            "() => { const y = Math.max(document.body?.scrollHeight||0, "
            "document.documentElement.scrollHeight||0); window.scrollTo(0, y); }",
            timeout=10.0,
        )
        await page.wait_for_timeout(350)
    except (Exception, asyncio.TimeoutError):
        pass
def _mime_console_failures(console_lines: list[str]) -> list[str]:
    keys = (
        "MIME type",
        "module script",
        "Failed to load module",
        "text/html",
        "Strict MIME",
    )
    return [x for x in console_lines if any(k in x for k in keys)]


async def _broken_images_report(page: Any) -> tuple[int, list[str]]:
    """主文档 + 各 frame 统计 complete 且 naturalWidth==0 的 img src（抽样）。"""
    js = """() => {
      const imgs = Array.from(document.querySelectorAll('img'));
      const bad = [];
      for (const im of imgs) {
        try {
          if (im.complete && im.naturalWidth === 0 && (im.src || im.currentSrc)) {
            const s = (im.currentSrc || im.src || '').slice(0, 160);
            if (s) bad.push(s);
          }
        } catch (e) {}
      }
      return { total: imgs.length, bad: bad.slice(0, 20) };
    }"""

    total = 0
    bad_all: list[str] = []
    for fr in page.frames:
        try:
            r = await asyncio.wait_for(fr.evaluate(js), timeout=12.0)
            total += int(r.get("total") or 0)
            for s in r.get("bad") or []:
                if s not in bad_all:
                    bad_all.append(s)
        except (Exception, asyncio.TimeoutError):
            continue
    return total, bad_all[:25]


async def _ext_visible_text_pat(page: Any, pat: str, *, timeout_ms: float = 3500) -> bool:
    """扩展用例专用：单模式可见性布尔。勿与 P1 的 _visible_any(patterns→tuple) 同名，否则会覆盖导致解包失败。"""
    try:
        loc = page.get_by_text(re.compile(pat, re.I)).first
        await loc.wait_for(state="visible", timeout=int(timeout_ms))
        return True
    except Exception:
        return False


async def _run_ext_game_list(page: Any) -> tuple[str, str]:
    try:
        await page.mouse.wheel(0, 900)
        await page.wait_for_timeout(350)
    except Exception:
        pass
    patterns = [
        r"Tongits\s*King",
        r"Royal\s*Pusoy",
        r"Texas\s*Holdem",
        r"Bingo",
        r"Party",
    ]
    found: list[str] = []
    missing: list[str] = []
    for pat in patterns:
        ok = await _ext_visible_text_pat(page, pat, timeout_ms=2800)
        if ok:
            found.append(pat)
        else:
            missing.append(pat)
    if len(found) >= 3:
        return ("PASS", f"至少命中 {len(found)}/{len(patterns)} 组关键词：{', '.join(found)}")
    return (
        "FAIL",
        f"可见游戏/模块文案不足（需≥3）。命中：{found or '无'}；未命中：{', '.join(missing)}。"
        "请先在大厅首页并滚到游戏区。",
    )


async def _run_ext_images(page: Any) -> tuple[str, str]:
    total, bad = await _broken_images_report(page)
    if total == 0:
        return ("SKIP", "页面上未统计到 img 节点（可能图为 background 或 canvas）。")
    if bad:
        return (
            "FAIL",
            f"发现 {len(bad)} 个疑似裂图（complete 且 naturalWidth=0），共扫描约 {total} 个 img。"
            f" 示例：{bad[0][:120]}…" if bad else "",
        )
    return ("PASS", f"抽样检查：{total} 个 img 未发现典型裂图（naturalWidth=0）。")


async def _run_ext_static_console(console_bucket: list[str]) -> tuple[str, str]:
    bad = _mime_console_failures(console_bucket)
    if bad:
        return (
            "FAIL",
            "Console 存在与 MIME/模块脚本相关的错误（可能静态资源被 HTML 替代或缓存异常）："
            + " | ".join(bad[:3]),
        )
    if not console_bucket:
        return ("PASS", "抽样 Console error 中未见典型 MIME/模块脚本类报错。")
    return ("PASS", f"有 {len(console_bucket)} 条 Console error 抽样，但无 MIME/模块脚本关键字命中。")


async def _try_click_bottom_exact_in_frames(
    page: Any,
    label: str,
    *,
    log: Callable[[str], None] | None,
    force: bool = True,
) -> bool:
    """
    底栏点击多策略：0) JS 从 label div 向上找 item/button 再 click（KalaroKo）
    ① class 容器 Home+Share ② nav/footer ③ Playwright force click nth ④ role ⑤ a/button。
    """
    label_re = re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)

    def lg(m: str) -> None:
        if log:
            log("  [诊断·耗时] " + m)

    for fi, fr in enumerate(page.frames):
        fu = ""
        try:
            fu = (fr.url or "")[:130]
        except Exception:
            pass
        try:
            if "about:blank" in (fr.url or "").lower() and fi > 0:
                continue
        except Exception:
            pass

        # 策略 0：文案节点非可点击层，向上找父级 item / button / [role=tab] 再 DOM click
        try:
            nlab = await fr.get_by_text(label, exact=True).count()
            if nlab >= 1:
                loc0 = fr.get_by_text(label, exact=True).last
                await loc0.wait_for(state="attached", timeout=4000)
                info = await loc0.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                lg(f"frame[{fi}] 策略0 JS 底栏「{label}」→ {info} | {fu!r}")
                await page.wait_for_timeout(350)
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略0「{label}」：{_exc_tail(e)}")

        # 策略 A：常见 tabbar class + 同时含 Home 与 Share（主导航条）
        for shell_sel in BOTTOM_SHELL_SELECTORS:
            try:
                base = fr.locator(shell_sel)
                bc = await base.count()
                if bc == 0:
                    continue
                shell = base.filter(has=fr.get_by_text("Home", exact=True))
                if await shell.count() == 0:
                    lg(f"frame[{fi}] 策略A {shell_sel!r} 有{bc}个但无 Home 子树")
                    continue
                shell2 = shell.filter(has=fr.get_by_text("Share", exact=True))
                if await shell2.count() > 0:
                    bar = shell2.last
                else:
                    bar = shell.last
                loc = bar.get_by_text(label, exact=True).first
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略A 容器{shell_sel!r} 内 first「{label}」 ok | {fu!r}")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略A {shell_sel!r}「{label}」：{_exc_tail(e)}")

        # 策略 B：nav/footer + Profile/Task 锚定
        try:
            shell = fr.locator("nav, footer").filter(has=fr.get_by_text("Home", exact=True))
            cn = await shell.count()
            lg(f"frame[{fi}] 策略B nav/footer∩Home 容器数={cn}")
            if cn > 0:
                bar = shell.filter(has=fr.get_by_text("Profile", exact=True)).last
                if await bar.count() == 0:
                    bar = shell.filter(has=fr.get_by_text("Task", exact=True)).last
                if await bar.count() == 0:
                    bar = shell.last
                loc = bar.get_by_text(label, exact=True).first
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略B nav/footer 内「{label}」 | {fu!r}")
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略B：{_exc_tail(e)}")

        # 策略 C：精确文案，按 nth 全尝试（last/first/其余）
        try:
            n = await fr.get_by_text(label, exact=True).count()
        except Exception as e:
            lg(f"frame[{fi}] 策略C 统计「{label}」失败：{_exc_tail(e)}")
            continue
        if n < 1:
            lg(f"frame[{fi}] 策略C 无 exact「{label}」，跳过本 frame | {fu!r}")
            continue
        order: list[int] = []
        for i in (n - 1, 0):
            if i >= 0 and i not in order:
                order.append(i)
        for i in range(n):
            if i not in order:
                order.append(i)
        lg(f"frame[{fi}] 策略C exact「{label}」×{n}，尝试 nth 顺序={order[:6]}{'…' if len(order) > 6 else ''}")
        for idx in order:
            try:
                loc = fr.get_by_text(label, exact=True).nth(idx)
                await loc.wait_for(state="attached", timeout=4000)
                try:
                    info = await loc.evaluate(_JS_CLICK_TAB_FROM_LABEL)
                    lg(f"frame[{fi}] 策略C-JS nth({idx})「{label}」→ {info}")
                    await page.wait_for_timeout(350)
                    return True
                except Exception as e_js:
                    lg(f"frame[{fi}] 策略C-JS nth({idx}) 失败：{_exc_tail(e_js, 100)}")
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略C nth({idx})「{label}」成功")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略C nth({idx})：{_exc_tail(e)}")

        # 策略 D：link / button 无障碍名
        for role in ("link", "button"):
            try:
                loc_all = fr.get_by_role(role, name=label_re)
                rn = await loc_all.count()
                if rn == 0:
                    continue
                loc = loc_all.last
                await _click_attached_force(loc, force=force)
                lg(f"frame[{fi}] 策略D role={role} 匹配×{rn} 取 last")
                return True
            except Exception as e:
                lg(f"frame[{fi}] 策略D role={role}：{_exc_tail(e)}")

        # 策略 E：可点击标签 + 文案
        try:
            row = fr.locator("a,button,[role='button']").filter(has_text=label_re)
            rn = await row.count()
            lg(f"frame[{fi}] 策略E a/button/[role=button] 匹配={rn}")
            if rn > 0:
                await _click_attached_force(row.last, force=force)
                lg(f"frame[{fi}] 策略E .last 成功")
                return True
        except Exception as e:
            lg(f"frame[{fi}] 策略E：{_exc_tail(e)}")

    lg(f"全部 frame 均未点到「{label}」")
    return False


# SPA 底栏切换：勿用 networkidle（长连接/轮询会导致几乎永远不达 idle，耗时≈超时）
_WAIT_PARTY_SWITCH_JS = """(before) => {
  const href = location.href;
  if (/party-hubs|party\\/hub/i.test(href)) return true;
  if (typeof before === 'string' && href !== before) return true;
  for (const row of document.querySelectorAll('[class*="tabbar_item"], [class*="_app_tabbar_item"]')) {
    const t = (row.innerText || '').trim();
    if (!t.includes('Party')) continue;
    const c = (row.className && row.className.toString()) || '';
    if (/\\bactive\\b/i.test(c)) return true;
    /* Party 选中态有时只在子节点；勿用 [class*="active"]（会误匹配 inactive） */
    const subs = row.getElementsByTagName('*');
    for (let i = 0; i < subs.length; i++) {
      const cn = (subs[i].className && subs[i].className.toString()) || '';
      if (/\\bactive\\b/i.test(cn)) return true;
    }
  }
  return false;
}"""


async def _wait_party_switch_settled(
    page: Any,
    url_before_click: str,
    *,
    timeout_ms: float,
    log: Callable[[str], None],
) -> tuple[bool, str]:
    """
    等待 Home→Party 在可观测层面的「切换完成」：Party 路由 URL、或 URL 已变、或底栏 Party 带 active。
    与 networkidle 解耦，避免把「网络安静」误当成「Tab 响应」。
    """
    to = max(500.0, float(timeout_ms))
    try:
        # Playwright Python：arg 必须为关键字参数，不可作第二位置参数
        await page.wait_for_function(
            _WAIT_PARTY_SWITCH_JS,
            arg=url_before_click,
            timeout=to,
        )
        log(f"  [诊断·耗时] Party 切换就绪（URL 或底栏 active，非 networkidle），上限等待 {to:.0f} ms")
        return True, "settled"
    except Exception as e:
        log(f"  [诊断·耗时] {to:.0f} ms 内未观察到 Party 路由/active：{_exc_tail(e, 160)}")
        return False, "timeout"


async def _run_ext_response_time(
    page: Any,
    threshold_ms: float,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    """先回站点根路径（含离开 /party-hubs）→ 露底栏 → Home → Party，测可观测切换耗时（非 networkidle）。"""
    log("  [诊断·耗时] 确保离开 party-hubs /my 等非大厅路由…")
    await _ensure_on_home_feed(page, target_url, log)
    await page.wait_for_timeout(400)
    await _log_bottom_nav_context(page, log, tag="响应·起测前")
    await _scroll_tabbar_into_view(page)
    await _log_bottom_nav_context(page, log, tag="响应·滚底后")
    ok_home = await _try_click_bottom_exact_in_frames(page, "Home", log=log, force=True)
    if not ok_home:
        log("  [诊断·耗时] 未能点击 Home（可能已在大厅）；继续尝试 Party…")
    await page.wait_for_timeout(500)
    await _scroll_tabbar_into_view(page)

    url_before = page.url
    t0 = time.monotonic()
    clicked = await _try_click_bottom_exact_in_frames(page, "Party", log=log, force=True)
    if not clicked:
        return ("SKIP", "未能点击底栏「Party」，跳过耗时统计（请先在大厅页或检查底栏是否在 iframe）。")
    ok_wait, _why = await _wait_party_switch_settled(
        page, url_before, timeout_ms=threshold_ms, log=log
    )
    dt_ms = (time.monotonic() - t0) * 1000
    if not ok_wait:
        return (
            "FAIL",
            f"{threshold_ms:.0f} ms 内未完成 Party 切换（未命中 party-hubs 等 URL、URL 未变且底栏 Party 无 active）。",
        )
    if dt_ms <= threshold_ms:
        return (
            "PASS",
            f"点击 Party 后至路由/active 就绪约 {dt_ms:.0f} ms（阈值 {threshold_ms:.0f} ms，非 networkidle）。",
        )
    return (
        "FAIL",
        f"切换耗时 {dt_ms:.0f} ms 超过阈值 {threshold_ms:.0f} ms（弱网或主线程卡顿可调高 --switch-ms）",
    )


async def _run_ext_no_more_data(
    page: Any, target_url: str, log: Callable[[str], None]
) -> tuple[str, str]:
    """「No More Data」仅出现在 Home 大厅列表底部：先 goto/离开 party-hubs 等再滚底。"""
    log("  [诊断·NoMoreData] 回到 Home 大厅根路径（避免当前在 Party Hubs 等子页）…")
    await _ensure_on_home_feed(page, target_url, log)
    await page.wait_for_timeout(450)
    try:
        await _eval_timeout(page, "() => window.scrollTo(0, 0)", timeout=10.0)
        await page.wait_for_timeout(200)
    except (Exception, asyncio.TimeoutError):
        pass
    await _scroll_to_bottom(page)
    ok = await _ext_visible_text_pat(page, r"No\s*More\s*Data", timeout_ms=4000)
    if ok:
        return ("PASS", "列表底部可见「No More Data」类文案。")
    ok2 = await _ext_visible_text_pat(page, r"没有更多|沒有更多|no\s+more", timeout_ms=1500)
    if ok2:
        return ("PASS", "列表底部可见「没有更多」类中文/变体文案。")
    return (
        "FAIL",
        "滚底后未见「No More Data」或常见中文无更多提示（若列表未分页到底则属环境差异）。",
    )


async def _run_ext_copy_light(page: Any) -> tuple[str, str]:
    try:
        txt = await asyncio.wait_for(
            page.evaluate(
                "() => (document.body && document.body.innerText) ? document.body.innerText.slice(0, 80000) : ''"
            ),
            timeout=30.0,
        )
    except asyncio.TimeoutError:
        return ("FAIL", "读取 innerText 超时（30s），页面主线程可能过重或卡死。")
    except Exception as e:
        return ("SKIP", f"无法读取正文：{e}")
    if not txt:
        return ("SKIP", "正文为空。")
    if "\ufffd" in txt or "\uFFFD" in txt:
        return ("FAIL", "正文出现 Unicode 替换字符 U+FFFD（可能编码/乱码）。")
    return ("PASS", "正文前 80k 字符未见 U+FFFD 替换符（非完整错别字审计）。")


async def _run_ext_layout_light(page: Any) -> tuple[str, str]:
    try:
        r = await _eval_timeout(
            page,
            """() => {
              const de = document.documentElement;
              const b = document.body;
              const sw = Math.max(de.scrollWidth, b ? b.scrollWidth : 0);
              const cw = de.clientWidth;
              return { scrollWidth: sw, clientWidth: cw, ratio: cw ? sw / cw : 1 };
            }""",
            timeout=12.0,
        )
        ratio = float(r.get("ratio") or 1)
        if ratio > 1.35:
            return (
                "FAIL",
                f"主文档横向 scrollWidth/clientWidth 比 ≈ {ratio:.2f}，可能存在明显横向溢出（轻量启发式）。",
            )
        return ("PASS", f"横向比例 ≈ {ratio:.2f}（轻量，非视觉回归）。")
    except Exception as e:
        return ("SKIP", f"无法测量布局：{e}")


async def _run_ext_scroll_light(page: Any) -> tuple[str, str]:
    try:
        h0 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        for _ in range(4):
            await _eval_timeout(
                page,
                "() => { window.scrollBy(0, 1400); }",
                timeout=8.0,
            )
            await page.wait_for_timeout(350)
        h1 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        await _scroll_to_bottom(page)
        h2 = await _eval_timeout(
            page,
            "Math.max(document.body?.scrollHeight||0, document.documentElement.scrollHeight||0)",
            timeout=12.0,
        )
        if h2 < (h0 or 0) * 0.5:
            return ("FAIL", f"滚底后 scrollHeight 异常收缩（{h0} → {h2}），可能存在布局闪动。")
        return (
            "PASS",
            f"滚动后高度 {h0} → {h1} → {h2}（轻量：未检测剧烈收缩）。",
        )
    except asyncio.TimeoutError:
        return ("FAIL", "滚动/scrollHeight 检测超时，页面脚本可能无响应（已避免无限挂起）。")
    except Exception as e:
        return ("SKIP", f"滚动测试异常：{e}")


async def _run_ext_case(
    case_id: str,
    page: Any,
    *,
    console_bucket: list[str],
    switch_ms: float,
    log: Callable[[str], None],
    target_url: str,
) -> tuple[str, str]:
    if case_id == "ext_game_list":
        return await _run_ext_game_list(page)
    if case_id == "ext_images":
        return await _run_ext_images(page)
    if case_id == "ext_static_console":
        return await _run_ext_static_console(console_bucket)
    if case_id == "ext_response_time":
        return await _run_ext_response_time(page, switch_ms, log, target_url)
    if case_id == "ext_no_more_data":
        return await _run_ext_no_more_data(page, target_url, log)
    if case_id == "ext_copy_light":
        return await _run_ext_copy_light(page)
    if case_id == "ext_layout_light":
        return await _run_ext_layout_light(page)
    if case_id == "ext_scroll_light":
        return await _run_ext_scroll_light(page)
    return ("BLOCKED", f"未知用例：{case_id}")


# Slow 3G 近似（与需求一致）
_THROTTLE = {
    "offline": False,
    "latency": 400,
    "downloadThroughput": 400 * 1024 // 8,
    "uploadThroughput": 200 * 1024 // 8,
}

_NORMAL_NET = {
    "offline": False,
    "latency": 0,
    "downloadThroughput": -1,
    "uploadThroughput": -1,
}

_JS_FRAME_WEAKNET_PROBE = """() => {
  const sels = [
    '[class*="skeleton" i]', '[class*="Skeleton"]', '[class*="loading" i]', '[class*="spinner" i]',
    '[aria-busy="true"]', '[class*="shimmer" i]', '[class*="placeholder" i]',
    '.ant-skeleton', '.el-skeleton', '[class*="n-skeleton" i]'
  ];
  let hit = null;
  for (const s of sels) {
    try { const n = document.querySelector(s); if (n) { hit = s; break; } } catch (e) {}
  }
  const t = (document.body && document.body.innerText) ? document.body.innerText : '';
  const len = t.replace(/\\s+/g, ' ').trim().length;
  const navSels = [
    'header', 'nav', '[role="navigation"]', '[class*="header" i]', '[class*="Header"]',
    '[class*="nav-bar" i]', '[class*="NavBar" i]', '[class*="topbar" i]', '[class*="TopBar" i]',
    '[class*="app-bar" i]', '[class*="AppBar" i]'
  ];
  let hasTopNav = false;
  for (const s of navSels) {
    try {
      const el = document.querySelector(s);
      if (!el) continue;
      const r = el.getBoundingClientRect();
      if (r.width < 2 || r.height < 2) continue;
      if (r.top < 220) { hasTopNav = true; break; }
    } catch (e) {}
  }
  const de = document.documentElement;
  const docElLen = (de && de.innerText) ? de.innerText.replace(/\\s+/g, ' ').trim().length : 0;
  const bc = document.body ? document.body.children.length : 0;
  let rootHtmlLen = 0;
  for (const sel of ['#root', '#app', '[data-app]', '[data-v-app]', '#__next']) {
    try {
      const n = document.querySelector(sel);
      if (n && n.innerHTML) rootHtmlLen = Math.max(rootHtmlLen, n.innerHTML.length);
    } catch (e) {}
  }
  const nIframes = document.querySelectorAll('iframe').length;
  return {
    hasSkeleton: !!hit,
    selectorHit: hit,
    bodyTextLen: len,
    docElTextLen: docElLen,
    bodyChildCount: bc,
    rootHtmlLen,
    iframeCount: nIframes,
    ready: document.readyState,
    hasTopNav
  };
}"""

_JS_PHASE_C = """() => {
  const imgs = Array.from(document.querySelectorAll('img'));
  let pending = 0;
  for (const im of imgs) {
    try {
      if (!im.complete || (im.naturalWidth === 0 && (im.src || im.currentSrc))) pending++;
    } catch (e) { pending++; }
  }
  const t = (document.title || '').trim();
  const body = (document.body && document.body.innerText) ? document.body.innerText : '';
  return {
    titleLen: t.length,
    bodyLen: body.replace(/\\s+/g, ' ').trim().length,
    totalImages: imgs.length,
    pendingImages: pending,
  };
}"""

_RE_WEAK_NET_UX = re.compile(
    r"重试|网络|不给力|加载失败|连接.*超时|请稍后|try\s*again|retry|offline|unstable",
    re.I,
)

# 阶梯观测用阈值（不作唯一判死依据）
_T_CORE_LADDER = 100  # 3s 性能优秀：足量正文字
_T_MAIN_SHALLOW = 35  # 主 frame 过短、子 frame 有内容 时作 iframe 标注
_T_ROOT_HTML_SHELL = 80  # SPA 根壳 innerHTML 长度视为「结构已就绪」
_T_TEXT_MIN_SIGNAL = 8  # 聚合正文字数弱信号下限
_WEAK_POLL_MAX_SEC = 75.0  # 首内容信号最长等待


async def _probe_all_frames_weaknet(page: Any) -> dict[str, Any]:
    """
    汇总 page.frames 上同一套弱网探测。单 frame 失败不拖死全量，缺省 false/0。
    """
    out_frames: list[dict[str, Any]] = []
    try:
        flist = list(page.frames)
    except Exception:
        flist = []
    for i, fr in enumerate(flist):
        row: dict[str, Any] = {
            "frameIndex": i,
            "isMain": i == 0,
            "url": "",
            "hasSkeleton": False,
            "selectorHit": None,
            "bodyTextLen": 0,
            "docElTextLen": 0,
            "bodyChildCount": 0,
            "rootHtmlLen": 0,
            "iframeCount": 0,
            "ready": "",
            "hasTopNav": False,
        }
        try:
            row["url"] = (fr.url or "")[:200]
        except Exception:
            pass
        try:
            d = await asyncio.wait_for(
                fr.evaluate(_JS_FRAME_WEAKNET_PROBE), timeout=3.0
            )
            if isinstance(d, dict):
                row["hasSkeleton"] = bool(d.get("hasSkeleton"))
                row["selectorHit"] = d.get("selectorHit")
                row["bodyTextLen"] = int(d.get("bodyTextLen") or 0)
                row["docElTextLen"] = int(d.get("docElTextLen") or 0)
                row["bodyChildCount"] = int(d.get("bodyChildCount") or 0)
                row["rootHtmlLen"] = int(d.get("rootHtmlLen") or 0)
                row["iframeCount"] = int(d.get("iframeCount") or 0)
                row["ready"] = (d.get("ready") or "") or ""
                row["hasTopNav"] = bool(d.get("hasTopNav"))
        except (Exception, asyncio.TimeoutError):
            pass
        out_frames.append(row)
    if not out_frames:
        return {
            "per_frame": [],
            "max_body": 0,
            "max_text_signal": 0,
            "max_root_html": 0,
            "has_struct_shell": False,
            "has_skeleton_any": False,
            "has_top_nav_any": False,
            "main_body": 0,
            "sub_max_body": 0,
            "iframe_suspected_main_shell": False,
        }

    def _text_sig(x: dict[str, Any]) -> int:
        return max(
            int(x.get("bodyTextLen") or 0),
            int(x.get("docElTextLen") or 0),
        )

    main_body = int(out_frames[0].get("bodyTextLen") or 0)
    sub_bodies = [int(x.get("bodyTextLen") or 0) for x in out_frames[1:]]
    sub_max = max(sub_bodies) if sub_bodies else 0
    has_sk = any(x.get("hasSkeleton") for x in out_frames)
    has_nav = any(x.get("hasTopNav") for x in out_frames)
    max_b = max(int(x.get("bodyTextLen") or 0) for x in out_frames)
    max_text_signal = max(_text_sig(x) for x in out_frames)
    max_root = max(int(x.get("rootHtmlLen") or 0) for x in out_frames)
    has_struct = any(int(x.get("rootHtmlLen") or 0) >= _T_ROOT_HTML_SHELL for x in out_frames)
    iframe_shell = (main_body < _T_MAIN_SHALLOW) and (
        sub_max >= 40
        or any(x.get("hasSkeleton") for x in out_frames[1:])
    )
    return {
        "per_frame": out_frames,
        "max_body": max_b,
        "max_text_signal": max_text_signal,
        "max_root_html": max_root,
        "has_struct_shell": has_struct,
        "has_skeleton_any": has_sk,
        "has_top_nav_any": has_nav,
        "main_body": main_body,
        "sub_max_body": sub_max,
        "iframe_suspected_main_shell": iframe_shell,
    }


def _zh_rate_weak_net_sec(sec: float | None) -> str:
    if sec is None:
        return "未测得首可感知内容时刻"
    if sec < 3.5:
        return "优秀（首包/首屏可感知较快）"
    if sec < 8.0:
        return "良好"
    if sec < 16.0:
        return "可接受"
    if sec < 35.0:
        return "偏慢（弱网环境下仍较常见）"
    return "过慢（建议关注首包体积与串行资源）"


def _weak_content_signal(
    snap: dict[str, Any], title: str
) -> tuple[bool, str]:
    """是否出现「可感知的打开进度」；SPA/iframe/仅标题亦算，避免与 FAIL 强绑定。"""
    if not isinstance(snap, dict):
        return False, ""
    if snap.get("has_skeleton_any"):
        return True, "skeleton"
    if int(snap.get("max_text_signal") or 0) >= _T_TEXT_MIN_SIGNAL:
        return True, "text_or_docel"
    if snap.get("has_struct_shell"):
        return True, "spa_root_shell"
    if (title or "").strip() and len((title or "").strip()) >= 2:
        return True, "document_title"
    for x in snap.get("per_frame") or []:
        if not isinstance(x, dict):
            continue
        if int(x.get("bodyChildCount") or 0) >= 2 and int(x.get("rootHtmlLen") or 0) > 20:
            return True, "dom_children_with_shell"
        if int(x.get("iframeCount") or 0) >= 1 and int(x.get("rootHtmlLen") or 0) > 30:
            return True, "iframe_with_shell"
    return False, ""


def _is_timed_out_console_line(text: str) -> bool:
    t = (text or "")
    if "ERR_CONNECTION" in t or "TIMED_OUT" in t or "net::" in t:
        return True
    return "timed out" in t.lower()


async def handle_weak_network_test(
    page: Any,
    target_url: str,
    *,
    log: Callable[[str], None],
) -> tuple[str, str, dict[str, Any]]:
    """
    CDP 弱网：记录限速下导航/首可感知时间并评价；**默认 PASS**（仅真正导航失败为 FAIL）。聚合 page.frames。
    在 finally 中恢复网络、关闭「禁用缓存」并 clearBrowserCache。

    **统合冒烟常见假阳性**：此前若页面已在 ``target_url``，同址 ``goto`` 常命中内存/BFCache，
    几乎无网络字节，Slow3G 限速不生效。此处先 ``setCacheDisabled`` + ``clearBrowserCache``，
    再 ``about:blank`` 后重新 ``goto(target_url)``，使测量反映真实弱网。

    返回 (verdict, detail, extra_observations)
    """
    observations: dict[str, Any] = {
        "phase_a": None,  # 阶梯 3/5/6/10s + 各 frame 快照
        "phase_b": None,
        "phase_c": None,
        "media_degraded": False,
        "console_timed_out_hits": [],
        "performance_suboptimal_3s": False,
        "basic_5s_met": None,
    }
    cdp: Any = None
    console_bucket: list[str] = []

    def on_console(msg: Any) -> None:
        try:
            text = f"{getattr(msg, 'type', '')}: {getattr(msg, 'text', '')[:800]}"
        except Exception:
            text = str(msg)[:400]
        console_bucket.append(text)
        if _is_timed_out_console_line(text):
            observations["console_timed_out_hits"].append(text[:400])
            if len(observations["console_timed_out_hits"]) > 20:
                observations["console_timed_out_hits"].pop(0)

    async def _restore_and_cache() -> None:
        if not cdp:
            return
        try:
            await cdp.send("Network.emulateNetworkConditions", _NORMAL_NET)
        except Exception as re:
            log(f"  [弱网] 恢复无节流失败（可忽略）：{_brief_exc(re)}")
        try:
            await cdp.send("Network.setCacheDisabled", {"cacheDisabled": False})
        except Exception:
            pass
        try:
            await cdp.send("Network.clearBrowserCache", {})
        except Exception as ce:
            log(f"  [弱网] clearBrowserCache 失败（可忽略）：{_brief_exc(ce)}")

    out: tuple[str, str, dict[str, Any]] = (
        "FAIL",
        "未执行到结算",
        observations,
    )
    try:
        try:
            page.on("console", on_console)
        except Exception:
            pass
        cdp = await page.context.new_cdp_session(page)
        await cdp.send("Network.enable", {})
        # 统合冒烟跑到此条时页面通常已在 target_url：同 URL goto 常走内存/BFCache，
        # 几乎无网络字节，Slow3G 限速形同虚设（0.x s「加载」即此类假阳性）。
        try:
            await cdp.send("Network.setCacheDisabled", {"cacheDisabled": True})
        except Exception as e:
            log(f"  [弱网] Network.setCacheDisabled 失败（继续）：{_brief_exc(e)}")
        try:
            await cdp.send("Network.clearBrowserCache", {})
        except Exception as e:
            log(f"  [弱网] 导航前 clearBrowserCache：{_brief_exc(e)}")
        log("  [弱网] Network.emulateNetworkConditions（Slow3G 类）…")
        await cdp.send("Network.emulateNetworkConditions", _THROTTLE)
        log("  [弱网] 冷导航：about:blank → 目标页（强制走网络，使限速生效）…")
        try:
            await page.goto("about:blank", wait_until="commit", timeout=15_000)
        except Exception as e:
            log(f"  [弱网] about:blank 提示：{_brief_exc(e)}")
        observations["weak_net_cold_navigation"] = True

        # —— 限速下导航 + 首包耗时 + 仅观测用阶梯（不 6s/10s 判死）——
        t0 = time.monotonic()
        try:
            await page.goto(
                target_url, wait_until="domcontentloaded", timeout=90_000
            )
        except Exception as e:
            gerr = _brief_exc(e)
            observations["phase_a"] = {"goto_error": gerr}
            return "FAIL", f"弱网下无法完成导航/打开首屏：{gerr}", observations

        nav_sec = time.monotonic() - t0
        observations["weak_net_nav_until_domcontentloaded_sec"] = round(nav_sec, 2)
        nav_load_rating = _zh_rate_weak_net_sec(nav_sec)
        observations["weak_net_nav_duration_rating_zh"] = nav_load_rating
        log(
            f"  [弱网] 加载耗时：自开始导航至 **domcontentloaded** 为 **{nav_sec:.1f} s**"
            f"（弱网 Slow3G 限速）；**对该加载耗时评价**：{nav_load_rating}"
        )

        excellent_3s = False
        m3 = m5 = False
        last_snap: dict[str, Any] = {}
        first_sig_sec: float | None = None
        signal_kind = ""
        t_poll_end = t0 + _WEAK_POLL_MAX_SEC
        title_for_sig = ""

        while time.monotonic() < t_poll_end:
            el = time.monotonic() - t0
            try:
                last_snap = await _probe_all_frames_weaknet(page)
            except Exception as e:
                last_snap = {"per_frame": [], "error": _brief_exc(e)}

            try:
                title_for_sig = (await page.title() or "").strip()
            except Exception:
                title_for_sig = ""

            sk = bool(last_snap.get("has_skeleton_any"))
            mtxt = int(
                last_snap.get("max_text_signal")
                or last_snap.get("max_body")
                or 0
            )
            nav = bool(last_snap.get("has_top_nav_any"))
            shell = bool(last_snap.get("iframe_suspected_main_shell"))
            n_frames = len(last_snap.get("per_frame") or [])

            ok_s, knd = _weak_content_signal(last_snap, title_for_sig)
            if ok_s and first_sig_sec is None:
                first_sig_sec = el
                signal_kind = knd

            if el <= 3.0 and (sk or mtxt >= _T_CORE_LADDER):
                excellent_3s = True

            if not m3 and el >= 3.0:
                m3 = True
                fast_enough_3s = (first_sig_sec is not None and first_sig_sec <= 3.0) or (
                    sk or mtxt >= _T_CORE_LADDER
                )
                if not fast_enough_3s:
                    observations["performance_suboptimal_3s"] = True
                    log(
                        "  [弱网] Observation: Performance sub-optimal"
                        "（3s 参考线未先出现可感知内容；SPA/弱网可晚于 3s 出现）"
                    )
                if shell and n_frames > 1:
                    observations["iframe_content_delayed_note"] = (
                        "主 document 正文字数偏少，子 frame 含更多文案/骨架；以首包与人工目视为准。"
                    )
                    log(
                        f"  [弱网] Observation: 多 frame（n={n_frames}）主壳偏空、"
                        f"子区 max_text≈{last_snap.get('sub_max_body')}，已单独标注。"
                    )

            if not m5 and el >= 5.0:
                m5 = True
                basic_ok = sk and nav
                observations["basic_5s_met"] = basic_ok
                if not basic_ok:
                    log(
                        "  [弱网] 5s 参考档：骨架+顶栏未同时出现"
                        f"（skeleton={sk} topNav={nav}；仅作记录）"
                    )

            if el >= 10.2 and m3 and m5 and first_sig_sec is not None:
                break
            await asyncio.sleep(0.1)

        if first_sig_sec is None:
            ok, kind = _weak_content_signal(last_snap, title_for_sig)
            if ok:
                first_sig_sec = min(
                    time.monotonic() - t0, _WEAK_POLL_MAX_SEC
                )
                signal_kind = kind
            else:
                first_sig_sec = nav_sec
                signal_kind = "nav_only_soft"
                umsg = (
                    "未稳定探测到内联正文/根壳（可能为跨域 iframe、Canvas、Shadow 等）；"
                    f"但 domcontentloaded 已于 {nav_sec:.1f}s 达成。本项不判 FAIL，建议实机目视。"
                )
                observations["weak_net_signal_uncertain"] = umsg
                log(f"  [弱网] Observation: {umsg}")
                log(
                    f"  [弱网] 以导航完成时刻作为参考首包：{first_sig_sec:.1f}s（{signal_kind}）"
                )

        rating = _zh_rate_weak_net_sec(first_sig_sec)
        observations["weak_net_sec_to_first_signal"] = (
            round(first_sig_sec, 2) if first_sig_sec is not None else None
        )
        observations["weak_net_signal_kind"] = signal_kind
        observations["weak_net_experience_rating_zh"] = rating
        log(
            f"  [弱网] 首可感知内容出现：约 **{first_sig_sec:.1f} s**"
            f"（信号：{signal_kind}）；**对该首现耗时评价**：{rating}"
        )
        if (
            first_sig_sec is not None
            and abs(float(first_sig_sec) - float(nav_sec)) < 0.2
        ):
            log(
                "  [弱网] 说明：首可感知与 domcontentloaded 几乎同时，"
                "上列「加载耗时」与「首现耗时」数值接近、评价可一并参考。"
            )
        else:
            log(
                f"  [弱网] 汇总：主文档 domcontentloaded **{nav_sec:.1f} s** → {nav_load_rating}；"
                f" 首可感知 **{first_sig_sec:.1f} s** → {rating}。"
            )

        observations["phase_a"] = {
            "ok": True,
            "ladder": {
                "tier_3s_excellent_core_or_skeleton": excellent_3s,
                "tier_3s_observation_performance_sub_optimal": bool(
                    observations.get("performance_suboptimal_3s")
                ),
                "tier_5s_met": observations.get("basic_5s_met"),
                "last_snap_summary": {
                    "max_body": last_snap.get("max_body"),
                    "max_text_signal": last_snap.get("max_text_signal"),
                    "max_root_html": last_snap.get("max_root_html"),
                    "has_skeleton_any": last_snap.get("has_skeleton_any"),
                    "has_top_nav_any": last_snap.get("has_top_nav_any"),
                    "frames": len(last_snap.get("per_frame") or []),
                },
            },
            "per_frame_tail": (last_snap.get("per_frame") or [])[:8],
        }

        # —— 阶段 B：load 为加分项，不因弱网长载 FAIL ——
        phase_b_ok = True
        phase_b_note = "弱网下继续等待 load 事件"
        try:
            await page.wait_for_load_state("load", timeout=45_000)
            phase_b_note = "已触发 load 事件"
        except (Exception, asyncio.TimeoutError):
            phase_b_ok = False
            try:
                html_snip = await asyncio.wait_for(
                    page.evaluate(
                        "() => (document.body && document.body.innerText) ? "
                        "document.body.innerText.slice(0, 1200) : ''"
                    ),
                    timeout=5.0,
                )
            except Exception:
                html_snip = ""
            if html_snip and _RE_WEAK_NET_UX.search(html_snip):
                phase_b_note = "load 未在 45s 内到达，但可见网络/重试类引导"
            else:
                phase_b_note = (
                    "load 在 45s 内未到达（资源仍在拉取常见）；不单独判 FAIL。摘要："
                    f"{html_snip[:200]!r}"
                )
        observations["phase_b"] = {"ok": phase_b_ok, "note": phase_b_note}

        # —— 首图：15s 内无解码完成则记 Media degraded（仍可为 PASS）——
        media_degraded = False
        try:
            await asyncio.wait_for(
                page.wait_for_function(
                    """() => {
                      const imgs = document.querySelectorAll('img');
                      if (!imgs.length) return true;
                      for (const im of Array.from(imgs)) {
                        try {
                          if (im.src && im.complete && im.naturalWidth > 0) return true;
                        } catch (e) {}
                      }
                      return false;
                    }""",
                    timeout=15_000,
                ),
                timeout=16.0,
            )
        except (Exception, asyncio.TimeoutError):
            media_degraded = True
            observations["media_degraded"] = True
            log(
                "  [弱网] Observation: Media resource degraded"
                "（图 15s 内未出现完整解码，记录为降级）"
            )

        c_stats = await page.evaluate(_JS_PHASE_C)
        snap_c = await _probe_all_frames_weaknet(page)
        bl_main = int(c_stats.get("bodyLen") or 0)
        bl_agg = max(bl_main, int(snap_c.get("max_body") or 0))
        c_stats["bodyLen_main"] = bl_main
        c_stats["bodyLen"] = bl_agg
        c_stats["bodyLen_aggregate_max"] = bl_agg
        observations["phase_c"] = c_stats
        n_title = int(c_stats.get("titleLen") or 0)
        n_body = bl_agg
        pend = int(c_stats.get("pendingImages") or 0)

        c_notes: list[str] = []
        if n_title < 2 and n_body < 20:
            c_notes.append(
                f"内联正文字/标题仍偏少（titleLen={n_title} bodyAgg={n_body}）"
                "，可能为 iframe/只读区限制；不单独转 FAIL"
            )
        if pend > 0 and n_body < 30:
            c_notes.append("媒体仍多 pending，与弱网/懒加载相关；不单独转 FAIL")
        for cn in c_notes:
            log(f"  [弱网] Observation: {cn}")

        w_nav = observations.get("weak_net_nav_until_domcontentloaded_sec")
        w_nav_r = observations.get("weak_net_nav_duration_rating_zh", "")
        w_ftt = observations.get("weak_net_sec_to_first_signal")
        w_rate = observations.get("weak_net_experience_rating_zh", "")
        w_kind = observations.get("weak_net_signal_kind", "")
        detail = (
            f"弱网：至 domcontentloaded 加载 {w_nav!s} s，对该加载时间评价：{w_nav_r!s}；"
            f"首可感知 {w_ftt!s} s（{w_kind!s}）对该时间评价：{w_rate!s}；"
            f"title/body(聚合) {n_title}/{n_body}，图 pending {pend}/"
            f"{c_stats.get('totalImages', '?')}"
        )
        if media_degraded:
            detail += " Observation: Media resource degraded（仍计 PASS）。"
        n_hits = len(observations.get("console_timed_out_hits") or [])
        if n_hits:
            detail += f" Console 曾出现连接/超时类行 {n_hits} 条（已记录）。"
        out = "PASS", detail, observations
        return out
    except Exception as e:
        out = "FAIL", f"弱网用例异常：{_brief_exc(e)}", observations
        return out
    finally:
        try:
            page.remove_listener("console", on_console)
        except Exception:
            pass
        await _restore_and_cache()
        if cdp:
            try:
                await cdp.detach()
            except Exception:
                pass


async def _async_main(args: argparse.Namespace) -> int:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("请先安装：pip install playwright && playwright install chromium", file=sys.stderr)
        return 2

    cdp = _kalaroko_cdp(args.cdp_http or None)
    target_url = (args.target_url or DEFAULT_TARGET).strip()
    host = _host_from_url(target_url)

    def log(msg: str) -> None:
        if args.verbose or not args.quiet:
            print(msg, flush=True)

    log("———————— K11 平台冒烟 · 统合版（P0+P1+扩展+弱网）————————")
    log(f"CDP：{cdp}  目标：{target_url}  Party 切换阈值：{args.switch_ms} ms")
    _skip_c = bool(getattr(args, "skip_browser_compat", False))
    _skip_g = bool(getattr(args, "skip_game_open_smoke", False))
    log(
        f"用例主流程 {len(UNIFIED_CASE_DEFS)} 条；"
        f"结束后{'不' if _skip_c else '将'}子进程跑浏览器兼容并并入结果；"
        f"随后{'不' if _skip_g else '将'}在同一 CDP 页签跑 herontest 游戏开门探活并并入结果"
    )
    log("")

    console_bucket: list[str] = []

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(cdp)
        navigate_if_no_tab = not getattr(args, "require_existing_tab", False)
        page, pick_err = await _acquire_cdp_target_page(
            browser,
            host=host,
            target_url=target_url,
            navigate_if_no_tab=navigate_if_no_tab,
            log=log,
        )
        if page is None:
            print(f"[失败] {pick_err or '无法获取目标页签'}", file=sys.stderr)
            return 2

        def _on_console(msg: Any) -> None:
            try:
                if msg.type == "error":
                    console_bucket.append(f"{msg.type}: {msg.text[:600]}")
            except Exception:
                pass

        page.on("console", _on_console)

        ok_env, env_detail = await _ensure_target_page(
            page,
            target_url,
            log=log,
            navigate_if_no_tab=navigate_if_no_tab,
            host=host,
        )
        if not ok_env:
            print(f"[失败] {env_detail}", file=sys.stderr)
            return 2

        log("准备：若在 party-hubs、app_tabbar=no 或个人中心，先回站点首页。")
        await _ensure_on_home_feed(page, target_url, log)
        log("")

        results: list[dict[str, Any]] = []
        for i, (cid, tier, title_zh) in enumerate(UNIFIED_CASE_DEFS, start=1):
            log(f"【{i}/{len(UNIFIED_CASE_DEFS)}】[{tier}] {title_zh}（{cid}）")
            await _dismiss_kalaroko_notification_prompt(page, log=log)
            v, detail = "BLOCKED", f"未执行：{cid}"
            try:
                if cid == "p0_env_access":
                    v, detail = "PASS", env_detail
                elif cid.startswith("p0_"):
                    v, detail = await _run_p0_case(
                        cid,
                        page,
                        log=log,
                        target_url=target_url,
                        console_bucket=console_bucket,
                        p0_play_now_really_click=bool(
                            getattr(args, "p0_play_now_really_click", False)
                        ),
                    )
                elif cid.startswith("p1_"):
                    await _ensure_on_home_feed(page, target_url, log)
                    v, detail = await _run_p1_case(
                        cid, page, log=log, target_url=target_url
                    )
                elif cid.startswith("ext_"):
                    await _ensure_on_home_feed(page, target_url, log)
                    v, detail = await _run_ext_case(
                        cid,
                        page,
                        console_bucket=console_bucket,
                        switch_ms=float(args.switch_ms),
                        log=log,
                        target_url=target_url,
                    )
                elif cid == "p2_weak_network":
                    v, detail, _obs = await handle_weak_network_test(
                        page, target_url, log=log
                    )
                else:
                    v, detail = "BLOCKED", f"未知用例：{cid}"
            except (KeyboardInterrupt, asyncio.CancelledError):
                raise
            except Exception as e:
                v, detail = "FAIL", f"用例执行异常：{_brief_exc(e, 480)}"
                if log:
                    log(f"  [错误] 本用例将记为 FAIL 并继续后续用例，避免整进程以 exit=3 退出。原因：{detail}")

            vzh = VERDICT_ZH.get(v, v)
            log(f"  观察说明：{detail}")
            log(f"  结论：{vzh}（{v}）")
            log("")
            results.append(
                {
                    "case": cid,
                    "tier": tier,
                    "case_title_zh": title_zh,
                    "verdict": v,
                    "verdict_zh": vzh,
                    "detail": detail,
                }
            )

            if cid == "p0_release_checklist":
                log("—— P1 前：回大厅 ——")
                try:
                    await _ensure_on_home_feed(page, target_url, log)
                except Exception as e:
                    if log:
                        log(f"  [警告] P1 前回大厅失败（已记录，继续跑）：{_brief_exc(e, 200)}")
            elif cid == "p1_party_status":
                log("—— 扩展项前：回大厅并轻滚 ——")
                try:
                    await _ensure_on_home_feed(page, target_url, log)
                except Exception as e:
                    if log:
                        log(f"  [警告] 扩展前回大厅失败（已记录，继续跑）：{_brief_exc(e, 200)}")
                try:
                    await _eval_timeout(
                        page, "() => window.scrollBy(0, 400)", timeout=6.0
                    )
                except (Exception, asyncio.TimeoutError):
                    pass
                await page.wait_for_timeout(400)

        bad_console = _filter_console_errors(console_bucket)
        if bad_console:
            log("———————— Console error（P0 过滤后抽样，最多 15 条）————————")
            for s in bad_console[:15]:
                log("  · " + s[:400])
            log("")

        mime_hits = _mime_console_failures(console_bucket)
        if mime_hits:
            log("———————— Console · MIME/模块脚本（供对照）————————")
            for s in mime_hits[:8]:
                log("  · " + s[:300])
            log("")

        compat_info: dict[str, Any] = {"skipped": True}
        if not getattr(args, "skip_browser_compat", False):
            log(
                "———————— 附加：浏览器兼容（子进程 test_k11_p2_compat_weaknet_playwright.py --only-compat）————————"
            )
            extra, crc, cerr = _run_p2_only_compat_subprocess(
                target_url=target_url,
                project_root=ROOT,
                log=log,
                headless=bool(getattr(args, "browser_compat_headless", False)),
                quiet_p2=bool(args.quiet),
            )
            compat_info = {
                "skipped": False,
                "subprocess_exit_code": crc,
                "merge_error": cerr,
                "appended_count": len(extra),
            }
            for row in extra:
                results.append(row)
                v = str(row.get("verdict", ""))
                vzh = str(row.get("verdict_zh", v))
                log(
                    f"  [compat] 并入用例 {row.get('case', '')!r}："
                    f"{row.get('case_title_zh', '')} → {vzh}（{v}）"
                )
            if not extra:
                log("  [compat] 未向 results 追加行（子进程无归并项）。")
            log("")
        else:
            log("  [compat] 已跳过：--skip-browser-compat")
            log("")

        game_open_info: dict[str, Any] = {"skipped": True}
        if not getattr(args, "skip_game_open_smoke", False):
            log(
                "———————— 附加：游戏模块开门探活（herontest · test_k11_game_open_smoke）————————"
            )
            gmod = _load_game_open_smoke_module(log=log)
            if gmod is None or not hasattr(gmod, "run_game_open_smoke_on_existing_page"):
                log(
                    "  [game_open] 未找到或无法加载 test_k11_game_open_smoke.py（"
                    f"路径 {_resolve_game_open_smoke_script_path()}），跳过。"
                )
                game_open_info = {"skipped": False, "error": "module_missing"}
            else:
                try:
                    raw_rows: list[dict[str, Any]] = await gmod.run_game_open_smoke_on_existing_page(  # type: ignore[misc]
                        page,
                        verbose=bool(args.verbose),
                        log=log,
                    )
                    game_open_info = {
                        "skipped": False,
                        "target": str(getattr(gmod, "TARGET_HOME", "")),
                        "games_run": len(raw_rows),
                    }
                    for row in raw_rows:
                        v = str(row.get("verdict", "FAIL")).upper()
                        if v not in ("PASS", "FAIL", "SKIP", "BLOCKED"):
                            v = "FAIL"
                        vzh = VERDICT_ZH.get(v, v)
                        gid = str(row.get("game_id") or "")
                        cid = f"game_open_{gid}" if gid else "game_open_unknown"
                        title = str(row.get("game_title") or gid or "游戏开门")
                        detail = str(row.get("detail") or "")
                        if row.get("load_ms") is not None:
                            detail = f"{detail} (load_ms={row.get('load_ms')})"
                        results.append(
                            {
                                "case": cid,
                                "tier": "游戏开门",
                                "case_title_zh": title,
                                "verdict": v,
                                "verdict_zh": vzh,
                                "detail": detail,
                            }
                        )
                        log(
                            f"  [game_open] 并入 {cid!r}：{title} → {vzh}（{v}）"
                        )
                except (KeyboardInterrupt, asyncio.CancelledError):
                    raise
                except Exception as e:
                    ber = _brief_exc(e, 480)
                    log(
                        f"  [game_open] 执行异常（已记 FAIL 占位，不阻断收尾）：{ber}"
                    )
                    game_open_info = {
                        "skipped": False,
                        "error": ber,
                        "games_run": 0,
                    }
                    results.append(
                        {
                            "case": "game_open_suite",
                            "tier": "游戏开门",
                            "case_title_zh": "游戏开门探活（整段异常）",
                            "verdict": "FAIL",
                            "verdict_zh": "失败",
                            "detail": ber,
                        }
                    )
            log("")
        else:
            log("  [game_open] 已跳过：--skip-game-open-smoke")
            log("")

        # 收尾快照需在关闭 Playwright 前完成，但不得因 CDP/页签异常让「写盘+飞书」整段不执行
        page_url_final = ""
        page_title_final = ""
        try:
            page_url_final = (page.url or "").strip()
        except Exception as e:
            log(
                f"  [警告] 无法读取 page.url（将写空字符串；不阻断飞书）：{_brief_exc(e, 200)}"
            )
        try:
            page_title_final = (await page.title() or "").strip()
        except Exception as e:
            log(
                f"  [警告] 无法读取 page.title（将写空字符串；不阻断飞书）：{_brief_exc(e, 200)}"
            )

    # 以下不依赖 CDP/浏览器：多轮 L3 调度下，每轮子进程结束时都应执行，避免只跑用例、无声退出
    out = {
        "schema": "k11_unified_platform_smoke_playwright/v2",
        "ts_utc": datetime.now(timezone.utc).isoformat(),
        "cdp": cdp,
        "target_url": target_url,
        "switch_ms_threshold": args.switch_ms,
        "page_url_final": page_url_final,
        "page_title_final": page_title_final,
        "console_errors_filtered_sample": bad_console[:30],
        "browser_compat_subprocess": compat_info,
        "game_open_smoke": game_open_info,
        "results": results,
    }
    if args.json_out:
        outp = Path(args.json_out)
        outp.parent.mkdir(parents=True, exist_ok=True)
        outp.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        log(f"JSON：{outp.resolve()}")

    if args.write_local_xlsx:
        xlsx_p = (
            args.xlsx_report
            if args.xlsx_report is not None
            else _default_k11_xlsx_report_path()
        )
        log("")
        log("—— 本地 Excel（可选）——")
        write_k11_unified_results_to_xlsx(Path(xlsx_p), results, log=log)

    if not args.no_lark_report:
        _lark_path = _resolve_k11_lark_smoke_report_path()
        k11_lark: Any = None
        if _lark_path.is_file():
            _spec = importlib.util.spec_from_file_location(
                "k11_lark_smoke_report", _lark_path
            )
            if _spec and _spec.loader:
                k11_lark = importlib.util.module_from_spec(_spec)
                try:
                    _spec.loader.exec_module(k11_lark)
                except Exception as e:
                    k11_lark = None
                    log(f"  [lark] 加载 k11_lark_smoke_report 失败，跳过：{_brief_exc(e, 320)}")
        if k11_lark:
            # 先注入侧车内嵌 Lark/K11 键，再读 os.environ（frozen 下内嵌覆盖陈旧环境，避免走错机器人）
            try:
                from l3_node.packaged_lark_env import apply_packaged_lark_to_os_environ

                apply_packaged_lark_to_os_environ()
            except Exception:
                pass
            _wiki = (
                (args.lark_wiki_url or "").strip()
                or (os.environ.get("K11_SMOKE_LARK_WIKI_URL") or "").strip()
                or K11_DEFAULT_LARK_WIKI_URL
            )
            if not _wiki and hasattr(k11_lark, "_DEFAULT_WIKI_URL"):
                _wiki = str(k11_lark._DEFAULT_WIKI_URL)
            log("")
            log("—— 飞书：同步到 Wiki/表格 ——")
            log(f"  目标：{_wiki}")
            _aid = (os.environ.get("K11_SMOKE_LARK_APP_ID") or "").strip()
            _sec = (os.environ.get("K11_SMOKE_LARK_APP_SECRET") or "").strip()
            _tbl = (os.environ.get("K11_SMOKE_LARK_TABLE_ID") or "").strip() or None
            _chat = (os.environ.get("K11_SMOKE_LARK_NOTIFY_CHAT_ID") or "").strip()
            log("")
            try:
                _nw = k11_lark.write_k11_unified_results_to_lark_bitable(  # type: ignore[attr-defined]
                    case_to_item_key=UNIFIED_CASE_TO_XLSX_TEST_ITEM_KEY,
                    results=results,
                    wiki_url=_wiki,
                    app_id=_aid,
                    app_secret=_sec,
                    table_id=_tbl,
                    log=log,
                )
                k11_lark.send_k11_smoke_lark_notification(  # type: ignore[attr-defined]
                    results=results,
                    target_url=target_url,
                    wiki_url=_wiki,
                    lark_wrote=int(_nw or 0),
                    app_id=_aid,
                    app_secret=_sec,
                    chat_id=_chat,
                    log=log,
                )
            except Exception as e:
                log(f"  [lark] 同步表或发消息时异常（已记日志，不阻断汇总）：{_brief_exc(e, 480)}")
        else:
            log(
                f"  [lark] 未找到或无法加载 k11_lark_smoke_report（解析路径为 {_lark_path}），"
                "跳过飞书同步；打包/便携请把该 .py 与统合脚本同放入 scripts/ 或 JACHIN_APP_ROOT/scripts/。"
            )

    log("———————— 汇总 ————————")
    for r in results:
        m = "✓" if r["verdict"] == "PASS" else ("○" if r["verdict"] == "SKIP" else "✗")
        log(
            f"  {m} [{r.get('tier', '')}] {r['case_title_zh']} → {r['verdict_zh']}"
        )

    verdicts = {r["verdict"] for r in results}
    if "FAIL" in verdicts or "BLOCKED" in verdicts:
        log("\n最终结果：存在未通过或阻塞项，退出码 1。")
        return 1
    log("\n最终结果：无 FAIL/BLOCKED，退出码 0。")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="K11 统合冒烟：P0+P1+扩展+弱网（单次 CDP）")
    ap.add_argument("--target-url", default=DEFAULT_TARGET, help="站点 URL（匹配标签页 host）")
    ap.add_argument("--cdp-http", default="", help="覆盖 KALAROKO_CDP_ENDPOINT")
    ap.add_argument(
        "--require-existing-tab",
        action="store_true",
        help="必须已有含目标域的页签；默认允许在存活页签上自动 goto",
    )
    ap.add_argument(
        "--switch-ms",
        type=float,
        default=12000.0,
        help="扩展·响应时间：底栏 Party 切换就绪上限毫秒（默认 12000）",
    )
    ap.add_argument(
        "--p0-play-now-really-click",
        action="store_true",
        help="P0 Play Now：可见性通过后真实点击一次并立即 goto 回大厅（默认与单脚本一致：不点击）",
    )
    ap.add_argument("--json-out", type=Path, default=None)
    ap.add_argument(
        "--xlsx-report",
        type=Path,
        default=None,
        help="与 --write-local-xlsx 联用；K11 平台测试用例 xlsx 路径；默认 K11_XLSX_REPORT 或 ~/Downloads/…",
    )
    ap.add_argument(
        "--write-local-xlsx",
        action="store_true",
        help="写本机 K11 平台测试用例 xlsx；默认不写，仅飞书同步",
    )
    ap.add_argument(
        "--no-lark-report",
        action="store_true",
        help="不写入飞书 Wiki 多维表、不向会话发完成通知",
    )
    ap.add_argument(
        "--skip-browser-compat",
        action="store_true",
        help="不在末尾子进程跑 test_k11_p2_compat_weaknet_playwright.py --only-compat",
    )
    ap.add_argument(
        "--skip-game-open-smoke",
        action="store_true",
        help="不在浏览器兼容之后跑 test_k11_game_open_smoke（herontest 五款游戏开门探活）",
    )
    ap.add_argument(
        "--browser-compat-headless",
        action="store_true",
        help="浏览器兼容子进程内 Chrome/Edge 使用无头（传给 P2 的 --headless）",
    )
    ap.add_argument(
        "--lark-wiki-url",
        default="",
        help=(
            f"飞书 Wiki 节点（含内嵌表）；默认用环境 K11_SMOKE_LARK_WIKI_URL 或内置与 k11 脚本一致的链接（见 K11_DEFAULT_LARK_WIKI_URL；当前 {K11_DEFAULT_LARK_WIKI_URL!r}）"
        ),
    )
    ap.add_argument("--quiet", action="store_true")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()
    try:
        return asyncio.run(_async_main(args))
    except Exception as e:
        print(f"[失败] {type(e).__name__}: {e}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
