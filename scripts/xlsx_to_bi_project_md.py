#!/usr/bin/env python3
"""一次性/可复用：将 K11 干系人 xlsx 转为 docs/bi_daily_report/bi_project/*.md"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

root = Path(__file__).resolve().parents[1]
if str(root) not in sys.path:
    sys.path.insert(0, str(root))


def esc(s: object) -> str:
    if s is None:
        return ""
    t = str(s).replace("|", "\\|").replace("\n", " ")
    return t


def xlsx_to_md(src: Path, dest: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(src, data_only=True)
    lines: list[str] = [
        "# K11 需求池 — 干系人（人员表）",
        "",
        "> 来源：公司内部表 `K11 需求池_干系人.xlsx`，由脚本转为 Markdown 便于 BI/Agent 检索。",
        "",
    ]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## 工作表：{sheet_name}")
        lines.append("")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("（空表）")
            lines.append("")
            continue
        header = [esc(c) for c in rows[0]]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for r in rows[1:]:
            if r and all(v is None or str(v).strip() == "" for v in r):
                continue
            cells = []
            for i in range(len(header)):
                v = r[i] if i < len(r) else None
                cells.append(esc(v))
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")
    wb.close()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--src",
        type=Path,
        default=Path(r"d:\zzz\bi\项目文档\k11项目文档\K11 需求池_干系人.xlsx"),
    )
    p.add_argument(
        "--dest",
        type=Path,
        default=root / "docs" / "bi_daily_report" / "bi_project" / "K11_需求池_干系人.md",
    )
    args = p.parse_args()
    if not args.src.is_file():
        print(f"源文件不存在: {args.src}", file=sys.stderr)
        return 1
    xlsx_to_md(args.src, args.dest)
    print(f"已写入 {args.dest} ({args.dest.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
