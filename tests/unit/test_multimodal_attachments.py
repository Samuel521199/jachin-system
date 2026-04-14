"""多模态附件组装（纯函数，无网络）。"""
from __future__ import annotations

import base64

from l3_node.intent_gateway.multimodal_attachments import build_openai_user_content


def test_no_attachments_returns_plain_text():
    assert build_openai_user_content("hello", []) == "hello"


def test_png_base64_produces_multimodal_list():
    # 1x1 PNG minimal
    png_b64 = (
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )
    raw = base64.b64decode(png_b64)
    assert len(raw) < 100
    out = build_openai_user_content(
        "看图",
        [{"name": "x.png", "mime": "image/png", "base64": png_b64}],
    )
    assert isinstance(out, list)
    assert out[0]["type"] == "image_url"
    assert out[0]["image_url"]["url"].startswith("data:image/jpeg;base64,")
    assert out[1]["type"] == "text"
    assert "看图" in out[1]["text"]


def test_txt_attachment_prefixes_text():
    raw = "line1\nline2".encode("utf-8")
    b64 = base64.b64encode(raw).decode("ascii")
    out = build_openai_user_content("用户问题", [{"name": "a.txt", "mime": "text/plain", "base64": b64}])
    assert isinstance(out, str)
    assert "[附件: a.txt 内容]" in out
    assert "line1" in out
    assert "用户问题" in out


def test_docx_base64_without_local_path_extracts_text_not_zip_garbage():
    """仅 Base64、无 path、mime=octet-stream 时须按 .docx 解析，不能把 ZIP 当 UTF-8（PK… 乱码）。"""
    from io import BytesIO

    from docx import Document

    bio = BytesIO()
    d = Document()
    d.add_paragraph("HelloDocxUniqueMarker")
    d.save(bio)
    b64 = base64.b64encode(bio.getvalue()).decode("ascii")
    out = build_openai_user_content(
        "请分析附件",
        [{"name": "Q3 战略.docx", "mime": "application/octet-stream", "base64": b64}],
    )
    assert isinstance(out, str)
    assert "HelloDocxUniqueMarker" in out
    assert "PK\x03\x04" not in out
    assert "[Content_Types].xml" not in out


def test_xlsx_base64_without_local_path_extracts_cell_text():
    """仅 Base64、mime=octet-stream 时须按 .xlsx 用 openpyxl 解析，不能把 ZIP 当 UTF-8。"""
    from io import BytesIO

    from openpyxl import Workbook

    bio = BytesIO()
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "HelloXlsxUniqueMarker"
    ws["B2"] = 42
    wb.save(bio)
    b64 = base64.b64encode(bio.getvalue()).decode("ascii")
    out = build_openai_user_content(
        "请分析表格",
        [{"name": "报表.xlsx", "mime": "application/octet-stream", "base64": b64}],
    )
    assert isinstance(out, str)
    assert "HelloXlsxUniqueMarker" in out
    assert "42" in out
    assert "=== Sheet: Data ===" in out
    assert "PK\x03\x04" not in out


def test_openai_nested_image_url_dict_with_data_uri():
    """image_url 为 {\"url\": \"data:image/...\"} 时必须识别，勿 str(dict) 导致丢图。"""
    png_b64 = (
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )
    data_url = f"data:image/png;base64,{png_b64}"
    out = build_openai_user_content(
        "图片讲述了什么内容",
        [{"name": "x.png", "mime": "image/png", "image_url": {"url": data_url}}],
    )
    assert isinstance(out, list)
    assert out[0]["type"] == "image_url"
    assert out[0]["image_url"]["url"] == data_url
    assert out[1]["type"] == "text"
    assert "图片讲述了什么内容" in out[1]["text"]


def test_top_level_string_data_image_url():
    """顶层 image_url 字符串为 data:image/... 时直接并入多模态列表。"""
    png_b64 = (
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )
    data_url = f"data:image/png;base64,{png_b64}"
    out = build_openai_user_content(
        "看图",
        [{"name": "y.png", "mime": "image/png", "image_url": data_url}],
    )
    assert isinstance(out, list)
    assert out[0]["image_url"]["url"] == data_url
